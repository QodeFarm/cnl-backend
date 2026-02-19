from uuid import UUID
from django.http import Http404
from django.shortcuts import get_object_or_404, render
from requests import request
from rest_framework import viewsets, generics, mixins as mi
from apps import customer
from apps.auditlogs.utils import log_user_action
from apps.customer.filters import CustomerCreditLimitReportFilter, CustomerOrderHistoryReportFilter, CustomerSummaryReportFilter, LedgerAccountsFilters, CustomerFilters, CustomerAddressesFilters, CustomerAttachmentsFilters, CustomerLedgerReportFilter, CustomerOutstandingReportFilter
from apps.customfields.models import CustomField, CustomFieldValue
from apps.customfields.serializers import CustomFieldSerializer, CustomFieldValueSerializer
from apps.sales.filters import SaleOrderFilter
from apps.sales.models import SaleInvoiceOrders, SaleOrder
from config.utils_db_router import set_db
from config.utils_filter_methods import filter_response, list_filtered_objects
from .models import *
from .serializers import *
from apps.company.models import Companies
from config.utils_methods import *
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
import logging
from django.db import transaction
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.serializers import ValidationError
from django.core.exceptions import  ObjectDoesNotExist
from django.db.models import Sum, F, DecimalField, ExpressionWrapper,When,Value,Case,FloatField,Max
from django.db.models.functions import Coalesce
from .models import Customer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
import openpyxl
import os

# Set up basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Create a logger object
logger = logging.getLogger(__name__) 

# Create your views here.

class LedgerAccountsViews(viewsets.ModelViewSet):
    queryset = LedgerAccounts.objects.all().order_by('is_deleted', '-created_at')
    serializer_class = LedgerAccountsSerializers
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = LedgerAccountsFilters
    ordering_fields = ['name', 'created_at', 'updated_at']
    
    #log actions
    log_actions = True
    log_module_name = "Ledger Accounts"
    log_pk_field = "ledger_account_id"
    log_display_field = "code"

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, LedgerAccounts, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)

class CustomerViews(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by('is_deleted', '-created_at')	
    serializer_class = CustomerSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = CustomerFilters
    ordering_fields = ['name', 'created_at', 'updated_at']

    def list(self, request, *args, **kwargs):
        # return create_instance(self, request, *args, **kwargs)
        summary = request.query_params.get('summary', 'false').lower() == 'true'
        if summary:
            customers = self.filter_queryset(self.get_queryset())
            data = CustomerOptionSerializer.get_customer_summary(customers)
            
            Result = Response(data, status=status.HTTP_200_OK)
        else:
            Result = list_all_objects(self, request, *args, **kwargs)
        
        return Result

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    
class CustomerAddressesViews(viewsets.ModelViewSet):
    queryset = CustomerAddresses.objects.all()
    serializer_class = CustomerAddressesSerializers
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = CustomerAddressesFilters 
    ordering_fields = ['created_at', 'updated_at']

    def list(self, request, *args, **kwargs):
        return list_all_objects(self, request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)

class CustomerAttachmentsViews(viewsets.ModelViewSet):
    queryset = CustomerAttachments.objects.all()
    serializer_class = CustomerAttachmentsSerializers
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = CustomerAttachmentsFilters 
    ordering_fields = ['attachment_name', 'created_at', 'updated_at']
    
    def list(self, request, *args, **kwargs):
        return list_all_objects(self, request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)

#==========================================================================  
    
class CustomerCreateViews(APIView):
    def get_object(self, pk):
        try:
            return Customer.objects.get(pk=pk)
        except Customer.DoesNotExist:
            logger.warning(f"Customer with ID {pk} does not exist.")
            return None

    def get(self, request, *args, **kwargs):
        """Handles GET requests to retrieve customers with optional filters and reports."""
        try:
            if "pk" in kwargs:
                result = validate_input_pk(self, kwargs['pk'])
                return result if result else self.retrieve(self, request, *args, **kwargs)

            if request.query_params.get("summary", "false").lower() == "true":
                return self.get_customer_summary(request)
            
            if request.query_params.get("customer_summary_report", "false").lower() == "true":
                return self.get_customer_summary_report(request)
            
            if request.query_params.get("customer_ledger_report", "false").lower() == "true":
                return self.get_customer_ledger_report(request)
            
            if request.query_params.get("customer_outstanding_report", "false").lower() == "true":
                return self.get_customer_outstanding_report(request)
            
            if request.query_params.get("customer_order_history", "false").lower() == "true":
             return self.get_customer_order_history(request)
         
            if request.query_params.get("credit_limit_report", "false").lower() == "true":
             return self.get_customer_credit_limit_report(request)
            
            if request.query_params.get("customer_payment_report", "false").lower() == "true":
             return self.get_customer_payment_report(request)
         
            return self.get_customers(request)

        except Exception as e:
            logger.error(f"An unexpected error occurred: {str(e)}")
            return build_response(0, "An error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_pagination_params(self, request):
        """Extracts pagination parameters from the request."""
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 10))
        return page, limit
    
    def get_customers(self, request):
        """Applies filters, pagination, and retrieves customers."""
        logger.info("Retrieving all customers")

        page, limit = self.get_pagination_params(request)
        using_db = self.resolve_db_from_request(request)

        queryset = Customer.objects.using(using_db).all().order_by('is_deleted', '-created_at')

        if request.query_params:
            filterset = CustomerFilters(request.GET, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
                
        # ✅ Enforce ordering again after filters
        queryset = queryset.order_by('is_deleted', '-created_at')

        total_count = Customer.objects.using(using_db).count()
        serializer = CustomerSerializer(queryset, many=True)

        return filter_response(queryset.count(), "Success", serializer.data, page, limit, total_count, status.HTTP_200_OK)


    def get_customer_summary(self, request):
        """Fetches customer summary data."""
        logger.info("Retrieving customer summary")

        page, limit = self.get_pagination_params(request)
        # using_db = self.resolve_db_from_request(request)

        queryset = Customer.objects.all().order_by('is_deleted', '-created_at')

        if request.query_params:
            filterset = CustomerFilters(request.GET, queryset=queryset.order_by('is_deleted', '-created_at'))
            if filterset.is_valid():
                queryset = filterset.qs

        total_count = Customer.objects.count()
        serializer = CustomerOptionSerializer(queryset, many=True)

        return filter_response(queryset.count(), "Success", serializer.data, page, limit, total_count, status.HTTP_200_OK)

    # def resolve_db_from_request(self, request):
    #     """
    #     Determines the DB to use:
    #     - If `sale_type=Other` --> use 'mstcnl'
    #     - If `bill_type=Others` --> use 'mstcnl' 
    #     - Else → use 'default'
    #     """
    #     bill_type = request.query_params.get("bill_type")
    #     sale_type = request.query_params.get("sale_type")

    #     if (bill_type and bill_type.lower() == 'others') or (sale_type and sale_type.lower() == 'other'):
    #         return 'mstcnl'

    #     return 'default'

    def get_customer_summary_report(self, request):
        """Fetches a summary report of total sales and outstanding payments per customer."""
        logger.info("Retrieving customer summary report data")
        page, limit = self.get_pagination_params(request)
        
       # Annotate each customer with aggregated values.
        queryset = Customer.objects.annotate(
            total_sales=Sum('saleorder__item_value', output_field=DecimalField(max_digits=18, decimal_places=2)),
            total_advance=Sum('saleorder__advance_amount', output_field=DecimalField(max_digits=18, decimal_places=2)),
            outstanding_payments=Sum(
                ExpressionWrapper(
                    F('saleorder__item_value') - F('saleorder__advance_amount'),
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                )
            )
        ).order_by('name')
        if request.query_params:
            filterset = CustomerSummaryReportFilter(request.GET, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
                
        total_count = Customer.objects.count()
        serializer = CustomerSummaryReportSerializer(queryset, many=True)
        return filter_response(queryset.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK)

    
    def get_customer_order_history(self, request):
        """Fetches past purchases by customers."""
        logger.info("Retrieving customer order history report data")

        page, limit = self.get_pagination_params(request)

         # Retrieve all sale invoice orders (sorted by invoice_date descending)
        queryset = SaleInvoiceOrders.objects.all().order_by('-invoice_date')
        
        if request.query_params:
            filterset = CustomerOrderHistoryReportFilter(request.GET, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
                
        total_count = SaleInvoiceOrders.objects.count()        
        serializer = CustomerOrderHistoryReportSerializer(queryset, many=True)
        return filter_response(queryset.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK)

    def get_customer_credit_limit_report(self, request):
        """Fetches assigned credit limits vs. usage per customer."""
        logger.info("Retrieving customer credit limit report data")

        page, limit = self.get_pagination_params(request)
        queryset = Customer.objects.annotate(
            credit_usage=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('saleinvoiceorders__total_amount') - F('saleinvoiceorders__advance_amount'),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    )
                ),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            )
        ).annotate(
            remaining_credit=ExpressionWrapper(
                Coalesce(F('credit_limit'), Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))) - F('credit_usage'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        ).order_by('name')
        # total_count = queryset.count()
        if request.query_params:
            filterset = CustomerCreditLimitReportFilter(request.GET, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
                
        total_count = Customer.objects.count()
        serializer = CustomerCreditLimitReportSerializer(queryset, many=True)
        return filter_response(queryset.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK)
    
    def get_customer_ledger_report(self, request):
        """Fetches all financial transactions with a specific customer and calculates running balance."""
        logger.info("Retrieving customer ledger report data")
        page, limit = self.get_pagination_params(request)
        
        from apps.finance.models import JournalEntryLines
        
        # Get customer_id filter from request params if provided
        customer_id = request.query_params.get('customer_id')
        
        # Start with base queryset filtering only for entries related to customers
        queryset = JournalEntryLines.objects.filter(customer_id__isnull=False).select_related(
            'journal_entry_id', 'customer_id'
        ).order_by('journal_entry_id__entry_date')
        
        # Apply customer filter if provided
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
            
        # Apply filters from request
        if request.query_params:
            filterset = CustomerLedgerReportFilter(request.GET, queryset=queryset)
            if filterset.is_valid():
                queryset = filterset.qs
        
        # Calculate total records before pagination
        total_count = queryset.count()
        
        # Calculate running balance for each transaction
        # First, get all the records to process
        records = list(queryset)
        running_balance = 0
        
        # Process each record to calculate running balance
        for record in records:
            # For customer ledger: debit increases balance, credit decreases
            running_balance += record.debit - record.credit
            record.running_balance = running_balance
        
        # Apply pagination manually after calculating running balance
        if page and limit:
            start_idx = (page - 1) * limit
            end_idx = start_idx + limit
            records = records[start_idx:end_idx]
        
        # Serialize the data with the running balance included
        serializer = CustomerLedgerReportSerializer(records, many=True)
        return filter_response(len(records), "Success", serializer.data, page, limit, total_count, status.HTTP_200_OK)
    
    def get_customer_outstanding_report(self, request):
        """
        Fetches a simplified report of pending payments per customer.
        """
        logger.info("Retrieving simplified customer outstanding report")
        page, limit = self.get_pagination_params(request)
        
        from django.utils import timezone
        from django.db.models import Sum, F, DecimalField, Value, ExpressionWrapper
        from django.db.models.functions import Coalesce
        
        # Start with a base queryset of customers
        queryset = Customer.objects.all()
        
        # Calculate total sales and payments
        customers = queryset.annotate(
            # Total sales (invoiced amount)
            total_sales=Coalesce(
                Sum('saleinvoiceorders__total_amount', output_field=DecimalField(max_digits=18, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
            # Total payments received
            total_paid=Coalesce(
                Sum('saleinvoiceorders__advance_amount', output_field=DecimalField(max_digits=18, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
            # Get the most recent payment date
            last_payment_date=Max('saleinvoiceorders__due_date')
        )
        
        # Calculate total pending amount
        customers = customers.annotate(
            total_pending=ExpressionWrapper(
                F('total_sales') - F('total_paid'),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        )
        
        # Filter out customers with no outstanding payments
        customers = customers.filter(total_pending__gt=0).order_by('-total_pending')
        
        # Apply any additional filters from the request
        if request.query_params:
            filterset = CustomerOutstandingReportFilter(request.GET, queryset=customers)
            if filterset.is_valid():
                customers = filterset.qs
        
        total_count = customers.count()
        
        serializer = CustomerOutstandingReportSerializer(customers, many=True)
        return filter_response(customers.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK )
    
    def retrieve(self, request, *args, **kwargs):
        """
        Retrieves a sale order and its related data (items, attachments, and shipments).
        """
        try:
            pk = kwargs.get('pk')
            if not pk:
                logger.error("Primary key not provided in request.")
                return build_response(0, "Primary key not provided", [], status.HTTP_400_BAD_REQUEST)

            # Retrieve the SaleOrder instance
            customer_data = get_object_or_404(Customer, pk=pk)
            customer_serializer = CustomerSerializer(customer_data)

            # Retrieve related data
            attachments_data = self.get_related_data(CustomerAttachments, CustomerAttachmentsSerializers, 'customer_id', pk)
            addresses_data = self.get_related_data(CustomerAddresses, CustomerAddressesSerializers, 'customer_id', pk)
            
            # Retrieve custom field values
            custom_field_values_data = self.get_related_data(CustomFieldValue, CustomFieldValueSerializer, 'custom_id', pk)


            # Customizing the response data
            custom_data = {
                "customer_data": customer_serializer.data,
                "customer_attachments": attachments_data,
                "customer_addresses": addresses_data,
                "custom_field_values": custom_field_values_data  # Add custom field values
            }
            logger.info("Customers and related data retrieved successfully.")
            return build_response(1, "Success", custom_data, status.HTTP_200_OK) 

        except Http404:
            logger.error("Sale order with pk %s does not exist.", pk)
            return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("An error occurred while retrieving sale order with pk %s: %s", pk, str(e))
            return build_response(0, "An error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_related_data(self, model, serializer_class, filter_field, filter_value):
        try:
            related_data = model.objects.filter(**{filter_field: filter_value})
            serializer = serializer_class(related_data, many=True)
            logger.debug(
                f"Retrieved related data for model {model.__name__} with filter {filter_field}={filter_value}."
            )
            return serializer.data
        except Exception as e:
            logger.exception(
                f"Error retrieving related data for model {model.__name__} with filter {filter_field}={filter_value}: {str(e)}"
            )
            return []
      
    # @transaction.atomic
    # def delete(self, request, pk, *args, **kwargs):
    #     """
    #     Handles the deletion of a sale order and its related attachments and shipments.
    #     """
    #     try:
    #         # Get the Customer instance
    #         instance = Customer.objects.get(pk=pk)
    #         update_field = {'custom_id': 'customer_id'}
    #         # Delete related CustomerAttachments and CustomerAddresses
    #         if not delete_multi_instance(pk, Customer, CustomerAttachments, main_model_field_name='customer_id'):
    #             return build_response(0, "Error deleting related order attachments", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
    #         if not delete_multi_instance(pk, Customer, CustomerAddresses, main_model_field_name='customer_id'):
    #             return build_response(0, "Error deleting related order shipments", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
    #         if not delete_multi_instance(pk, Customer, CustomFieldValue, main_model_field_name='custom_id'):
    #             return build_response(0, "Error deleting related order shipments", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    #         # Delete from mstcnl DB
    #         try:
    #             mst_customer = CustomersMstModel.objects.using('mstcnl').filter(customer_id=pk).first()
    #             if mst_customer:
    #                 mst_customer.delete(using='mstcnl')
    #                 logger.info(f"Customer with ID {pk} deleted from mstcnl database.")
    #             else:
    #                 logger.warning(f"No mstcnl customer found with ID {pk}.")
    #         except Exception as mstcnl_error:
    #             logger.error(f"Error deleting mstcnl customer with ID {pk}: {str(mstcnl_error)}")
    #             return build_response(0, "Error deleting record from mstcnl database", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    #         # Delete the main Customer instance
    #         instance.delete()

    #         logger.info(f"Customer with ID {pk} deleted successfully.")
    #         return build_response(1, "Record deleted successfully", [], status.HTTP_204_NO_CONTENT)
    #     except Customer.DoesNotExist:
    #         logger.warning(f"Customer with ID {pk} does not exist.")
    #         return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
    #     except Exception as e:
    #         logger.error(f"Error deleting Customer with ID {pk}: {str(e)}")
    #         return build_response(0, "Record deletion failed due to an error", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @transaction.atomic
    def delete(self, request, pk, *args, **kwargs):
        """
        Soft-delete the Customer and cleanup related child objects.
        """
        try:
            # Get the Customer instance
            instance = Customer.objects.get(pk=pk)

            # Delete related CustomerAttachments, CustomerAddresses, etc.
            if not delete_multi_instance(pk, Customer, CustomerAttachments, main_model_field_name='customer_id'):
                return build_response(0, "Error deleting related order attachments", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
            if not delete_multi_instance(pk, Customer, CustomerAddresses, main_model_field_name='customer_id'):
                return build_response(0, "Error deleting related order shipments", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
            if not delete_multi_instance(pk, Customer, CustomFieldValue, main_model_field_name='custom_id'):
                return build_response(0, "Error deleting related custom fields", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Delete from mstcnl DB (optional: you may soft delete here too)
            try:
                mst_customer = CustomersMstModel.objects.using('mstcnl').filter(customer_id=pk).first()
                if mst_customer:
                    # 👉 Soft delete here too if needed:
                    mst_customer.is_deleted = True
                    mst_customer.save(using='mstcnl')
                    logger.info(f"Customer with ID {pk} soft-deleted in mstcnl database.")
                else:
                    logger.warning(f"No mstcnl customer found with ID {pk}.")
            except Exception as mstcnl_error:
                logger.error(f"Error deleting mstcnl customer with ID {pk}: {str(mstcnl_error)}")
                return build_response(0, "Error deleting record from mstcnl database", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

            # 🔑 👉 Instead of real delete:
            instance.is_deleted = True
            instance.save()

            logger.info(f"Customer with ID {pk} marked as is_deleted successfully.")
            return build_response(1, "Record deleted successfully", [], status.HTTP_204_NO_CONTENT)

        except Customer.DoesNotExist:
            logger.warning(f"Customer with ID {pk} does not exist.")
            return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error deleting Customer with ID {pk}: {str(e)}")
            return build_response(0, "Record deletion failed due to an error", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def patch(self, request, pk, *args, **kwargs):
        """
        Restores a soft-deleted Customer record (is_deleted=True → is_deleted=False).
        """
        try:
            instance = Customer.objects.get(pk=pk)

            if not instance.is_deleted:
                logger.info(f"Customer with ID {pk} is already active.")
                return build_response(0, "Record is already active", [], status.HTTP_400_BAD_REQUEST)

            instance.is_deleted = False
            instance.save()

            logger.info(f"Customer with ID {pk} restored successfully.")
            return build_response(1, "Record restored successfully", [], status.HTTP_200_OK)

        except Customer.DoesNotExist:
            logger.warning(f"Customer with ID {pk} does not exist.")
            return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error restoring Customer with ID {pk}: {str(e)}")
            return build_response(0, "Record restoration failed due to an error", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    # Handling POST requests for creating
    def post(self, request, *args, **kwargs):   #To avoid the error this method should be written [error : "detail": "Method \"POST\" not allowed."]
        return self.create(request, *args, **kwargs)
    
    # def create(self, request, *args, **kwargs):
    #     given_data = request.data
    #     print("Given data:", given_data)
    #     # db_name = set_db('default')
        
    #     # print("Db name : ", db_name)
    #     # Extract customer_data from the request
    #     customer_data = given_data.pop('customer_data', None)

    #     # Validate customer_data
    #     if customer_data:
    #         # Check if 'picture' exists in customer_data and is a list
    #         picture_data = customer_data.get('picture', None)
    #         if picture_data:
    #             if not isinstance(picture_data, list):
    #                 return build_response(0, "'picture' field in customer_data must be a list.", [], status.HTTP_400_BAD_REQUEST)

    #             for attachment in picture_data:
    #                 if not all(key in attachment for key in ['uid', 'name', 'attachment_name', 'file_size', 'attachment_path']):
    #                     return build_response(0, "Missing required fields in some picture data.", [], status.HTTP_400_BAD_REQUEST)
            
    #         # Validate the rest of customer_data
    #         customer_error = validate_payload_data(self, customer_data, CustomerSerializer)
    #     else:
    #         customer_error = ["customer_data is required."]

    #     # Validate customer_attachments
    #     attachments_data = given_data.pop('customer_attachments', None)
    #     if attachments_data:
    #         attachment_error = validate_multiple_data(self, attachments_data, CustomerAttachmentsSerializers, ['customer_id'])
    #     else:
    #         attachment_error = []

    #     # Validate customer_addresses
    #     addresses_data = given_data.pop('customer_addresses', None)
    #     if addresses_data:
    #         addresses_error = validate_multiple_data(self, addresses_data, CustomerAddressesSerializers, ['customer_id'])
    #     else:
    #         addresses_error = []
            
    #     # Validate custom_field_values
    #     custom_fields_data = given_data.get('custom_field_values', None)
    #     if custom_fields_data:
    #         custom_fields_error = validate_multiple_data(self, custom_fields_data, CustomFieldValueSerializer, ['custom_id'])
    #     else:
    #         custom_fields_error = []

    #     # Check for mandatory fields
    #     if not customer_data or not addresses_data:
    #         logger.error("Customers, Customer Addresses data are mandatory but not provided.")
    #         return build_response(0, "Customers, Customer Addresses data are mandatory", [], status.HTTP_400_BAD_REQUEST)

    #     # Collect validation errors
    #     errors = {}
    #     if customer_error:
    #         errors["customer_data"] = customer_error
    #     if attachment_error:
    #         errors['customer_attachments'] = attachment_error
    #     if addresses_error:
    #         errors['customer_addresses'] = addresses_error
    #     if custom_fields_error:
    #         errors['custom_field_values'] = custom_fields_error
    #     if errors:
    #         return build_response(0, "ValidationError:", errors, status.HTTP_400_BAD_REQUEST)

    #     # Create Customer Data
    #     new_customer_data = generic_data_creation(self, [customer_data], CustomerSerializer)
    #     customer_id = new_customer_data[0].get("customer_id", None)
    #     logger.info('Customer - created*')

    #     # Create CustomerAttachment Data
    #     update_fields = {'customer_id': customer_id}
    #     if attachments_data:
    #         attachments_data = generic_data_creation(self, attachments_data, CustomerAttachmentsSerializers, update_fields)
    #         logger.info('CustomerAttachments - created*')
    #     else:
    #         attachments_data = []

    #     # Create CustomerAddress Data
    #     addresses_data = generic_data_creation(self, addresses_data, CustomerAddressesSerializers, update_fields)
    #     logger.info('CustomerAddresses - created*')

    #     # Handle CustomFieldValues
    #     custom_fields_data = given_data.pop('custom_field_values', None)
    #     if custom_fields_data:
    #         # Link each custom field value to the new customer
    #         for custom_field in custom_fields_data:
    #             custom_field['custom_id'] = customer_id  # Set the newly created customer ID
    #         custom_fields_data = generic_data_creation(self, custom_fields_data, CustomFieldValueSerializer)
    #         logger.info('CustomFieldValues - created*')
    #     else:
    #         custom_fields_data = []

    #     # Build the response with all created data
    #     custom_data = [
    #         {"customer_data": new_customer_data[0]},
    #         {"customer_attachments": attachments_data},
    #         {"customer_addresses": addresses_data},
    #         {"custom_field_values": custom_fields_data}
    #     ]

    #     return build_response(1, "Record created successfully", custom_data, status.HTTP_201_CREATED)
    
    def create(self, request, *args, **kwargs):
        given_data = request.data
        print("Given data:", given_data)
        # use_db = 'default'
        # Extract customer_data from the request
        customer_data = given_data.pop('customer_data', None)

        # Validate customer_data
        if customer_data:
            # Check if 'picture' exists in customer_data and is a list
            picture_data = customer_data.get('picture', None)
            if picture_data:
                if not isinstance(picture_data, list):
                    return build_response(0, "'picture' field in customer_data must be a list.", [], status.HTTP_400_BAD_REQUEST)

                for attachment in picture_data:
                    if not all(key in attachment for key in ['uid', 'name', 'attachment_name', 'file_size', 'attachment_path']):
                        return build_response(0, "Missing required fields in some picture data.", [], status.HTTP_400_BAD_REQUEST)

            # Validate the rest of customer_data
            customer_error = validate_payload_data(self, customer_data, CustomerSerializer)
        else:
            customer_error = ["customer_data is required."]

        # Validate customer_attachments
        attachments_data = given_data.pop('customer_attachments', None)
        if attachments_data:
            attachment_error = validate_multiple_data(self, attachments_data, CustomerAttachmentsSerializers, ['customer_id'], using_db=set_db('default'))
        else:
            attachment_error = []

        # Validate customer_addresses
        addresses_data = given_data.pop('customer_addresses', None)
        if addresses_data:
            addresses_error = validate_multiple_data(self, addresses_data, CustomerAddressesSerializers, ['customer_id'], using_db=set_db('default'))
        else:
            addresses_error = []

        # Validate custom_field_values
        custom_fields_data = given_data.get('custom_field_values', None)
        if custom_fields_data:
            custom_fields_error = validate_multiple_data(self, custom_fields_data, CustomFieldValueSerializer, ['custom_id'], using_db=set_db('default'))
        else:
            custom_fields_error = []

        # Check for mandatory fields
        if not customer_data or not addresses_data:
            logger.error("Customers, Customer Addresses data are mandatory but not provided.")
            return build_response(0, "Customers, Customer Addresses data are mandatory", [], status.HTTP_400_BAD_REQUEST)

        # Collect validation errors
        errors = {}
        if customer_error:
            errors["customer_data"] = customer_error
        if attachment_error:
            errors['customer_attachments'] = attachment_error
        if addresses_error:
            errors['customer_addresses'] = addresses_error
        if custom_fields_error:
            errors['custom_field_values'] = custom_fields_error
        if errors:
            return build_response(0, "ValidationError: {errors}", errors, status.HTTP_400_BAD_REQUEST)

        # Step 1: Create Customer Data in devcnl
        new_customer_data = generic_data_creation(self, [customer_data], CustomerSerializer, using=set_db('default'))
        customer_id = new_customer_data[0].get("customer_id", None)
        logger.info('Customer - created*')

        # # Step 2: Create entry in mstcnl.customers table
        # from your_app.models import Companies, CustomersMstModel  # Adjust this import based on your project structure

        customer_name = new_customer_data[0].get("name")
        phone = new_customer_data[0].get("phone")
        email = new_customer_data[0].get("email")
        company_id = new_customer_data[0].get("company_id")

        try:
            company = Companies.objects.first()
            print("-"*20)
            print("company : ", company)
            print("-"*20)
            company_id = company.company_id
            company_name = company.name
            phone = addresses_data[0].get("phone") if addresses_data else None
            print("phone : ", phone)
            email = addresses_data[0].get("email") if addresses_data else None
            print("email : ", email)
            # Create the record in mstcnl DB
            CustomersMstModel.objects.using('mstcnl').create(
                customer_id=customer_id,
                name=customer_name,
                phone=phone,
                email=email,
                company_id=company_id,
                company_name=company_name
            )
            logger.info("Customer also created in mstcnl.customers table")

        except Companies.DoesNotExist:
            logger.warning(f"Company with ID {company_id} not found. Skipping mstcnl customer creation.")

        # Step 3: Create CustomerAttachment Data
        update_fields = {'customer_id': customer_id}
        if attachments_data:
            attachments_data = generic_data_creation(self, attachments_data, CustomerAttachmentsSerializers, update_fields, using=set_db('default'))
            logger.info('CustomerAttachments - created*')
        else:
            attachments_data = []

        # Step 4: Create CustomerAddress Data
        addresses_data = generic_data_creation(self, addresses_data, CustomerAddressesSerializers, update_fields, using=set_db('default'))
        logger.info('CustomerAddresses - created*')

        # Step 5: Handle CustomFieldValues
        custom_fields_data = given_data.pop('custom_field_values', None)
        if custom_fields_data:
            for custom_field in custom_fields_data:
                custom_field['custom_id'] = customer_id
            custom_fields_data = generic_data_creation(self, custom_fields_data, CustomFieldValueSerializer, using=set_db('default'))
            logger.info('CustomFieldValues - created*')
        else:
            custom_fields_data = []

        # Step 6: Build response
        custom_data = [
            {"customer_data": new_customer_data[0]},
            {"customer_attachments": attachments_data},
            {"customer_addresses": addresses_data},
            {"custom_field_values": custom_fields_data}
        ]
        
        customer_name = customer_data.get("name")
        
        log_user_action(
            set_db('default'),
            request.user,
            "CREATE",
            "Customers",
            customer_id,
            f"{customer_name} - Customer record created by {request.user.username}"
        )

        return build_response(1, "Record created successfully", custom_data, status.HTTP_201_CREATED)


    
    def put(self, request, pk, *args, **kwargs):

            #----------------------------------- D A T A  V A L I D A T I O N -----------------------------#
            """
            All the data in request will be validated here. it will handle the following errors:
            - Invalid data types
            - Invalid foreign keys
            - nulls in required fields
            """
            # Get the given data from request
            given_data = request.data  

            # Vlidated Customer Data
            customer_data = given_data.pop('customer_data', None)
            if customer_data:                
                customer_error = validate_payload_data(self, customer_data , CustomerSerializer)

            # Vlidated CustomerAttachment Data
            attachments_data = given_data.pop('customer_attachments', None)
            if attachments_data:
                exclude_fields = ['customer_id']
                attachments_error = validate_put_method_data(self, attachments_data,CustomerAttachmentsSerializers, exclude_fields, CustomerAttachments, current_model_pk_field='attachment_id')
            else:
                attachments_error = [] # Since 'CustomerAttachment' is optional, so making an error is empty list

            # Vlidated CustomerAddresses Data
            addresses_data = given_data.pop('customer_addresses', None)
            if addresses_data:
                exclude_fields = ['customer_id']
                addresses_error = validate_put_method_data(self, addresses_data,CustomerAddressesSerializers, exclude_fields, CustomerAddresses, current_model_pk_field='customer_addresses_id')
                
            # Validated CustomFieldValues Data
            custom_field_values_data = given_data.pop('custom_field_values', None)
            if custom_field_values_data:
                exclude_fields = ['entity_data_id']
                custom_field_values_error = validate_put_method_data(self, custom_field_values_data, CustomFieldValueSerializer, exclude_fields, CustomFieldValue, current_model_pk_field='custom_field_value_id')
            else:
                custom_field_values_error = []  # Optional, so initialize as an empty list

            # Ensure mandatory data is present
            if not customer_data or not addresses_data:
                logger.error("Customer data and Customer addresses & CustomFields data are mandatory but not provided.")
                return build_response(0, "Customer and Customer addresses & CustomFields are mandatory", [], status.HTTP_400_BAD_REQUEST)
            
            errors = {}
            if customer_error:
                errors["customer_data"] = customer_error
            if attachments_error:
                errors["customer_attachments"] = attachments_error
            if addresses_error:
                errors['customer_addresses'] = addresses_error
            if custom_field_values_error:
                errors['custom_field_values'] = custom_field_values_error
            if errors:
                return build_response(0, "ValidationError :",errors, status.HTTP_400_BAD_REQUEST)
            
            # ------------------------------ D A T A   U P D A T I O N -----------------------------------------#
            if customer_data:
                update_fields = []# No need to update any fields
                customer_data = update_multi_instances(self, pk, [customer_data], Customer, CustomerSerializer, update_fields,main_model_related_field='customer_id', current_model_pk_field='customer_id')

            # Update CustomerAttachment Data
            update_fields = {'customer_id':pk}
            attachments_data = update_multi_instances(self,pk, attachments_data,CustomerAttachments,CustomerAttachmentsSerializers, update_fields, main_model_related_field='customer_id', current_model_pk_field='attachment_id')

            # Update CustomerAddress Data
            addresses_data = update_multi_instances(self,pk, addresses_data,CustomerAddresses, CustomerAddressesSerializers, update_fields, main_model_related_field='customer_id', current_model_pk_field='customer_address_id')
            
            # Update CustomFieldValues Data
            if custom_field_values_data:
                custom_field_values_data = update_multi_instances(self, pk, custom_field_values_data, CustomFieldValue, CustomFieldValueSerializer, {}, main_model_related_field='custom_id', current_model_pk_field='custom_field_value_id')

            # --------------------------- M S T C N L   S Y N C ----------------------------------- #
            try:
                mst_customer = CustomersMstModel.objects.using('mstcnl').filter(customer_id=pk).first()

                if mst_customer:
                    customer_name = customer_data[0].get('name')
                    company = Companies.objects.first()
                    phone = addresses_data[0].get("phone") if addresses_data else None
                    email = addresses_data[0].get("email") if addresses_data else None
                    
                    # Update fields explicitly
                    mst_customer.name = customer_name
                    mst_customer.phone = phone
                    mst_customer.email = email
                    mst_customer.company_id = company.company_id if company else None
                    mst_customer.company_name = company.name if company else None
                    mst_customer.save(using='mstcnl')

                    logger.info("Customer updated successfully in mstcnl database.")
                else:
                    logger.warning(f"No customer found in mstcnl with customer_id = {pk}")
            except Exception as e:
                logger.warning(f"Error updating mstcnl record: {str(e)}")
                
            custom_data = [
                {"customer_data":customer_data},
                {"customer_attachments":attachments_data if attachments_data else []},
                {"customer_addresses":addresses_data if addresses_data else []},
                {"custom_field_values": custom_field_values_data if custom_field_values_data else []}  # Add custom field values to response
            ]
            
            customername = customer_data[0].get("name") if customer_data else None

            log_user_action(
                set_db('default'),
                request.user,
                "UPDATE",
                "Customers",
                pk,
                f"{customername} - Customer record Updated by {request.user.username}"
            )

            return build_response(1, "Records updated successfully", custom_data, status.HTTP_200_OK)
        
    def get_customer_payment_report(self, request):
        """
        Fetches a report of all payments received from customers.
        Shows payment details with customer and invoice information.
        """
        logger.info("Retrieving customer payment report data")
        page, limit = self.get_pagination_params(request)
        
        try:
            # Import required finance models
            from apps.finance.models import PaymentTransaction
            from apps.customer.filters import CustomerPaymentReportFilter
            
            # Base query - filter payments related to sales (received from customers)
            queryset = PaymentTransaction.objects.filter(
                order_type='Sale',
                transaction_type='Credit'  # Credit transactions represent payments received from customers
            )
            
            # Order by payment date (most recent first)
            queryset = queryset.order_by('-payment_date')
            
            # Apply any additional filters from the request
            if request.query_params:
                filterset = CustomerPaymentReportFilter(request.GET, queryset=queryset)
                if filterset.is_valid():
                    queryset = filterset.qs
            
            # Get total count for pagination
            total_count = queryset.count()
            
            # Apply pagination
            if page and limit:
                start_idx = (page - 1) * limit
                end_idx = start_idx + limit
                queryset = queryset[start_idx:end_idx]
            
            # Serialize the filtered data
            serializer = CustomerPaymentReportSerializer(queryset, many=True)
            
            return filter_response(queryset.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error generating customer payment report: {str(e)}")
            return build_response(0, "An error occurred while generating the report", [], status.HTTP_500_INTERNAL_SERVER_ERROR)



class CustomerBulkUpdateView(BaseBulkUpdateView):
    """
    PATCH /customers/bulk-update/
    
    Body:
    {
        "ids": ["uuid1", "uuid2", ...],
        "update_data": {
            "customer_category_id": "uuid",
            "territory_id": "uuid",
            "credit_limit": 50000
        }
    }
    
    Only filled fields are updated. Empty fields are ignored.
    """
    MODEL = Customer
    PK_FIELD = "customer_id"
    MODULE_NAME = "Customers"
    MAX_SELECTION = 100

    # Whitelist: field_name → FK Model (None = plain value field)
    ALLOWED_FIELDS = {
        'customer_category_id': CustomerCategories,
        'territory_id':         Territory,
        'firm_status_id':       FirmStatuses,
        'gst_category_id':      GstCategories,
        'payment_term_id':      CustomerPaymentTerms,
        'price_category_id':    PriceCategories,
        'transporter_id':       Transporters,
        'tax_type':             None,
        'credit_limit':         None,
        'max_credit_days':      None,
        'interest_rate_yearly':  None,
        'tds_applicable':       None,
        'tds_on_gst_applicable': None,
        'gst_suspend':          None,
    }


class CustomerBalanceView(APIView):
    def post(self, request, customer_id, remaining_payment): 
        '''update_customer_balance_after payment transaction. This is used in the apps.sales.view.PaymentTransactionAPIView class.'''  
        try:
            customer_instance = Customer.objects.get(customer_id=customer_id)
            customer_balance = CustomerBalance.objects.filter(customer_id=customer_instance)

            if customer_balance.exists():
                # Update balance if customer balance already exists
                for balance in customer_balance:
                    customer_balance.update(balance_amount= balance.balance_amount + remaining_payment)
            else:
                # Create a new CustomerBalance entry if it doesn't exist
                CustomerBalance.objects.create(customer_id=customer_instance, balance_amount=remaining_payment)
    
        except ObjectDoesNotExist as e:
            return build_response(1, f"Customer with ID {customer_id} does not exist.", str(e), status.HTTP_404_NOT_FOUND)

        return build_response(1, "Balance Updated In Customer Balance Table", [], status.HTTP_201_CREATED)
    
    def get(self, request, pk=None):
        if pk:
            try:
                customer_id = get_object_or_404(Customer, pk=pk)
                balance = get_object_or_404(CustomerBalance, customer_id=customer_id)
                serializer = CustomerBalanceSerializer(balance)
                return build_response(1, "Customer Balance", serializer.data, status.HTTP_200_OK)
            except Exception as e:
                return build_response(1, "Something Went Wrong", str(e), status.HTTP_403_FORBIDDEN)
        else:
            balances = CustomerBalance.objects.all()
            serializer = CustomerBalanceSerializer(balances, many=True)
            return build_response(len(serializer.data), "Customer Balance", serializer.data, status.HTTP_200_OK)

    def update():
        pass


   
# ------------------------------
# Lookup Field Mappings
# ------------------------------
CUSTOMER_LOOKUP_FIELDS = {
    'firm_status_id': 'firm_status',
    'territory_id': 'territory',
    'customer_category_id': 'customer_category',
    'gst_category_id': 'gst_category',
    'payment_term_id': 'payment_term',
    'price_category_id': 'price_category',
    'transporter_id': 'transporter',
    'ledger_account_id': 'ledger_account',
}

ADDRESS_LOOKUP_FIELDS = {
    'city_id': 'city',
    'state_id': 'state',
    'country_id': 'country',
}

REQUIRED_CUSTOMER_FIELDS = ["name", "print_name", "customer_category"]
REQUIRED_ADDRESS_FIELDS = []

def download_customer_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "CustomerTemplate"

    # ------------------------------
    # 1. CUSTOMER HEADERS
    # ------------------------------
    customer_fields = []
    for field in Customer._meta.get_fields():
        if field.auto_created or field.name in ['customer_id', 'created_at', 'updated_at', 'picture']:
            continue
        if field.name in CUSTOMER_LOOKUP_FIELDS:
            customer_fields.append(CUSTOMER_LOOKUP_FIELDS[field.name])
        else:
            customer_fields.append(field.name)


    # ------------------------------
    # 2. ADDRESS HEADERS
    # ------------------------------
    def get_address_fields(prefix):
        result_fields = []
        for field in CustomerAddresses._meta.get_fields():
            if field.auto_created or field.name in ['customer_address_id', 'customer_id', 'created_at', 'updated_at', 'address_type', 'route_map']:
                continue
            if field.name in ADDRESS_LOOKUP_FIELDS:
                result_fields.append(f"{prefix}_{ADDRESS_LOOKUP_FIELDS[field.name]}")
            else:
                result_fields.append(f"{prefix}_{field.name}")
        return result_fields

    billing_fields = get_address_fields("billing")
    shipping_fields = get_address_fields("shipping")

    # Add ledger_group next to ledger_account
    ledger_account_index = customer_fields.index('ledger_account')
    customer_fields.insert(ledger_account_index + 1, 'ledger_group')
    
    # ------------------------------
    # 3. Combine All Headers (explicitly NOT including shipping_same_as_billing)
    # ------------------------------
    headers = customer_fields + billing_fields + shipping_fields
    ws.append(headers)

    # ------------------------------
    # 4. Header Styling + Required Marking
    # ------------------------------
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center")

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = font_bold
        cell.alignment = align_center

        is_required = False
        if column_title in REQUIRED_CUSTOMER_FIELDS:
            is_required = True
        elif column_title.startswith("billing_") or column_title.startswith("shipping_"):
            for req in REQUIRED_ADDRESS_FIELDS:
                if column_title.endswith(req.replace("_id", "")) or column_title.endswith(req):
                    is_required = True
                    break

        if is_required:
            cell.comment = Comment("This field is required", "System")

        ws.column_dimensions[get_column_letter(col_num)].width = len(str(column_title)) + 4

    # ------------------------------
    # 5. Return Response
    # ------------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Customer_Import_Template.xlsx'
    wb.save(response)
    return response



# ------------------------------
# Customer Excel Import
# ------------------------------
class CustomerExcelImport(BaseExcelImportExport):
    """
    Customer Excel import/export functionality
    """
    MODEL_CLASS = Customer
    SERIALIZER_CLASS = CustomerSerializer
    REQUIRED_COLUMNS = ["name", "print_name", "customer_category"]
    TEMPLATE_FILENAME = "Customer_Import_Template.xlsx"
    
    FIELD_MAP = {
        "name": "name",
        "print_name": "print_name",
        "identification": "identification",
        "code": "code",
        "ledger_account": ("ledger_account_id", LedgerAccounts),
        "ledger_group": "ledger_group",  # For creating LedgerAccounts with proper group
        "customer_common_for_sales_purchase": "customer_common_for_sales_purchase",
        "is_sub_customer": "is_sub_customer",
        "firm_status": ("firm_status_id", FirmStatuses),
        "territory": ("territory_id", Territory),
        "customer_category": ("customer_category_id", CustomerCategories),
        "contact_person": "contact_person",
        "gst": "gst",
        "registration_date": "registration_date",
        "cin": "cin",
        "pan": "pan",
        "gst_category": ("gst_category_id", GstCategories),
        "gst_suspend": "gst_suspend",
        "tax_type": "tax_type",
        "distance": "distance",
        "tds_on_gst_applicable": "tds_on_gst_applicable",
        "tds_applicable": "tds_applicable",
        "website": "website",
        "facebook": "facebook",
        "skype": "skype",
        "twitter": "twitter",
        "linked_in": "linked_in",
        "payment_term": ("payment_term_id", CustomerPaymentTerms),
        "price_category": ("price_category_id", PriceCategories),
        "batch_rate_category": "batch_rate_category",
        "transporter": ("transporter_id", Transporters),
        "credit_limit": "credit_limit",
        "max_credit_days": "max_credit_days",
        "interest_rate_yearly": "interest_rate_yearly",
    }
    
    BOOLEAN_FIELDS = [
        "customer_common_for_sales_purchase", 
        "is_sub_customer", 
        "gst_suspend", 
        "tds_applicable", 
        "tds_on_gst_applicable"
    ]
    
    # ============================================================
    # FK BULK CONFIG - Clean, DRY configuration for bulk imports
    # ============================================================
    FK_BULK_CONFIG = {
        # Basic FKs with name field
        'customer_category': {
            'model': CustomerCategories,
            'name_field': 'name',
            'create_fields': {'code': lambda name: name[:3].upper()}
        },
        'firm_status': {
            'model': FirmStatuses,
            'name_field': 'name',
            'create_fields': {}
        },
        'territory': {
            'model': Territory,
            'name_field': 'name',
            'create_fields': {'code': lambda name: name[:3].upper()}
        },
        'gst_category': {
            'model': GstCategories,
            'name_field': 'name',
            'create_fields': {}
        },
        'payment_term': {
            'model': CustomerPaymentTerms,
            'name_field': 'name',
            'create_fields': {}
        },
        'price_category': {
            'model': PriceCategories,
            'name_field': 'name',
            'create_fields': {}
        },
        'transporter': {
            'model': Transporters,
            'name_field': 'name',
            'create_fields': {'code': lambda name: name[:3].upper()}
        },
        'ledger_group': {
            'model': LedgerGroups,
            'name_field': 'name',
            'create_fields': {}
        },
        # Location FKs
        'country': {
            'model': Country,
            'name_field': 'country_name',
            'create_fields': {'country_code': lambda name: name[:3].upper()}
        },
        'state': {
            'model': State,
            'name_field': 'state_name',
            'create_fields': {}  # Requires country_id, handle separately
        },
        'city': {
            'model': City,
            'name_field': 'city_name',
            'create_fields': {}  # Requires state_id, handle separately
        },
    }
    
    # Special handling for foreign keys that require additional fields
    FK_REQUIRED_FIELDS = {
        "LedgerAccounts": {
            "ledger_group_id": lambda: LedgerGroups.objects.first()
        }
    }
    
    @classmethod
    def create_record(cls, row_data, field_map=None, boolean_fields=None, get_or_create_funcs=None):
        """Create customer record with addresses from Excel data"""
        field_map = field_map or cls.FIELD_MAP
        boolean_fields = boolean_fields or cls.BOOLEAN_FIELDS
        
        with transaction.atomic():
            # 1. Process the customer data
            customer_data = {}
            
            # Special handling for ledger account with group
            ledger_account_name = row_data.get("ledger_account")
            ledger_group_name = row_data.get("ledger_group")
            
            if ledger_account_name:
                if ledger_group_name:
                    # Get or create the ledger group first
                    ledger_group = LedgerGroups.objects.filter(name__iexact=ledger_group_name).first()
                    if not ledger_group:
                        logger.info(f"Creating LedgerGroup with name='{ledger_group_name}'")
                        ledger_group = LedgerGroups.objects.create(name=ledger_group_name)
                        
                    # Now get or create ledger account with the right group
                    ledger_account = LedgerAccounts.objects.filter(name__iexact=ledger_account_name).first()
                    if not ledger_account:
                        logger.info(f"Creating LedgerAccount with name='{ledger_account_name}' and group='{ledger_group_name}'")
                        ledger_account = LedgerAccounts.objects.create(
                            name=ledger_account_name,
                            ledger_group_id=ledger_group
                        )
                else:
                    # If no group specified, use default group
                    default_group = LedgerGroups.objects.first()
                    ledger_account = LedgerAccounts.objects.filter(name__iexact=ledger_account_name).first()
                    if not ledger_account:
                        if not default_group:
                            raise ValueError("Cannot create LedgerAccount without a LedgerGroup")
                        logger.info(f"Creating LedgerAccount with name='{ledger_account_name}' and default group")
                        ledger_account = LedgerAccounts.objects.create(
                            name=ledger_account_name,
                            ledger_group_id=default_group
                        )
                
                customer_data['ledger_account_id'] = ledger_account
            
            # Process other fields
            for excel_col, mapping in field_map.items():
                # Skip already processed fields
                if excel_col in ['ledger_account', 'ledger_group']:
                    continue
                    
                value = row_data.get(excel_col)
                
                # Skip empty values
                if value is None or value == '':
                    continue
                
                if isinstance(mapping, tuple):
                    # Handle foreign key fields
                    field_name, model = mapping
                    customer_data[field_name] = cls.get_or_create_fk(model, value)
                else:
                    # Handle regular fields
                    if mapping in boolean_fields:
                        customer_data[mapping] = cls.parse_boolean(value)
                    elif mapping == 'tax_type':
                        # Special validation for tax_type to ensure it matches allowed choices
                        valid_tax_types = ['Inclusive', 'Exclusive', 'Both']
                        if value not in valid_tax_types:
                            raise ValueError(f"Invalid tax_type '{value}'. Must be one of: {', '.join(valid_tax_types)}")
                        customer_data[mapping] = value
                    else:
                        customer_data[mapping] = value
            
            # Create the customer record
            customer = Customer.objects.create(**customer_data)
            
            # 2. Create billing address - check if any billing field has data
            billing_fields = ['billing_address', 'billing_country', 'billing_state', 'billing_city', 
                            'billing_pin_code', 'billing_phone', 'billing_email', 
                            'billing_longitude', 'billing_latitude']
            has_billing_data = any(row_data.get(field) for field in billing_fields)
            
            if has_billing_data:
                # Get or create location fields INDEPENDENTLY
                # Each field is saved regardless of whether others exist
                country = cls.get_or_create_country(row_data.get("billing_country"))
                state = cls.get_or_create_state(row_data.get("billing_state"), country)
                # Pass country as fallback for city creation when state is missing
                city = cls.get_or_create_city(row_data.get("billing_city"), state, country)
                
                try:
                    CustomerAddresses.objects.create(
                        customer_id=customer,
                        address_type="Billing",
                        address=row_data.get("billing_address") or None,
                        city_id=city,
                        state_id=state,
                        country_id=country,
                        pin_code=row_data.get("billing_pin_code") or None,
                        phone=row_data.get("billing_phone") or None,
                        email=row_data.get("billing_email") or None,
                        longitude=row_data.get("billing_longitude") or None,
                        latitude=row_data.get("billing_latitude") or None
                    )
                    logger.info(f"Created billing address for customer {customer.name}")
                except Exception as e:
                    logger.warning(f"Could not create billing address for customer {customer.name}: {e}")
            
            # 3. Create shipping address - check if any shipping field has data
            shipping_fields = ['shipping_address', 'shipping_country', 'shipping_state', 'shipping_city',
                             'shipping_pin_code', 'shipping_phone', 'shipping_email',
                             'shipping_longitude', 'shipping_latitude']
            has_shipping_data = any(row_data.get(field) for field in shipping_fields)
            
            if has_shipping_data:
                # Get or create location fields INDEPENDENTLY
                # Each field is saved regardless of whether others exist
                s_country = cls.get_or_create_country(row_data.get("shipping_country"))
                s_state = cls.get_or_create_state(row_data.get("shipping_state"), s_country)
                # Pass country as fallback for city creation when state is missing
                s_city = cls.get_or_create_city(row_data.get("shipping_city"), s_state, s_country)
                
                try:
                    CustomerAddresses.objects.create(
                        customer_id=customer,
                        address_type="Shipping",
                        address=row_data.get("shipping_address") or None,
                        city_id=s_city,
                        state_id=s_state,
                        country_id=s_country,
                        pin_code=row_data.get("shipping_pin_code") or None,
                        phone=row_data.get("shipping_phone") or None,
                        email=row_data.get("shipping_email") or None,
                        longitude=row_data.get("shipping_longitude") or None,
                        latitude=row_data.get("shipping_latitude") or None
                    )
                    logger.info(f"Created shipping address for customer {customer.name}")
                except Exception as e:
                    logger.warning(f"Could not create shipping address for customer {customer.name}: {e}")
                
            return customer

    @classmethod
    def bulk_create_records(cls, validated_rows, field_map=None, boolean_fields=None):
        """
        Bulk create customer records for maximum performance.
        
        Uses helper methods from BaseExcelImportExport for clean, DRY code.
        5-10x faster than create_record() by using bulk_create and pre-loaded FK lookups.
        
        Args:
            validated_rows: List of row dictionaries with '_excel_row' for tracking
            field_map: Field mapping dictionary
            boolean_fields: List of boolean field names
            
        Returns:
            dict with 'success' count and 'errors' list
        """
        field_map = field_map or cls.FIELD_MAP
        boolean_fields = boolean_fields or cls.BOOLEAN_FIELDS
        
        errors = []
        customers_to_create = []
        total_rows = len(validated_rows)
        BATCH_SIZE = 500
        
        logger.info(f"BULK CREATE: Starting for {total_rows} customers")
        
        # ========================================
        # PHASE 1: Load all FK lookups using helper
        # ========================================
        logger.info("PHASE 1: Loading FK lookups...")
        
        # Add ledger_account to config (special case - uses 'name' field)
        fk_config = {
            **cls.FK_BULK_CONFIG,
            'ledger_account': {'model': LedgerAccounts, 'name_field': 'name'}
        }
        fk_lookups = cls.bulk_load_fk_lookups(fk_config)
        
        # ========================================
        # PHASE 2: Collect and create missing FKs using helper
        # ========================================
        logger.info("PHASE 2: Creating missing FKs...")
        
        # Only auto-create these simple FKs (location FKs need special handling)
        simple_fk_config = {k: v for k, v in cls.FK_BULK_CONFIG.items() 
                           if k not in ['country', 'state', 'city']}
        
        missing_fks = cls.bulk_collect_missing_fks(validated_rows, simple_fk_config, fk_lookups)
        fk_lookups = cls.bulk_create_missing_fks(missing_fks, simple_fk_config, fk_lookups)
        
        logger.info("PHASE 2 complete")
        
        # ========================================
        # PHASE 3: Handle LedgerAccounts (special - requires ledger_group)
        # ========================================
        logger.info("PHASE 3: Creating missing Ledger Accounts...")
        
        missing_ledger_accounts = {}
        for row_data in validated_rows:
            ledger_name = row_data.get('ledger_account')
            if ledger_name and str(ledger_name).strip():
                ledger_lower = str(ledger_name).strip().lower()
                if ledger_lower not in fk_lookups['ledger_account']:
                    group_name = row_data.get('ledger_group', '')
                    missing_ledger_accounts[ledger_lower] = {
                        'name': str(ledger_name).strip(),
                        'group': str(group_name).strip() if group_name else None
                    }
        
        if missing_ledger_accounts:
            with transaction.atomic():
                default_group = LedgerGroups.objects.first()
                for ledger_lower, info in missing_ledger_accounts.items():
                    group = fk_lookups['ledger_group'].get(info['group'].lower()) if info['group'] else None
                    group = group or default_group
                    
                    if group:
                        ledger = LedgerAccounts.objects.create(name=info['name'], ledger_group_id=group)
                        fk_lookups['ledger_account'][ledger_lower] = ledger
                
                logger.info(f"Created {len(missing_ledger_accounts)} Ledger Accounts")
        
        logger.info("PHASE 3 complete")
        
        # ========================================
        # PHASE 4: Build Customer objects
        # ========================================
        logger.info("PHASE 4: Building Customer objects...")
        
        for idx, row_data in enumerate(validated_rows):
            excel_row = row_data.get('_excel_row', idx + 2)
            
            try:
                customer_data = {}
                
                # Process each field
                for excel_col, mapping in field_map.items():
                    if excel_col in ['ledger_account', 'ledger_group']:
                        continue  # Handle separately
                    
                    value = row_data.get(excel_col)
                    if value is None or value == '':
                        continue
                    
                    if isinstance(mapping, tuple):
                        # FK field - use helper
                        field_name = mapping[0]
                        fk_obj = cls.get_fk_object(fk_lookups, excel_col, value)
                        if fk_obj:
                            customer_data[field_name] = fk_obj
                    else:
                        # Regular field
                        if mapping in boolean_fields:
                            customer_data[mapping] = cls.parse_boolean(value)
                        elif mapping == 'tax_type':
                            if value in ['Inclusive', 'Exclusive', 'Both']:
                                customer_data[mapping] = value
                        else:
                            customer_data[mapping] = value
                
                # Handle ledger account
                ledger_obj = cls.get_fk_object(fk_lookups, 'ledger_account', row_data.get('ledger_account'))
                if ledger_obj:
                    customer_data['ledger_account_id'] = ledger_obj
                
                customer = Customer(**customer_data)
                customer._excel_row = excel_row
                customer._row_data = row_data  # Store for address creation
                customers_to_create.append(customer)
                
            except Exception as e:
                errors.append({"row": excel_row, "error": f"Failed to prepare customer: {str(e)}"})
        
        logger.info(f"PHASE 4 complete: {len(customers_to_create)} customers prepared")
        
        # ========================================
        # PHASE 5: Bulk create customers
        # ========================================
        logger.info("PHASE 5: Bulk creating customers...")
        
        success_count = 0
        created_customers = []
        
        try:
            with transaction.atomic():
                for batch_start in range(0, len(customers_to_create), BATCH_SIZE):
                    batch = customers_to_create[batch_start:batch_start + BATCH_SIZE]
                    Customer.objects.bulk_create(batch, batch_size=BATCH_SIZE)
                    created_customers.extend(batch)
                    
                    batch_num = (batch_start // BATCH_SIZE) + 1
                    total_batches = (len(customers_to_create) + BATCH_SIZE - 1) // BATCH_SIZE
                    logger.info(f"Batch {batch_num}/{total_batches} created ({len(batch)} customers)")
                
                success_count = len(created_customers)
                
        except Exception as e:
            logger.error(f"Bulk create failed: {str(e)}")
            for customer in customers_to_create:
                errors.append({"row": getattr(customer, '_excel_row', 0), "error": f"Bulk create failed: {str(e)}"})
            return {"success": 0, "errors": errors, "total": total_rows}
        
        logger.info(f"PHASE 5 complete: {success_count} customers created")
        
        # ========================================
        # PHASE 6: Bulk create addresses
        # ========================================
        logger.info("PHASE 6: Creating addresses...")
        
        billing_addresses = []
        shipping_addresses = []
        
        for customer in created_customers:
            row_data = getattr(customer, '_row_data', {})
            
            # Billing address
            if any(row_data.get(f) for f in ['billing_address', 'billing_country', 'billing_state', 'billing_city']):
                billing_addresses.append(CustomerAddresses(
                    customer_id=customer,
                    address_type="Billing",
                    address=row_data.get("billing_address"),
                    city_id=cls.get_fk_object(fk_lookups, 'city', row_data.get('billing_city')),
                    state_id=cls.get_fk_object(fk_lookups, 'state', row_data.get('billing_state')),
                    country_id=cls.get_fk_object(fk_lookups, 'country', row_data.get('billing_country')),
                    pin_code=row_data.get("billing_pin_code"),
                    phone=row_data.get("billing_phone"),
                    email=row_data.get("billing_email"),
                    longitude=row_data.get("billing_longitude"),
                    latitude=row_data.get("billing_latitude")
                ))
            
            # Shipping address
            if any(row_data.get(f) for f in ['shipping_address', 'shipping_country', 'shipping_state', 'shipping_city']):
                shipping_addresses.append(CustomerAddresses(
                    customer_id=customer,
                    address_type="Shipping",
                    address=row_data.get("shipping_address"),
                    city_id=cls.get_fk_object(fk_lookups, 'city', row_data.get('shipping_city')),
                    state_id=cls.get_fk_object(fk_lookups, 'state', row_data.get('shipping_state')),
                    country_id=cls.get_fk_object(fk_lookups, 'country', row_data.get('shipping_country')),
                    pin_code=row_data.get("shipping_pin_code"),
                    phone=row_data.get("shipping_phone"),
                    email=row_data.get("shipping_email"),
                    longitude=row_data.get("shipping_longitude"),
                    latitude=row_data.get("shipping_latitude")
                ))
        
        if billing_addresses:
            CustomerAddresses.objects.bulk_create(billing_addresses, batch_size=BATCH_SIZE)
            logger.info(f"Created {len(billing_addresses)} billing addresses")
        
        if shipping_addresses:
            CustomerAddresses.objects.bulk_create(shipping_addresses, batch_size=BATCH_SIZE)
            logger.info(f"Created {len(shipping_addresses)} shipping addresses")
        
        logger.info(f"PHASE 6 complete")
        
        return {"success": success_count, "errors": errors, "total": total_rows}

    @classmethod
    def generate_template(cls, extra_columns=None):
        """Generate Excel template with address fields"""
        wb = super().generate_template(extra_columns)
        ws = wb.active

        # Add data validation for GST classification type
        if 'gst_classification' in cls.FIELD_MAP:
            gst_col = list(cls.FIELD_MAP.keys()).index('gst_classification') + 1
            # Add dropdown validation with allowed values from the model
            dv = DataValidation(type="list", formula1='"HSN,SAC,Both"', allow_blank=True)
            dv.add(f"{get_column_letter(gst_col)}2:{get_column_letter(gst_col)}1000")
            ws.add_data_validation(dv)
            
            # Add a comment explaining valid values
            gst_cell = ws.cell(row=1, column=gst_col)
            comment = Comment('Valid values: HSN, SAC, Both', 'System')
            gst_cell.comment = comment
        
        # Add hints for constrained fields like tax_type
        if 'tax_type' in cls.FIELD_MAP:
            tax_type_col = list(cls.FIELD_MAP.keys()).index('tax_type') + 1
            # Add data validation for tax_type
            dv = DataValidation(type="list", formula1='"Inclusive,Exclusive,Both"', allow_blank=True)
            dv.add(f"{get_column_letter(tax_type_col)}2:{get_column_letter(tax_type_col)}1000")
            ws.add_data_validation(dv)
            
            # Add a comment explaining valid values
            tax_type_cell = ws.cell(row=1, column=tax_type_col)
            comment = Comment('Valid values: Inclusive, Exclusive, Both', 'System')
            tax_type_cell.comment = comment
        
        # Add address fields with same styling as optional fields
        address_headers = []
        
        # Billing address fields
        billing_fields = [
            "billing_address", "billing_country", "billing_state", 
            "billing_city", "billing_pin_code", "billing_phone", 
            "billing_email", "billing_longitude", "billing_latitude"
        ]
        address_headers.extend(billing_fields)
        
        # Shipping address fields (no shipping_same_as_billing field)
        shipping_fields = [
            "shipping_address", "shipping_country", "shipping_state", 
            "shipping_city", "shipping_pin_code", "shipping_phone", 
            "shipping_email", "shipping_longitude", "shipping_latitude"
        ]
        address_headers.extend(shipping_fields)
        
        # Style for address fields - Light blue (same as optional fields)
        address_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        address_font = Font(bold=True, color="000000")
        
        # Get the last row and append the address headers
        last_row = ws.max_row
        for col_num, header in enumerate(address_headers, 1 + len(list(cls.FIELD_MAP.keys()))):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = address_fill
            cell.font = address_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = len(str(header)) + 4
            
        return wb
    
    @classmethod
    def get_or_create_country(cls, name):
        """Get or create a country"""
        if not name:
            return None
            
        country = Country.objects.filter(country_name__iexact=name).first()
        if not country:
            logger.info(f"Creating Country with name='{name}'")
            country = Country.objects.create(
                country_name=name,
                country_code=name[:3].upper()
            )
        return country
        
    @classmethod
    def get_or_create_state(cls, name, country=None):
        """Get or create a state - country is optional for lookup but required for creation"""
        if not name:
            return None
        
        # First try to find existing state by name
        state = State.objects.filter(state_name__iexact=name).first()
        if state:
            # Update country if provided and state doesn't have one
            if country and not state.country_id:
                state.country_id = country
                state.save()
            return state
        
        # Need to create new state - requires country due to DB constraint
        if not country:
            logger.warning(f"Cannot create State '{name}' without country - state will be None")
            return None
            
        logger.info(f"Creating State with state_name='{name}'")
        state = State.objects.create(
            state_name=name,
            state_code=name[:3].upper(),
            country_id=country
        )
        return state
        
    @classmethod
    def get_or_create_city(cls, name, state=None, country=None):
        """
        Get or create a city DYNAMICALLY.
        
        Priority:
        1. If state provided → find city in THAT specific state (or create new)
        2. If no state but country provided → create default state, then city
        3. If nothing provided → use India as default country
        
        """
        if not name:
            return None
        
        name = str(name).strip()
        
        # CASE 1: State IS provided - look for city in THAT specific state
        if state:
            city = City.objects.filter(city_name__iexact=name, state_id=state).first()
            if city:
                logger.info(f"Found existing city '{city.city_name}' in state '{state.state_name}'")
                return city
            # City doesn't exist in this state - create it
            logger.info(f"Creating City '{name}' in state '{state.state_name}'")
            city = City.objects.create(
                city_name=name,
                city_code=name[:3].upper(),
                state_id=state
            )
            return city
        
        # CASE 2: No state provided - try to find any existing city with this name
        city = City.objects.filter(city_name__iexact=name).first()
        if city:
            logger.info(f"Found existing city '{city.city_name}' in state '{city.state_id.state_name if city.state_id else 'None'}'")
            return city
        
        # CASE 3: City doesn't exist and no state - need to create with default
        # Use India as default country if nothing provided (for Indian ERP context)
        if not country:
            country = Country.objects.filter(country_name__iexact='India').first()
            if not country:
                country = Country.objects.create(
                    country_name='India',
                    country_code='IND'
                )
                logger.info("Created default country 'India'")
        
        # Create default state using country name
        default_state_name = country.country_name
        state = State.objects.filter(state_name__iexact=default_state_name, country_id=country).first()
        if not state:
            state = State.objects.create(
                state_name=default_state_name,
                state_code=default_state_name[:3].upper(),
                country_id=country
            )
            logger.info(f"Created default state '{default_state_name}' for country '{country.country_name}'")
            
        logger.info(f"Creating City '{name}' in default state '{state.state_name}'")
        city = City.objects.create(
            city_name=name,
            city_code=name[:3].upper(),
            state_id=state
        )
        return city


    # ============================================================
    # UPDATE RECORD — Used by Excel re-import with mode=update
    # ============================================================
    @classmethod
    def update_record(cls, row_data, field_map=None, boolean_fields=None, get_or_create_funcs=None):
        """
        Update an existing customer from Excel row data.
        Matches by customer_id column. Only non-empty fields are updated.
        """
        field_map = field_map or cls.FIELD_MAP
        boolean_fields = boolean_fields or cls.BOOLEAN_FIELDS

        customer_id = row_data.get('customer_id')
        if not customer_id:
            raise ValueError("customer_id is required for update mode")

        try:
            customer = Customer.objects.get(pk=customer_id, is_deleted=False)
        except Customer.DoesNotExist:
            raise ValueError(f"Customer with ID {customer_id} not found or is deleted")

        with transaction.atomic():
            # --- Update customer fields ---
            updated_fields = []
            for excel_col, mapping in field_map.items():
                if excel_col in ['ledger_account', 'ledger_group']:
                    continue  # Handle separately below

                value = row_data.get(excel_col)
                if value is None or value == '':
                    continue  # Skip empty = no change

                if isinstance(mapping, tuple):
                    field_name, model = mapping
                    fk_obj = cls.get_or_create_fk(model, value)
                    if fk_obj:
                        setattr(customer, field_name, fk_obj)
                        updated_fields.append(field_name)
                else:
                    if mapping in boolean_fields:
                        setattr(customer, mapping, cls.parse_boolean(value))
                    elif mapping == 'tax_type':
                        if value in ['Inclusive', 'Exclusive', 'Both']:
                            setattr(customer, mapping, value)
                    else:
                        setattr(customer, mapping, value)
                    updated_fields.append(mapping)

            # Handle ledger account with group
            ledger_name = row_data.get('ledger_account')
            if ledger_name:
                group_name = row_data.get('ledger_group')
                if group_name:
                    group = LedgerGroups.objects.filter(name__iexact=group_name).first()
                    if not group:
                        group = LedgerGroups.objects.create(name=group_name)
                else:
                    group = LedgerGroups.objects.first()

                ledger = LedgerAccounts.objects.filter(name__iexact=ledger_name).first()
                if not ledger and group:
                    ledger = LedgerAccounts.objects.create(name=ledger_name, ledger_group_id=group)
                if ledger:
                    customer.ledger_account_id = ledger
                    updated_fields.append('ledger_account_id')

            if updated_fields:
                customer.save()
                logger.info(f"Updated customer {customer.name}: {', '.join(updated_fields)}")

            # --- Update addresses ---
            cls._update_address(customer, row_data, 'Billing', 'billing')
            cls._update_address(customer, row_data, 'Shipping', 'shipping')

        return customer

    @classmethod
    def _update_address(cls, customer, row_data, address_type, prefix):
        """Update or create a billing/shipping address from row data."""
        address_fields = [f'{prefix}_address', f'{prefix}_phone', f'{prefix}_email',
                          f'{prefix}_city', f'{prefix}_state', f'{prefix}_country',
                          f'{prefix}_pin_code', f'{prefix}_longitude', f'{prefix}_latitude']

        if not any(row_data.get(f) for f in address_fields):
            return  # No address data provided

        address = CustomerAddresses.objects.filter(
            customer_id=customer, address_type=address_type
        ).first()

        country = cls.get_or_create_country(row_data.get(f'{prefix}_country'))
        state = cls.get_or_create_state(row_data.get(f'{prefix}_state'), country)
        city = cls.get_or_create_city(row_data.get(f'{prefix}_city'), state, country)

        addr_data = {
            'address': row_data.get(f'{prefix}_address') or (address.address if address else None),
            'city_id': city or (address.city_id if address else None),
            'state_id': state or (address.state_id if address else None),
            'country_id': country or (address.country_id if address else None),
            'pin_code': row_data.get(f'{prefix}_pin_code') or (address.pin_code if address else None),
            'phone': row_data.get(f'{prefix}_phone') or (address.phone if address else None),
            'email': row_data.get(f'{prefix}_email') or (address.email if address else None),
            'longitude': row_data.get(f'{prefix}_longitude') or (address.longitude if address else None),
            'latitude': row_data.get(f'{prefix}_latitude') or (address.latitude if address else None),
        }

        if address:
            for key, val in addr_data.items():
                setattr(address, key, val)
            address.save()
        else:
            CustomerAddresses.objects.create(
                customer_id=customer, address_type=address_type, **addr_data
            )

    # ============================================================
    # EXPORT EXISTING CUSTOMERS TO EXCEL
    # ============================================================
    @classmethod
    def export_customers(cls, queryset=None):
        """
        Export existing customers to Excel (with customer_id for re-import update).
        Returns an openpyxl Workbook.
        """
        wb = cls.generate_template()
        ws = wb.active

        # Prepend customer_id as the first column and add address columns
        # Rebuild headers: customer_id + FIELD_MAP keys + address columns
        field_keys = list(cls.FIELD_MAP.keys())
        address_cols = [
            'billing_address', 'billing_country', 'billing_state', 'billing_city',
            'billing_pin_code', 'billing_phone', 'billing_email', 'billing_longitude', 'billing_latitude',
            'shipping_address', 'shipping_country', 'shipping_state', 'shipping_city',
            'shipping_pin_code', 'shipping_phone', 'shipping_email', 'shipping_longitude', 'shipping_latitude',
        ]
        all_headers = ['customer_id'] + field_keys + address_cols

        # Clear the sheet and rewrite headers
        ws.delete_rows(1, ws.max_row)
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        id_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for ID
        font_bold = Font(bold=True)
        align_center = Alignment(horizontal="center")

        for col_num, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = id_fill if header == 'customer_id' else header_fill
            cell.font = font_bold
            cell.alignment = align_center
            ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 4, 15)

        # Add comment to customer_id column
        ws.cell(row=1, column=1).comment = Comment(
            '⚠️ DO NOT MODIFY this column. Used to match records for update.', 'System'
        )

        # FK field → attribute name mapping for readable export
        FK_DISPLAY = {
            'ledger_account': ('ledger_account_id', 'name'),
            'firm_status': ('firm_status_id', 'name'),
            'territory': ('territory_id', 'name'),
            'customer_category': ('customer_category_id', 'name'),
            'gst_category': ('gst_category_id', 'name'),
            'payment_term': ('payment_term_id', 'name'),
            'price_category': ('price_category_id', 'name'),
            'transporter': ('transporter_id', 'name'),
        }

        if queryset is None:
            queryset = Customer.objects.filter(is_deleted=False).order_by('name')

        # Prefetch addresses for performance
        from django.db.models import Prefetch
        queryset = queryset.select_related(
            'ledger_account_id', 'ledger_account_id__ledger_group_id',
            'firm_status_id', 'territory_id',
            'customer_category_id', 'gst_category_id', 'payment_term_id',
            'price_category_id', 'transporter_id'
        ).prefetch_related(
            Prefetch(
                'customeraddresses_set',
                queryset=CustomerAddresses.objects.select_related('city_id', 'state_id', 'country_id'),
                to_attr='_prefetched_addresses'
            )
        )

        for customer in queryset:
            row = []
            # 1. customer_id
            row.append(str(customer.customer_id))

            # 2. FIELD_MAP columns
            for excel_col, mapping in cls.FIELD_MAP.items():
                if excel_col in FK_DISPLAY:
                    fk_attr, display_field = FK_DISPLAY[excel_col]
                    fk_obj = getattr(customer, fk_attr, None)
                    row.append(getattr(fk_obj, display_field, '') if fk_obj else '')
                elif excel_col == 'ledger_group':
                    try:
                        ledger = getattr(customer, 'ledger_account_id', None)
                        group = getattr(ledger, 'ledger_group_id', None) if ledger else None
                        row.append(getattr(group, 'name', '') if group else '')
                    except Exception:
                        row.append('')
                elif isinstance(mapping, tuple):
                    row.append('')  # Unknown FK — leave blank
                else:
                    val = getattr(customer, mapping, '')
                    if isinstance(val, bool):
                        val = 'Yes' if val else 'No'
                    row.append(val if val is not None else '')

            # 3. Address columns (from prefetched data — no extra queries)
            prefetched = getattr(customer, '_prefetched_addresses', [])
            billing = next((a for a in prefetched if a.address_type == 'Billing'), None)
            shipping = next((a for a in prefetched if a.address_type == 'Shipping'), None)

            for addr in [billing, shipping]:
                if addr:
                    row.extend([
                        addr.address or '', 
                        addr.country_id.country_name if addr.country_id else '',
                        addr.state_id.state_name if addr.state_id else '',
                        addr.city_id.city_name if addr.city_id else '',
                        addr.pin_code or '', addr.phone or '', addr.email or '',
                        str(addr.longitude) if addr.longitude else '',
                        str(addr.latitude) if addr.latitude else '',
                    ])
                else:
                    row.extend([''] * 9)

            ws.append(row)

        # Freeze header + ID column
        ws.freeze_panes = 'B2'
        return wb


# API Views for Customer Excel import/export
class CustomerTemplateAPIView(APIView):
    """
    API for downloading the customer import template.
    """
    
    def get(self, request, *args, **kwargs):
        return CustomerExcelImport.get_template_response(request)

class CustomerExportAPIView(APIView):
    """
    API for exporting existing customers to Excel (for re-import update).
    GET /export-customers/              → export all customers
    GET /export-customers/?ids=a,b,c    → export specific customers
    """
    def get(self, request, *args, **kwargs):
        try:
            ids_param = request.query_params.get('ids', '')
            queryset = Customer.objects.filter(is_deleted=False).order_by('name')

            if ids_param:
                id_list = [i.strip() for i in ids_param.split(',') if i.strip()]
                queryset = queryset.filter(customer_id__in=id_list)
                if not queryset.exists():
                    return build_response(0, "No matching customers found", [], status.HTTP_404_NOT_FOUND)

            wb = CustomerExcelImport.export_customers(queryset)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Customer_Export.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"Error exporting customers: {str(e)}")
            return build_response(0, f"Export failed: {str(e)}", [], status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomerExcelUploadAPIView(APIView):
    """
    API for importing customers from Excel files.
    
    Supports two modes (via query param ?mode=create or ?mode=update):
    - create (default): Creates new customers from Excel rows
    - update: Matches by customer_id column and updates existing records
    
    OPTIMIZED FOR LARGE IMPORTS:
    - Uses bulk_create for imports > 500 rows (5-10x faster)
    - Uses traditional save() loop for smaller imports (safer, signals fire)
    """
    parser_classes = (MultiPartParser, FormParser)
    
    # Threshold for switching to bulk mode
    BULK_THRESHOLD = 500
    
    def post(self, request, *args, **kwargs):
        import time
        start_time = time.time()
        import_mode_param = request.query_params.get('mode', 'create').lower()
        
        # --- AUTO-DETECT UPDATE MODE ---
        # If the uploaded Excel has a 'customer_id' column, it's an exported
        # file meant for re-import as updates — switch to update mode automatically.
        if import_mode_param == 'create':
            file = request.FILES.get('file')
            if file:
                peek_wb = None
                try:
                    peek_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                    peek_sheet = peek_wb.active
                    headers = [str(cell.value).strip().lower() for cell in next(peek_sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
                    if 'customer_id' in headers:
                        import_mode_param = 'update'
                        logger.info("Auto-detected update mode (customer_id column found in Excel)")
                except Exception:
                    pass
                finally:
                    if peek_wb is not None:
                        peek_wb.close()
                    file.seek(0)  # reset for downstream processing

        # --- UPDATE MODE ---
        if import_mode_param == 'update':
            return self._handle_update(request, start_time)
        
        # --- CREATE MODE (existing logic, untouched) ---
        try:
            # Upload and validate file
            file_path, status_code = CustomerExcelImport.upload_file(request)
            
            # If there was an error with the file
            if status_code != status.HTTP_200_OK:
                return build_response(0, file_path.get("error", "Unknown error"), [], status_code)
            
            # Get the file and count rows
            file = request.FILES.get('file')
            
            # Quick row count check
            temp_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            temp_sheet = temp_wb.active
            row_count = sum(1 for row in temp_sheet.iter_rows(min_row=2, values_only=True) 
                          if any(cell is not None for cell in row))
            temp_wb.close()
            
            # Reset file pointer
            file.seek(0)
            
            logger.info(f"Customer import: {row_count} rows detected")
            
            # Choose method based on row count
            if row_count >= self.BULK_THRESHOLD:
                # BULK MODE
                logger.info(f"Using BULK import mode for {row_count} customers")
                result, status_code = CustomerExcelImport.process_excel_file_bulk(
                    file,
                    CustomerExcelImport.bulk_create_records
                )
                import_mode = "bulk"
            else:
                # TRADITIONAL MODE
                logger.info(f"Using TRADITIONAL import mode for {row_count} customers")
                result, status_code = CustomerExcelImport.process_excel_file(
                    file,
                    CustomerExcelImport.create_record
                )
                import_mode = "traditional"
            
            # Calculate elapsed time
            elapsed_time = round(time.time() - start_time, 2)
            
            if status_code != status.HTTP_200_OK:
                error_msg = result.get("error", "Import failed")
                error_details = {}
                
                if "missing_columns" in result:
                    error_details["missing_columns"] = result["missing_columns"]
                if "unexpected_columns" in result:
                    error_details["unexpected_columns"] = result["unexpected_columns"]
                if "missing_expected_columns" in result:
                    error_details["missing_expected_columns"] = result["missing_expected_columns"]
                if "missing_data_rows" in result:
                    error_details["missing_data_rows"] = result["missing_data_rows"]
                    
                return build_response(0, error_msg, error_details, status_code)
            
            success_count = result.get("success", 0)
            errors = result.get("errors", [])
            total_count = result.get("total", success_count + len(errors))
            
            if success_count == 0 and errors:
                first_error = errors[0]["error"] if errors else "Unknown error during import"
                return build_response(
                    0,
                    f"Import failed: {first_error}",
                    {"errors": errors[:20], "total_errors": len(errors), "elapsed_time": f"{elapsed_time}s", "import_mode": import_mode},
                    status.HTTP_400_BAD_REQUEST
                )
            elif errors:
                return build_response(
                    success_count,
                    f"Partial import: {success_count} of {total_count} customers imported in {elapsed_time}s. {len(errors)} failed.",
                    {"success_count": success_count, "failed_count": len(errors), "errors": errors[:20], "elapsed_time": f"{elapsed_time}s", "import_mode": import_mode},
                    status.HTTP_200_OK
                )
            else:
                return build_response(
                    success_count,
                    f"{success_count} customers imported successfully in {elapsed_time}s ({import_mode} mode).",
                    {"success_count": success_count, "total_count": total_count, "elapsed_time": f"{elapsed_time}s", "import_mode": import_mode},
                    status.HTTP_200_OK
                )
            
        except Exception as e:
            logger.error(f"Error in customer Excel import: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return build_response(0, f"Import failed: {str(e)}", [], status.HTTP_400_BAD_REQUEST)

    def _handle_update(self, request, start_time):
        """
        Handle Excel-based bulk update.
        Reads uploaded Excel, matches by customer_id, updates only non-empty fields.
        """
        import time
        try:
            file = request.FILES.get('file')
            if not file:
                return build_response(0, "No file uploaded", [], status.HTTP_400_BAD_REQUEST)

            file_name = file.name.lower()
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                return build_response(0, "Invalid file format. Only .xlsx or .xls supported.", [], status.HTTP_400_BAD_REQUEST)

            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            try:
                sheet = wb.active

                raw_headers = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[1]]
                headers = [h.replace(' *', '').strip() for h in raw_headers]

                if 'customer_id' not in headers:
                    return build_response(0, "Excel must have a 'customer_id' column for update mode. Use the Export file as your template.", [], status.HTTP_400_BAD_REQUEST)

                success = 0
                failed = []
                total = 0

                for excel_row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_values = list(row)
                    if not any(cell is not None for cell in row_values):
                        continue  # Skip empty rows

                    total += 1
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row_values):
                            val = row_values[i]
                            if isinstance(val, str):
                                val = val.strip() if val.strip() else None
                            row_data[header] = val

                    if not row_data.get('customer_id'):
                        failed.append({"row": excel_row_idx, "error": "Missing customer_id"})
                        continue

                    try:
                        CustomerExcelImport.update_record(row_data)
                        success += 1
                    except Exception as e:
                        failed.append({"row": excel_row_idx, "error": str(e)})
            finally:
                wb.close()

            elapsed = round(time.time() - start_time, 2)

            if success == 0 and failed:
                return build_response(0, f"Update failed. {len(failed)} errors.",
                    {"errors": failed[:20], "elapsed_time": f"{elapsed}s"}, status.HTTP_400_BAD_REQUEST)
            elif failed:
                return build_response(success,
                    f"Partial update: {success}/{total} updated in {elapsed}s. {len(failed)} failed.",
                    {"success_count": success, "failed_count": len(failed), "errors": failed[:20], "elapsed_time": f"{elapsed}s"},
                    status.HTTP_200_OK)
            else:
                return build_response(success,
                    f"{success} customers updated successfully in {elapsed}s.",
                    {"success_count": success, "total_count": total, "elapsed_time": f"{elapsed}s"},
                    status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error in customer Excel update: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return build_response(0, f"Update failed: {str(e)}", [], status.HTTP_400_BAD_REQUEST)