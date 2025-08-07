import decimal
import logging
import time
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter
from rest_framework import viewsets
from rest_framework import status
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404
from django.http import Http404
from django.db import transaction
from apps.masters.models import City, Country, State
from config.utils_filter_methods import filter_response, list_filtered_objects
from config.utils_variables import *
from config.utils_methods import *
from apps.inventory.serializers import WarehouseLocationsSerializer
from apps.inventory.models import WarehouseLocations
from .serializers import *
from .models import *
from .filters import ColorFilter, ProductGroupsFilter, ProductCategoriesFilter, ProductStockUnitsFilter, ProductGstClassificationsFilter, ProductSalesGlFilter, ProductPurchaseGlFilter, ProductsFilter, ProductItemBalanceFilter, ProductVariationFilter, SizeFilter

# Set up basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Create a logger object
logger = logging.getLogger(__name__)

# Create your views here.
class ProductGroupsViewSet(viewsets.ModelViewSet):
    queryset = ProductGroups.objects.all().order_by('-created_at')	
    serializer_class = ProductGroupsSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductGroupsFilter
    ordering_fields = ['group_name','created_at']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductGroups,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)

class ProductCategoriesViewSet(viewsets.ModelViewSet):
    queryset = ProductCategories.objects.all().order_by('-created_at')	
    serializer_class = ProductCategoriesSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductCategoriesFilter
    ordering_fields = ['category_name','code','created_at']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductCategories,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)

class ProductStockUnitsViewSet(viewsets.ModelViewSet):
    queryset = ProductStockUnits.objects.all().order_by('-created_at')	
    serializer_class = ProductStockUnitsSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductStockUnitsFilter
    ordering_fields = ['stock_unit_name','quantity_code_id']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductStockUnits,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)
	
class ProductGstClassificationsViewSet(viewsets.ModelViewSet):
    queryset = ProductGstClassifications.objects.all().order_by('-created_at')	
    serializer_class = ProductGstClassificationsSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductGstClassificationsFilter
    ordering_fields = ['type','code','hsn_or_sac_code','created_at']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductGstClassifications,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)

class ProductSalesGlViewSet(viewsets.ModelViewSet):
    queryset = ProductSalesGl.objects.all().order_by('-created_at')	
    serializer_class = ProductSalesGlSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductSalesGlFilter
    ordering_fields = ['name','sales_accounts','code','type','account_no','rtgs_ifsc_code','address','pan','employee']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductSalesGl,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)
	
class ProductPurchaseGlViewSet(viewsets.ModelViewSet):
    queryset = ProductPurchaseGl.objects.all().order_by('-created_at')	
    serializer_class = ProductPurchaseGlSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductPurchaseGlFilter
    ordering_fields = ['name','purchase_accounts','code','type','account_no','rtgs_ifsc_code','address','pan','employee']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductPurchaseGl,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)

class productsViewSet(viewsets.ModelViewSet):
    queryset = Products.objects.all().order_by('-created_at')
    serializer_class = productsSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductsFilter
    ordering_fields = ['name','code','barcode','category_id','product_group_id','type_id','gst_classification_id','created_at']

    def list(self, request, *args, **kwargs):
        summary = request.query_params.get('summary', 'false').lower() == 'true'
        if summary:
            products = self.filter_queryset(self.get_queryset())
            data = ProductOptionsSerializer.get_product_summary(products)
            result = Response(data, status=status.HTTP_200_OK)
        else:
            result = list_all_objects(self, request, *args, **kwargs)
        
        return result
    
    def create(self, request, *args, **kwargs):
        # Extract the product ID from the request if needed
        request_product_id = request.data.get('product_id')

        # Check if 'picture' exists in request data and is a list
        if 'picture' in request.data and isinstance(request.data['picture'], list):
            attachment_data_list = request.data['picture']
            if attachment_data_list:
                # Ensure all items in the list have the required fields
                for attachment in attachment_data_list:
                    if not all(key in attachment for key in ['uid', 'name', 'attachment_name', 'file_size', 'attachment_path']):
                        return build_response(0, "Missing required fields in some picture data.", [], status.HTTP_400_BAD_REQUEST)
                
                # Set the picture field in request data as a list of objects
                request.data['picture'] = attachment_data_list

            else:
                # Handle case where 'picture' list is empty
                return build_response(0, "'picture' list is empty.", [], status.HTTP_400_BAD_REQUEST)

        # Proceed with creating the instance
        try:
            response = super().create(request, *args, **kwargs)
            
            # Format the response to include the picture data
            if isinstance(response.data, dict):
                picture_data = response.data.get('picture')
                if picture_data:
                    response.data['picture'] = picture_data
            return response
        
        except ValidationError as e:
            return build_response(1, "Creation failed due to validation errors.", e.detail, status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)
  
class ProductItemBalanceViewSet(viewsets.ModelViewSet):
    queryset = ProductItemBalance.objects.all().order_by('-created_at')	
    serializer_class = ProductItemBalanceSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ProductItemBalanceFilter
    ordering_fields = ['created_at']

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, ProductItemBalance,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)

class SizeViewSet(viewsets.ModelViewSet):
    queryset = Size.objects.all().order_by('-created_at')	
    serializer_class = SizeSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = SizeFilter
    ordering_fields = []

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, Size,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)  
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)  

class ColorViewSet(viewsets.ModelViewSet):
    queryset = Color.objects.all().order_by('-created_at')
    serializer_class = ColorSerializer
    filter_backends = [DjangoFilterBackend,OrderingFilter]
    filterset_class = ColorFilter
    ordering_fields = []

    def list(self, request, *args, **kwargs):
        return list_filtered_objects(self, request, Color,*args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)  
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        return soft_delete(instance)  

class ProductVariationViewSet(viewsets.ModelViewSet):
    queryset = ProductVariation.objects.all()
    serializer_class = ProductVariationSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = ProductVariationFilter
    ordering_fields = []

    def list(self, request, *args, **kwargs):
        return list_all_objects(self, request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        return create_instance(self, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return update_instance(self, request, *args, **kwargs)

#--------------------P R O D U C T - A P I-----------------------#

class ProductViewSet(APIView):
    """
    API ViewSet for handling Lead creation and related data.
    """
    def validate_sku(self, data):
        seen = set()
        for variation in data:
            sku = variation.get('sku')  # Get SKU, ignoring missing keys 
            if sku in seen:
                return build_response(0, f"Duplicate SKU found: {sku}", [], status.HTTP_400_BAD_REQUEST)
            seen.add(sku)
        return None  # No duplicates found, continue execution

    def get_object(self, pk):
        try:
            return Products.objects.get(pk=pk)
        except Products.DoesNotExist:
            logger.warning(f"Products with ID {pk} does not exist.")
            return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)

    def get(self, request, *args, **kwargs):
        if "pk" in kwargs:
            result = validate_input_pk(self, kwargs['pk'])
            return result if result else self.retrieve(self, request, *args, **kwargs)
        try:
            summary = request.query_params.get('summary', 'false').lower() == 'true' + '&' 
            if summary:
                product = Products.objects.all().order_by('-created_at')	
                data = ProductOptionsSerializer.get_product_summary(product)
                return build_response(len(data), "Success", data, status.HTTP_200_OK)
            else:
                logger.info("Retrieving all products")
                queryset = Products.objects.all().order_by('-created_at')	

                page = int(request.query_params.get('page', 1))  # Default to page 1 if not provided
                limit = int(request.query_params.get('limit', 10)) 
                total_count = Products.objects.count()

                # Apply filters manuallys
                if request.query_params:
                    filterset = ProductsFilter(request.GET, queryset=queryset)
                    if filterset.is_valid():
                        queryset = filterset.qs 

                serializer = ProductOptionsSerializer(queryset, many=True)
                logger.info("product data retrieved successfully.")
                # return build_response(queryset.count(), "Success", serializer.data, status.HTTP_200_OK)
                return filter_response(queryset.count(),"Success",serializer.data,page,limit,total_count,status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"An unexpected error occurred: {str(e)}")
            return build_response(0, "An error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
            """
            Retrieves a products and its related data (ProductVariation)
            """
            try:
                pk = kwargs.get('pk')
                if not pk:
                    logger.error("Primary key not provided in request.")
                    return build_response(0, "Primary key not provided", [], status.HTTP_400_BAD_REQUEST)

                # Retrieve the Products instance
                product = get_object_or_404(Products, pk=pk)
                products_serializer = productsSerializer(product)

                # Retrieve 'product_variations'
                product_variations_data = get_related_data(ProductVariation, ProductVariationSerializer, 'product_id', pk)

                # Retrieve 'product_item_balance'
                product_balance_data = get_related_data(ProductItemBalance, ProductItemBalanceSerializer, 'product_id', pk)

                # Customizing the response data
                custom_data = {
                    "products": products_serializer.data,
                    "product_variations":product_variations_data if product_variations_data else [],
                    "product_item_balance":product_balance_data if product_balance_data else [],
                    }
                logger.info("product and related data retrieved successfully.")
                return build_response(1, "Success", custom_data, status.HTTP_200_OK)

            except Http404:
                logger.error("product record with pk %s does not exist.", pk)
                return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
            except Exception as e:
                logger.exception(
                    "An error occurred while retrieving product with pk %s: %s", pk, str(e))
                return build_response(0, "An error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Handling POST requests for creating
    def post(self, request, *args, **kwargs):
        return self.create(request, *args, **kwargs)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        # Extracting data from the request
        given_data = request.data

        # ---------------------- D A T A   V A L I D A T I O N ----------------------------------#
        """
        All the data in request will be validated here. it will handle the following errors:
        - Invalid data types
        - Invalid foreign keys
        - nulls in required fields
        """
        errors = {}

        # Validate products Data
        products_data = given_data.pop('products', None)  # parent_data
        if products_data:
            # Check if 'picture' exists in products_data and is a list
            picture_data = products_data.get('picture', None)
            if picture_data:
                if not isinstance(picture_data, list):
                    return build_response(0, "'picture' field in products_data must be a list.", [], status.HTTP_400_BAD_REQUEST)

                for attachment in picture_data:
                    if not all(key in attachment for key in ['uid', 'name', 'attachment_name', 'file_size', 'attachment_path']):
                        return build_response(0, "Missing required fields in some picture data.", [], status.HTTP_400_BAD_REQUEST)
            product_error = validate_payload_data(self, products_data, productsSerializer)
            if product_error:
                errors["products"] = product_error
        else:
            logger.error("products data is mandatory but not provided.")
            return build_response(0, "products data is mandatory", [], status.HTTP_400_BAD_REQUEST)

        # Validate 'product_variations' data
        product_variations_data = given_data.pop('product_variations', None)
        if product_variations_data:
            sku_error = self.validate_sku(product_variations_data)
            if sku_error is None:
                variations_error = validate_multiple_data(self, product_variations_data, ProductVariationSerializer, ['product_id'])
                if variations_error:
                    errors["product_variations"] = variations_error
            else:
                return  sku_error

        # Validate product_item_balance Data
        product_item_balance_data = given_data.pop('product_item_balance', None)
        if product_item_balance_data:
            product_item_error = validate_multiple_data(self, product_item_balance_data, ProductItemBalanceSerializer, ['product_id'])
            if product_item_error:
                errors["product_item_balance"] = product_item_error

        if errors:
            return build_response(0, "ValidationError :",errors, status.HTTP_400_BAD_REQUEST)
        logger.info("*** (POST) Validation Passed***")

        # ---------------------- D A T A   C R E A T I O N ---------------------------- #
        """
        After the data is validated, this validated data is created as new instances.
        """

        # Hence the data is validated , further it can be created.
        custom_data = {
            'products':{},
            'product_variations':[],
            'product_item_balance':[]
            }

        # Create Products Data
        new_products_data = generic_data_creation(self, [products_data], productsSerializer)
        new_products_data = new_products_data[0]
        custom_data["products"] = new_products_data
        product_id = new_products_data.get("product_id", None)  # Fetch product_id from mew instance
        logger.info('Products - created*')

        # Create 'product_variations' data
        if product_variations_data:
            variations_data = generic_data_creation(self, product_variations_data, ProductVariationSerializer, {'product_id':product_id})
            custom_data["product_variations"] = variations_data
            logger.info('ProductVariation - created*')

        # Create 'product_item_balance' data
        if product_item_balance_data:
            update_fields = {'product_id':product_id}
            product_balance_data = generic_data_creation(self, product_item_balance_data, ProductItemBalanceSerializer, update_fields)
            custom_data["product_item_balance"] = product_balance_data
            logger.info('ProductItemBalance - created*') 

        return build_response(1, "Record created successfully", custom_data, status.HTTP_201_CREATED)

    def put(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    @transaction.atomic
    def update(self, request, pk, *args, **kwargs):

        # ----------------------------------- D A T A  V A L I D A T I O N -----------------------------#
        """
        All the data in request will be validated here. it will handle the following errors:
        - Invalid data types
        - Invalid foreign keys
        - nulls in required fields
        """
        # Get the given data from request
        given_data = request.data

        errors = {}

        # Validate 'products' Data
        products_data = given_data.pop('products', None)  # parent_data
        if products_data:
            products_data['product_id'] = pk
            products_error = validate_multiple_data(self, products_data, productsSerializer, [])
            if products_error:
                errors['products'] = products_error

        # Validate 'product_variations' data
        product_variations_data = given_data.pop('product_variations', None)
        if product_variations_data:
            mark = self.validate_sku(product_variations_data)

            if mark is None:
                variations_error = validate_put_method_data(self, product_variations_data, ProductVariationSerializer, ['product_id'],ProductVariation,'product_variation_id')
                if variations_error:
                    errors["product_variations"] = variations_error
            else:
                return mark # A duplicate SKU was found, return the error response

        # Validate 'product_item_balance' Data
        product_balance_data = given_data.pop('product_item_balance', None)
        if product_balance_data:
            bal_error = validate_multiple_data(self, product_balance_data, ProductItemBalanceSerializer, [])
            if products_error:
                errors['product_item_balance'] = bal_error

        if errors:
            return build_response(0, "ValidationError :", errors, status.HTTP_400_BAD_REQUEST)

        # ------------------------------ D A T A   U P D A T I O N -----------------------------------------#
        custom_data = {
            "products":{},
            "product_variations":[],
            "product_item_balance":[]
            }

        try:
            # update 'Products'
            product_data = update_multi_instances(self, pk, products_data, Products, productsSerializer, [], main_model_related_field='product_id', current_model_pk_field='product_id')
            product_data = product_data[0] if len(product_data)==1 else product_data
            custom_data['products'] = product_data

            update_fields = {'product_id':pk}

            # update 'product_variations'
            variations_data = update_multi_instances(self, pk, product_variations_data, ProductVariation, ProductVariationSerializer, update_fields, main_model_related_field='product_id', current_model_pk_field='product_variation_id')
            custom_data['product_variations'] = variations_data

            # update 'product_item_balance'
            item_bal_data = update_multi_instances(self, pk, product_balance_data, ProductItemBalance, ProductItemBalanceSerializer, update_fields, main_model_related_field='product_id', current_model_pk_field='product_item_balance_id')
            custom_data['product_item_balance'] = item_bal_data

        except Exception as e:
            logger.error(f'Error: {e}')
            return build_response(0, "An error occurred while updating the data. Please try again.", [], status.HTTP_400_BAD_REQUEST)
        return build_response(1, "Records updated successfully", custom_data, status.HTTP_200_OK)

    @transaction.atomic
    def delete(self, request, pk, *args, **kwargs):
        """
        Handles the deletion of a lead and its related assignments and interactions.
        """
        try:
            # delete the main instance
            instance = Products.objects.get(pk=pk) # Get the Products instance
            instance.is_deleted = True
            instance.save()
            # instance.delete() # Delete the main Products instance
            logger.info(f"Products with ID {pk} deleted successfully.")
            return build_response(1, "Record deleted successfully", [], status.HTTP_204_NO_CONTENT)
        
        except Products.DoesNotExist:
            logger.warning(f"Products with ID {pk} does not exist.")
            return build_response(0, "Record does not exist", [], status.HTTP_404_NOT_FOUND)
        
        except Exception as e:
            logger.error(f"Error deleting Leads with ID {pk}: {str(e)}")
            return build_response(0, f"Record deletion failed due to an error : {e}", [], status.HTTP_500_INTERNAL_SERVER_ERROR)
        
from openpyxl.worksheet.datavalidation import DataValidation



class ProductExcelImport(BaseExcelImportExport):
    """
    Product Excel import/export functionality
    
    This class handles importing product data from Excel files and 
    exporting product templates to Excel.
    """
    MODEL_CLASS = Products
    SERIALIZER_CLASS = productsSerializer  # You need to create this if not exists
    REQUIRED_COLUMNS = ["name", "print_name"]
    TEMPLATE_FILENAME = "Product_Import_Template.xlsx"
    
    FIELD_MAP = {
        # Basic Info
        "name": "name",
        "print_name": "print_name",
        "code": "code",  # This will be auto-generated, but included for completeness
        "barcode": "barcode",
        "packet_barcode": "packet_barcode",
        "status": "status",
        
        # Foreign Keys - Primary Classifications
        "product_group": ("product_group_id", ProductGroups, "group_name"),
        "category": ("category_id", ProductCategories, "category_name"),
        "product_type": ("type_id", ProductTypes, "type_name"),
        "brand": ("brand_id", ProductBrands, "brand_name"),
        
        # Foreign Keys - Units
        "stock_unit": ("stock_unit_id", ProductStockUnits, "stock_unit_name"),
        "pack_unit": ("pack_unit_id", ProductStockUnits, "stock_unit_name"),
        "g_pack_unit": ("g_pack_unit_id", ProductStockUnits, "stock_unit_name"),
        "unit_options": ("unit_options_id", UnitOptions, "unit_name"),
        # For nested relationship, we'll handle it specially in the code
        "quantity_code": ("quantity_code", ProductUniqueQuantityCodes, "quantity_code_name"),
        
        # Product Attributes
        "item_type": ("item_type_id", ProductItemType, "item_name"),
        "drug_type": ("drug_type_id", ProductDrugTypes, "drug_type_name"),
        "gst_classification": ("gst_classification_id", ProductGstClassifications, 'type'),
        
        # GL Accounts
        "sales_gl": ("sales_gl_id", ProductSalesGl, "name"),
        "purchase_gl": ("purchase_gl_id", ProductPurchaseGl, "name"),
        
        # Descriptions
        "sales_description": "sales_description",
        "purchase_description": "purchase_description",
        "salt_composition": "salt_composition",
        
        # Pricing
        "mrp": "mrp",
        "minimum_price": "minimum_price",
        "sales_rate": "sales_rate",
        "wholesale_rate": "wholesale_rate",
        "dealer_rate": "dealer_rate",
        "rate_factor": "rate_factor",
        "discount": "discount",
        "dis_amount": "dis_amount",
        "purchase_rate": "purchase_rate",
        "purchase_rate_factor": "purchase_rate_factor",
        "purchase_discount": "purchase_discount",
        
        # Inventory
        "minimum_level": "minimum_level",
        "maximum_level": "maximum_level",
        "balance": "balance",
        "pack_vs_stock": "pack_vs_stock",
        "g_pack_vs_pack": "g_pack_vs_pack",
        
        # Other attributes
        "hsn_code": "hsn_code",
        "gst_input": "gst_input",
        "print_barcode": "print_barcode",
        "weighscale_mapping_code": "weighscale_mapping_code",
        "purchase_warranty_months": "purchase_warranty_months", 
        "sales_warranty_months": "sales_warranty_months"
    }
    
    BOOLEAN_FIELDS = ["print_barcode"]
    
    # Additional fields for product variations
    VARIATION_HEADERS = [
        "size_category", "size_name", "color_name", "sku", "variation_price", "variation_quantity"
    ]
     # Additional fields for warehouse locations
    WAREHOUSE_HEADERS = ["warehouse_name", "location_name", "location_quantity"]        

    @classmethod
    def create_record(cls, row_data, field_map=None, boolean_fields=None, get_or_create_funcs=None):
        """
        Create a product record with variations from Excel data
        
        This method handles:
        1. Creating the main product record
        2. Processing all foreign key relationships
        3. Creating product variations (if provided)
        4. Setting default values when needed
        """
        field_map = field_map or cls.FIELD_MAP
        boolean_fields = boolean_fields or cls.BOOLEAN_FIELDS
        
        with transaction.atomic():
            # 1. Process the product data
            product_data = {}
            
            # Validate required fields
            for required_field in cls.REQUIRED_COLUMNS:
                if not row_data.get(required_field):
                    raise ValueError(f"Required field '{required_field}' is missing")
            
            # Process foreign key fields and regular fields
            for excel_col, mapping in field_map.items():
                value = row_data.get(excel_col)
                
                # Skip empty values
                if value is None or value == '':
                    continue
                
                try:
                    if isinstance(mapping, tuple):
                        # Handle foreign key fields
                        field_name, model_class, lookup_field = mapping if len(mapping) > 2 else (mapping[0], mapping[1], None)
                        
                        # Determine lookup field if not provided
                        if not lookup_field:
                            if hasattr(model_class, 'name'):
                                lookup_field = 'name'
                            elif hasattr(model_class, model_class.__name__.lower() + '_name'):
                                lookup_field = model_class.__name__.lower() + '_name'
                        
                        # Create lookup filters
                        lookup_kwargs = {f"{lookup_field}__iexact": value}
                        
                        # Handle the special case of quantity_code (nested foreign key)
                        if field_name == "quantity_code":
                            # This is not directly a field in the product, store it for later use
                            # when we process stock_unit, pack_unit, and g_pack_unit fields
                            quantity_code_obj = model_class.objects.filter(**lookup_kwargs).first()
                            if not quantity_code_obj:
                                # Create new quantity code if not found
                                create_kwargs = {lookup_field: value}
                                logger.info(f"Creating {model_class.__name__} with {lookup_field}='{value}'")
                                quantity_code_obj = model_class.objects.create(**create_kwargs)
                            
                            # Store the quantity_code object for later use
                            row_data['_quantity_code_obj'] = quantity_code_obj
                            
                            # We don't add this directly to product_data since it's not a field in Products
                            continue
                            
                        # Try to find existing object
                        obj = model_class.objects.filter(**lookup_kwargs).first()
                        
                        if not obj:
                            # Create new object if not found
                            create_kwargs = {lookup_field: value}
                            logger.info(f"Creating {model_class.__name__} with {lookup_field}='{value}'")
                            
                            # Special handling for code fields
                            if hasattr(model_class, 'code'):
                                create_kwargs['code'] = value[:3].upper()
                            
                            # Special handling for ProductStockUnits when we have a quantity_code
                            if model_class == ProductStockUnits and '_quantity_code_obj' in row_data:
                                create_kwargs['quantity_code_id'] = row_data['_quantity_code_obj']
                                
                            obj = model_class.objects.create(**create_kwargs)
                            
                        # Special handling for updating ProductStockUnits with quantity_code if needed
                        if model_class == ProductStockUnits and '_quantity_code_obj' in row_data and not obj.quantity_code_id:
                            obj.quantity_code_id = row_data['_quantity_code_obj']
                            obj.save()
                            
                        product_data[field_name] = obj
                    else:
                        # Handle regular fields
                        if mapping in boolean_fields:
                            product_data[mapping] = cls.parse_boolean(value)
                        # Inside the method that processes fields in the create_record method
                        elif excel_col == "gst_classification":
                            # Special handling for gst_classification to ensure it matches allowed values
                            valid_types = ['HSN', 'SAC', 'Both']
                            if value and value not in valid_types:
                                raise ValueError(f"Invalid gst_classification '{value}'. Must be one of: {', '.join(valid_types)}")
                            product_data[mapping] = value   
                        else:
                            # Convert decimal fields from string
                            if mapping in ['mrp', 'minimum_price', 'sales_rate', 'wholesale_rate', 
                                        'dealer_rate', 'rate_factor', 'discount', 'dis_amount',
                                        'purchase_rate', 'purchase_rate_factor', 'purchase_discount']:
                                try:
                                    product_data[mapping] = decimal.Decimal(str(value))
                                except (decimal.InvalidOperation, TypeError):
                                    logger.warning(f"Invalid decimal value '{value}' for field '{mapping}', skipping")
                                    continue
                            else:
                                product_data[mapping] = value
                except Exception as e:
                    logger.error(f"Error processing field {excel_col}: {str(e)}")
                    raise ValueError(f"Failed to process field {excel_col}: {str(e)}")
            
            # 2. Create the product record
            try:
                # The 'code' field will be auto-generated in the save() method
                # Always check the database for the highest product number
                # This ensures that after deleting all products, we start from 1 again
                latest_product = Products.objects.filter(code__startswith='PRD-').order_by('-code').first()
                
                if latest_product and latest_product.code:
                    try:
                        last_num = int(latest_product.code.split('-')[1])
                        cls._last_product_num = last_num
                    except (ValueError, IndexError):
                        cls._last_product_num = 0
                else:
                    # No products in database, start from 0
                    cls._last_product_num = 0
                
                # Store the row index for ordered creation timestamps
                if not hasattr(cls, '_row_index'):
                    cls._row_index = 0
                else:
                    cls._row_index += 1
                
                # Increment the counter for each new product
                cls._last_product_num += 1
                
                # Generate the new code with proper formatting
                new_product_code = f"PRD-{cls._last_product_num:05d}"
                logger.info(f"Generating product code: {new_product_code}")
                
                # Remove any code from Excel to use our generated one
                if 'code' in product_data:
                    del product_data['code']
                    
                # Create product with a specific code
                product = Products(**product_data)
                product.code = new_product_code
                product._skip_auto_code = True  # Add a flag to bypass auto-generation in save()
                product.save()
                
                # Update the created_at timestamp to be unique and in reverse order
                # This ensures the most recently imported products (last ones in Excel) appear first
                import time
                from django.utils import timezone
                import datetime
                
                # Delay creation timestamps by milliseconds in reverse order
                # Last rows in Excel get newer timestamps, appearing first in the list
                delay_seconds = 0.01 * (1000 - cls._row_index)  # Last rows get newer timestamps
                new_timestamp = timezone.now() - datetime.timedelta(seconds=delay_seconds)
                
                # Update the created_at field directly in the database
                Products.objects.filter(product_id=product.product_id).update(created_at=new_timestamp)
                
                # Refresh from DB to get the updated timestamp
                product.refresh_from_db()
                
                # Log the successful creation
                logger.info(f"Created product: {product.name} (ID: {product.product_id}, Code: {product.code})")
            except Exception as e:
                logger.error(f"Failed to create product: {str(e)}")
                raise ValueError(f"Failed to create product record: {str(e)}")
            
            # 3. Create product variations if provided
            has_variation = any(row_data.get(f) for f in cls.VARIATION_HEADERS)
            
            if has_variation:
                try:
                    # Check if we have enough variation data
                    if not row_data.get('size_name') or not row_data.get('color_name'):
                        raise ValueError("Both size_name and color_name are required for variations")
                    
                    # Get or create size
                    size_category = row_data.get('size_category', 'Standard')
                    size_name = row_data.get('size_name')
                    
                    size = Size.objects.filter(size_name__iexact=size_name).first()
                    if not size:
                        logger.info(f"Creating Size with name='{size_name}'")
                        size = Size.objects.create(
                            size_name=size_name,
                            size_category=size_category
                            )
                        
                        # Get or create color
                        color_name = row_data.get('color_name')
                        color = Color.objects.filter(color_name__iexact=color_name).first()
                        if not color:
                            logger.info(f"Creating Color with name='{color_name}'")
                            color = Color.objects.create(color_name=color_name)
                        
                        # Create product variation
                        sku = row_data.get('sku', f"{product.code}-{size_name}-{color_name}")
                        price = decimal.Decimal(str(row_data.get('variation_price', 0)))
                        quantity = int(row_data.get('variation_quantity', 0))
                        
                        variation = ProductVariation.objects.create(
                            product_id=product,
                            size_id=size,
                            color_id=color,
                            sku=sku,
                            price=price,
                            quantity=quantity
                        )
                        logger.info(f"Created product variation: {sku}")
                        
                except Exception as e:
                    logger.error(f"Error creating product variation: {str(e)}")
                    raise ValueError(f"Failed to create product variation: {str(e)}")
                
                # 4. Create product inventory balance if warehouse and location are provided
                warehouse_name = row_data.get('warehouse_name')
                location_name = row_data.get('location_name')
                quantity = row_data.get('location_quantity')
                
                if warehouse_name and location_name and quantity is not None:
                    try:
                        # Get or create the warehouse - using correct field name 'name'
                        warehouse = Warehouses.objects.filter(name__iexact=warehouse_name).first()
                        if not warehouse:
                            logger.info(f"Creating Warehouse with name='{warehouse_name}'")
                            # Since Warehouses requires city_id, state_id fields which are mandatory,
                            # you might need to create with default values or handle appropriately
                            try:
                                # Get default city/state/country if available
                                default_city = City.objects.first()
                                default_state = State.objects.first()
                                default_country = Country.objects.first()
                                
                                if not (default_city and default_state):
                                    raise ValueError("Cannot create warehouse without city and state. Please create them first.")
                                    
                                warehouse = Warehouses.objects.create(
                                    name=warehouse_name,
                                    code=warehouse_name[:10] if warehouse_name else None,
                                    city_id=default_city,
                                    state_id=default_state,
                                    country_id=default_country
                                )
                            except Exception as e:
                                logger.error(f"Failed to create warehouse: {str(e)}")
                                raise ValueError(f"Failed to create warehouse: {str(e)}")
                        
                        # Get or create the warehouse location
                        location = WarehouseLocations.objects.filter(
                            location_name__iexact=location_name,
                            warehouse_id=warehouse
                        ).first()
                        
                        if not location:
                            logger.info(f"Creating WarehouseLocation with name='{location_name}' in warehouse '{warehouse_name}'")
                            location = WarehouseLocations.objects.create(
                                location_name=location_name,
                                warehouse_id=warehouse
                            )
                        
                        # Create the product item balance
                        try:
                            quantity_int = int(float(quantity))
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid quantity value '{quantity}', defaulting to 0")
                            quantity_int = 0
                        
                        ProductItemBalance.objects.create(
                            product_id=product,
                            warehouse_location_id=location,
                            quantity=quantity_int
                        )
                        logger.info(f"Created product item balance: {quantity_int} units in {location_name}")
                        
                    except Exception as e:
                        logger.error(f"Error creating product inventory balance: {str(e)}")
                        raise ValueError(f"Failed to create product inventory balance: {str(e)}")
            
            # Move this return statement outside the if block so all products are returned
            return product
                
            
    @classmethod
    def generate_template(cls, extra_columns=None):
        """
        Generate Excel template with product-specific fields and variation fields
        """
        wb = super().generate_template(extra_columns)
        ws = wb.active

        # Add data validation for GST classification type
        if 'gst_classification' in cls.FIELD_MAP:
            gst_col = list(cls.FIELD_MAP.keys()).index('gst_classification') + 1
            # Add dropdown validation with allowed values from the model
            dv = DataValidation(type="list", formula1='"HSN,SAC,Both"', allow_blank=True)
            dv.add(f"{get_column_letter(gst_col)}2:{get_column_letter(gst_col)}1000")
            ws.add_data_validation(dv)
            
            # Add a comment explaining valid values
            gst_cell = ws.cell(row=1, column=gst_col)
            comment = Comment('Valid values: HSN, SAC, Both', 'System')
            gst_cell.comment = comment
        
        # Add variation headers after the main product fields
        variation_headers = cls.VARIATION_HEADERS
        
        # Get the last column and append variation headers with same color as main fields
        for col_num, header in enumerate(variation_headers, 1 + len(list(cls.FIELD_MAP.keys()))):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue color
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = len(str(header)) + 4
            
        # Add a helper note in row 2 for the quantity_code field
        if 'quantity_code' in cls.FIELD_MAP:
            col = list(cls.FIELD_MAP.keys()).index('quantity_code') + 1
            note_cell = ws.cell(row=2, column=col)
            note_cell.font = Font(italic=True, color="666666")
            note_cell.alignment = Alignment(horizontal="center")
        
        # Add warehouse headers after the variation fields with same color
        warehouse_headers = ["warehouse_name", "location_name", "location_quantity"]
        
        last_col = len(list(cls.FIELD_MAP.keys())) + len(cls.VARIATION_HEADERS)
        
        for col_num, header in enumerate(warehouse_headers, 1 + last_col):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue color
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = len(str(header)) + 4
        
        # Return statement should be outside of the loop
        return wb



class ProductTemplateAPIView(APIView):
    """
    API for downloading the product import template.
    """
    
    def get(self, request, *args, **kwargs):
        return ProductExcelImport.get_template_response(request)

class ProductExcelUploadAPIView(APIView):
    """
    API for importing products from Excel files.
    """
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request, *args, **kwargs):
        try:
            # Upload and validate file
            file_path, status_code = ProductExcelImport.upload_file(request)
            
            # If there was an error with the file
            if status_code != status.HTTP_200_OK:
                return build_response(0, file_path.get("error", "Unknown error"), [], status_code)
                
            # Process the Excel file
            result, status_code = ProductExcelImport.process_excel_file(
                request.FILES.get('file'),
                ProductExcelImport.create_record
            )
            
            # # Check for validation errors
            # if status_code != status.HTTP_200_OK:
            #     return build_response(0, result.get("error", "Import failed"), [], status_code)
            # Check for validation errors
            if status_code != status.HTTP_200_OK:
                error_msg = result.get("error", "Import failed")
                error_details = {}
                
                # Add more detailed error information
                if "missing_columns" in result:
                    error_details["missing_columns"] = result["missing_columns"]
                if "unexpected_columns" in result:
                    error_details["unexpected_columns"] = result["unexpected_columns"]
                if "missing_expected_columns" in result:
                    error_details["missing_expected_columns"] = result["missing_expected_columns"]
                if "missing_data_rows" in result:
                    error_details["missing_data_rows"] = result["missing_data_rows"]
                    
                return build_response(0, error_msg, error_details, status_code)
            
                
            # Check for processing errors
            success_count = result.get("success", 0)
            errors = result.get("errors", [])
            
            if errors:
                # Return the first error as the main message
                first_error = errors[0]["error"] if errors else "Unknown error during import"
                return build_response(0,f"Import failed: {first_error}",{"errors": errors},status.HTTP_400_BAD_REQUEST)
            else:
                # Success response
                return build_response(
                    success_count,
                    result.get("message", f"{success_count} products imported successfully."),
                    [],
                    status.HTTP_200_OK
                )
            
        except Exception as e:
            logger.error(f"Error in product Excel import: {str(e)}")
            return build_response(0, f"Import failed: {str(e)}", [], status.HTTP_400_BAD_REQUEST)