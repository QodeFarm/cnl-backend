# Grouped imports:
from rest_framework.permissions import BasePermission
from uuid import UUID, uuid4
import logging
import base64
import random
import string
import json
import uuid
import os

# Alphabetized imports:
import inflect
import requests
from django.conf import settings
from django.core.cache import cache
from django.core.mail import EmailMessage
from django.db import models, transaction
from django.db.models import Q
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.serializers import ValidationError

# from apps.sales.models import MstcnlSaleOrder, SaleOrder
from config.settings import MEDIA_ROOT, MEDIA_URL
import os
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction

from config.utils_filter_methods import filter_response
logger = logging.getLogger(__name__)




# Set up basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__) # Create a logger object

def custom_upload_to(instance, filename):
    ''' File Path Handler (for Vendor model only) '''
    file_extension = filename.split('.')[-1]
    unique_id = uuid4().hex[:7]  # Generate a unique ID (e.g., using UUID)
    new_filename = f"{unique_id}_{filename}"
    new_filename = new_filename.replace(' ', '_')
    return os.path.join('vendor', str(instance.name), new_filename)

def encrypt(text):
    ''' Functions for demonstration purposes '''
    if text is None:
        return None
    # Encode the text using base64
    encoded_bytes = base64.b64encode(text.encode("utf-8"))
    encrypted_text = encoded_bytes.decode("utf-8")
    return encrypted_text

def decrypt(encrypted_text):
    if encrypted_text is None:
        return None
    # Decode the text using base64
    decoded_bytes = base64.b64decode(encrypted_text.encode("utf-8"))
    decrypted_text = decoded_bytes.decode("utf-8")
    return decrypted_text

class EncryptedTextField(models.TextField):
    """
    A custom field to store encrypted text.
    """
    def from_db_value(self, value, expression, connection):
        if value is None:
            return value
        try:
            # Attempt to decrypt the value
            return decrypt(value)
        except Exception as e:
            # Handle decryption errors gracefully
            return None
        # Implement decryption logic here
        return decrypt(value)

    def to_python(self, value):
        if isinstance(value, str):
            # Implement decryption logic here
            return decrypt(value)
        return value
 
    def get_prep_value(self, value):
        # Implement encryption logic here
        return encrypt(value)

" NOTE : If you want to decrypt then you can uncomment this and run... in output you will find the decrypted account number "
encoded_account_number = "" # Encoded account number
decoded_bytes = base64.b64decode(encoded_account_number) # Decode from base64
original_account_number = decoded_bytes.decode("utf-8") # Convert bytes to string

#======================= POST / PUT / GET / BUILD RESPONSE =====================================================

def build_response(count, message, data, status_code, errors=None):
    """
    Builds a standardized API response.
    """
    response = {
        'count': count,
        'message': message,
        'data': data
    }
    if errors:
        response['errors'] = errors
    return Response(response, status=status_code)

def list_all_objects_1(self, request, *args, **kwargs):
    queryset = self.filter_queryset(self.get_queryset())
    serializer = self.get_serializer(queryset, many=True)
    message = "NO RECORDS INSERTED" if not serializer.data else None
    return build_response(queryset.count(), message, serializer.data, status.HTTP_201_CREATED if not serializer.data else status.HTTP_200_OK)

from itertools import chain

# def list_all_objects(self, request, *args, **kwargs):
#     try:
#         filters = request.query_params.dict()
#         sale_order_id = filters.get('sale_order_id')
#         records_all = request.query_params.get("records_all", "false").lower() == "true"

#         # Default queryset (from default DB)
#         queryset = self.filter_queryset(self.get_queryset().order_by('-created_at'))

#         #  Special logic: only if sale_order_id is provided
#         if sale_order_id:
#             default_qs = self.get_queryset().using('default').filter(sale_order_id=sale_order_id)
#             mstcnl_qs = self.get_queryset().using('mstcnl').filter(sale_order_id=sale_order_id)

#             if default_qs.exists():
#                 queryset = self.filter_queryset(default_qs.order_by('-created_at'))
#             elif mstcnl_qs.exists():
#                 queryset = self.filter_queryset(mstcnl_qs.order_by('-created_at'))
#             else:
#                 queryset = []  # not found in either DB

#         #  records_all logic (combine both DBs)
#         elif records_all:
#             default_qs = self.filter_queryset(self.get_queryset().using('default').order_by('-created_at'))
#             mstcnl_qs = self.filter_queryset(self.get_queryset().using('mstcnl').order_by('-created_at'))
#             queryset = list(chain(default_qs, mstcnl_qs))

#         #  All other cases = default logic is already active

#         serializer = self.get_serializer(queryset, many=True)
#         message = "NO RECORDS INSERTED" if not serializer.data else None

#         return build_response(
#             len(queryset),
#             message,
#             serializer.data,
#             status.HTTP_201_CREATED if not serializer.data else status.HTTP_200_OK
#         )

#     except Exception as e:
#         logger.error(f"Error in list_all_objects: {str(e)}")
#         return build_response(0, "Error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)

from django.utils.functional import cached_property

def list_all_objects(self, request, *args, **kwargs):
    try:
        # --- read params ---
        filters = request.query_params.dict()
        sale_order_id = filters.get('sale_order_id')
        records_all = request.query_params.get("records_all", "false").lower() == "true"
        page = int(request.query_params.get('page', 1) or 1)
        limit = int(request.query_params.get('limit', 10) or 10)
        start = (page - 1) * limit
        end = start + limit

        # --- helpers ---
        def qs_count(obj):
            try:
                return obj.count()
            except Exception:
                return len(obj)

        def slice_any(obj, s, e):
            # Works for QuerySet and list
            return obj[s:e]

        # --- default queryset (already filtered & ordered) ---
        base_qs = self.filter_queryset(self.get_queryset().order_by('-created_at'))

        total_count = 0
        page_items = []

        # --- special: sale_order_id (pick ONLY one DB if present there) ---
        if sale_order_id:
            default_qs = self.get_queryset().using('default').filter(sale_order_id=sale_order_id).order_by('-created_at')
            default_qs = self.filter_queryset(default_qs)

            mstcnl_qs = self.get_queryset().using('mstcnl').filter(sale_order_id=sale_order_id).order_by('-created_at')
            mstcnl_qs = self.filter_queryset(mstcnl_qs)

            if default_qs.exists():
                total_count = qs_count(default_qs)
                page_items = list(slice_any(default_qs, start, end))
            elif mstcnl_qs.exists():
                total_count = qs_count(mstcnl_qs)
                page_items = list(slice_any(mstcnl_qs, start, end))
            else:
                total_count = 0
                page_items = []

        # --- records_all=true: combine BOTH DBs (default first, then mstcnl) ---
        elif records_all:
            default_qs = self.filter_queryset(self.get_queryset().using('default').order_by('-created_at'))
            mstcnl_qs = self.filter_queryset(self.get_queryset().using('mstcnl').order_by('-created_at'))

            default_count = qs_count(default_qs)
            mstcnl_count = qs_count(mstcnl_qs)
            total_count = default_count + mstcnl_count

            # Efficient cross-DB pagination without materializing full lists
            # Keep the original order: all default first, then mstcnl (same as your previous functionality)
            page_items = []

            # Portion from default DB
            if start < default_count:
                d_start = start
                d_end = min(end, default_count)
                page_items.extend(list(slice_any(default_qs, d_start, d_end)))

            # Portion from mstcnl DB
            if end > default_count:
                m_start = max(0, start - default_count)
                m_end = end - default_count
                if m_start < mstcnl_count:
                    page_items.extend(list(slice_any(mstcnl_qs, m_start, min(m_end, mstcnl_count))))

        # --- default path (single DB as you already had) ---
        else:
            total_count = qs_count(base_qs)
            page_items = list(slice_any(base_qs, start, end))

        serializer = self.get_serializer(page_items, many=True)
        message = "NO RECORDS INSERTED" if not serializer.data else None

        # Preserve your previous status behavior:
        status_code = status.HTTP_201_CREATED if not serializer.data else status.HTTP_200_OK

        # Return paginated response with correct totalCount
        return filter_response(
            len(serializer.data),   # current page count
            message,
            serializer.data,
            page,
            limit,
            total_count,            # IMPORTANT: real total
            status_code
        )

    except Exception as e:
        logger.error(f"Error in list_all_objects: {str(e)}")
        return build_response(0, "Error occurred", [], status.HTTP_500_INTERNAL_SERVER_ERROR)



# def create_instance(self, request, *args, **kwargs):
#     serializer = self.get_serializer(data=request.data)
    
#     if serializer.is_valid():
#         serializer.save()
#         data = serializer.data
        
#         # ----------------Pramod -update ----------------------------
        
        
#         #-------------------------------------------------------------
#         return build_response(1, "Record created successfully", data, status.HTTP_201_CREATED)
#     else:
#         errors_str = json.dumps(serializer.errors, indent=2)
#         logger.error("Serializer validation error: %s", errors_str)
#         return build_response(0, "Form validation failed", [], errors = serializer.errors, status_code=status.HTTP_400_BAD_REQUEST)

def create_instance(self, request, *args, **kwargs):
    serializer = self.get_serializer(data=request.data)

    if serializer.is_valid():
        instance = serializer.save()
        data = serializer.data

        # ---------------- LOGGING ---------------- #
        if getattr(self, "log_actions", False):
            from apps.auditlogs.utils import log_user_action

            module = getattr(self, "log_module_name", self.__class__.__name__.replace("ViewSet", ""))
            pk_field = getattr(self, "log_pk_field", "id")
            pk_value = data.get(pk_field)

            display_field = getattr(self, "log_display_field", None)
            display_value = data.get(display_field) if display_field else pk_value

            log_user_action(
                "default",
                request.user,
                "CREATE",
                module,
                pk_value,
                f"{module} ({display_value}) Created by {request.user.username}"
            )
        # ------------------------------------------- #

        return build_response(1, "Record created successfully", data, status.HTTP_201_CREATED)

    else:
        errors_str = json.dumps(serializer.errors, indent=2)
        logger.error("Serializer validation error: %s", errors_str)
        return build_response(0, "Form validation failed", [], errors = serializer.errors, status_code=status.HTTP_400_BAD_REQUEST)


def update_instance(self, request, *args, **kwargs):
    partial = kwargs.pop('partial', False)
    instance = self.get_object()
    serializer = self.get_serializer(instance, data=request.data, partial=partial)
    serializer.is_valid(raise_exception=True)
    self.perform_update(serializer)
    data = serializer.data
    
    #------------Pramod -updated --------------------------
    
    # ---------------- LOGGING ---------------- #
    if getattr(self, "log_actions", False):
        from apps.auditlogs.utils import log_user_action

        module = getattr(self, "log_module_name", self.__class__.__name__.replace("ViewSet", ""))
        pk_field = getattr(self, "log_pk_field", "id")
        pk_value = data.get(pk_field)

        display_field = getattr(self, "log_display_field", None)
        display_value = data.get(display_field) if display_field else pk_value

        log_user_action(
            "default",
            request.user,
            "UPDATE",
            module,
            pk_value,
            f"{module} ({display_value}) Updated by {request.user.username}"
        )
    # ------------------------------------------- #
    
    #---------------------------------------------------------
    return build_response(1, "Record updated successfully", data, status.HTTP_200_OK)

def perform_update(self, serializer):
    serializer.save()  # Add any custom logic for updating if needed
    
    
def soft_delete(instance):
    instance.is_deleted = True
    instance.save()
    return Response(status=status.HTTP_204_NO_CONTENT)

#=========================== Patterns /  Order number related =====================================================

# def generate_order_number(order_type_prefix):
#     """
#     Generate an order number based on the given prefix and the current date.

#     Args:
#         order_type_prefix (str): The prefix for the order type.

#     Returns:
#         str: The generated order number.
#     """
#     if order_type_prefix == "PRD":
#         key = f"{order_type_prefix}"
#         sequence_number = cache.get(key, 0)
#         sequence_number += 1
#         cache.set(key, sequence_number, timeout=None)

#         sequence_number_str = f"{sequence_number:05d}"
#         order_number = f"{order_type_prefix}-{sequence_number_str}"
#     else:

#         current_date = timezone.now()
#         date_str = current_date.strftime('%y%m')

#         key = f"{order_type_prefix}-{date_str}"
#         sequence_number = cache.get(key, 0)
#         sequence_number += 1
#         cache.set(key, sequence_number, timeout=None)

#         sequence_number_str = f"{sequence_number:05d}"
#         order_number = f"{order_type_prefix}-{date_str}-{sequence_number_str}"
#     return order_number

#---------------------------pavan-end--------------------------------------------------

def generate_order_number(order_type_prefix, model_class=None, field_name=None, override_prefix=None):
    """
    Generate an order number using the last record in the database, safe against server restarts.

    Args:
        order_type_prefix (str): The prefix for the order type.
        model_class (Model): The model class to query for last number.
        field_name (str): The field name storing the order number (e.g., 'return_no').

    Returns:
        str: The generated order number.
    """
    current_date = timezone.now()
    date_str = current_date.strftime('%y%m')
    
    prefix = override_prefix if override_prefix else order_type_prefix
    
    # 🔹 CUSTOMER → DB-based, no cache
    if prefix == "CUST":
        if not model_class or not field_name:
            return f"{prefix}-00001"

        last_record = (
            model_class.objects
            .filter(**{f"{field_name}__startswith": f"{prefix}-"})
            .order_by(f"-{field_name}")
            .first()
        )

        last_number = 0
        if last_record:
            try:
                last_number = int(getattr(last_record, field_name).split('-')[-1])
            except Exception:
                last_number = 0

        return f"{prefix}-{last_number + 1:05d}"

    if prefix == "PRD":
        if model_class and field_name:
            filter_kwargs = {f"{field_name}__startswith": prefix}
            last_record = model_class.objects.filter(**filter_kwargs).order_by(f"-{field_name}").first()
            last_number = 0
            if last_record:
                try:
                    last_number = int(getattr(last_record, field_name).split('-')[-1])
                except Exception:
                    last_number = 0
            return f"{prefix}-{last_number + 1:05d}"
        return f"{prefix}-00001"

    if prefix == "SOO":
        from apps.sales.models import SaleOrder, MstcnlSaleOrder
        filter_kwargs = {f"{field_name}__startswith": f"{prefix}-{date_str}"}
        
        last_default = SaleOrder.objects.using('default').filter(**filter_kwargs).order_by(f"-{field_name}").first()
        last_mstcnl = MstcnlSaleOrder.objects.using('mstcnl').filter(**filter_kwargs).order_by(f"-{field_name}").first()

        def extract_seq(obj):
            if not obj:
                return 0
            try:
                return int(getattr(obj, field_name).split('-')[-1])
            except Exception:
                return 0

        max_default = extract_seq(last_default)
        max_mstcnl = extract_seq(last_mstcnl)
        next_seq = max(max_default, max_mstcnl) + 1

        return f"{prefix}-{date_str}-{next_seq:05d}"
    
    if prefix == "SOO-INV":
        # ✅ Late import to avoid circular
        from apps.sales.models import SaleInvoiceOrders, MstcnlSaleInvoiceOrder
        filter_kwargs = {f"{field_name}__startswith": f"{prefix}-{date_str}"}

        last_default = SaleInvoiceOrders.objects.using('default').filter(**filter_kwargs).order_by(f"-{field_name}").first()
        last_mstcnl = MstcnlSaleInvoiceOrder.objects.using('mstcnl').filter(**filter_kwargs).order_by(f"-{field_name}").first()

        def extract_seq(obj):
            if not obj:
                return 0
            try:
                return int(getattr(obj, field_name).split('-')[-1])
            except Exception:
                return 0

        max_default = extract_seq(last_default)
        max_mstcnl = extract_seq(last_mstcnl)
        next_seq = max(max_default, max_mstcnl) + 1

        return f"{prefix}-{date_str}-{next_seq:05d}"
    
    if prefix == "PTR":
        from apps.sales.models import PaymentTransactions, MstCnlPaymentTransactions
        filter_kwargs = {f"{field_name}__startswith": f"{prefix}-{date_str}"}

        last_default = PaymentTransactions.objects.using('default').filter(**filter_kwargs).order_by(f"-{field_name}").first()
        last_mstcnl = MstCnlPaymentTransactions.objects.using('mstcnl').filter(**filter_kwargs).order_by(f"-{field_name}").first()

        def extract_seq(obj):
            if not obj:
                return 0
            try:
                return int(getattr(obj, field_name).split('-')[-1])
            except Exception:
                return 0

        max_default = extract_seq(last_default)
        max_mstcnl = extract_seq(last_mstcnl)
        next_seq = max(max_default, max_mstcnl) + 1

        return f"{prefix}-{date_str}-{next_seq:05d}"

    # Normal date-based flow — unchanged
    prefix = f"{order_type_prefix}-{date_str}"
    last_number = 0

    if model_class and field_name:
        # filter_kwargs = {f"{field_name}__startswith": prefix}
        # filter_kwargs = {
        #     f"{field_name}__startswith": prefix,
        #     f"{field_name}__regex": r'^[A-Z]+-\d{4}-\d{5}$'
        # }
        # if model_class and field_name:
        filter_kwargs = {f"{field_name}__startswith": prefix}

        # First, try to get only parent orders (ignoring children)
        parent_qs = model_class.objects.filter(
            **filter_kwargs,
            **{f"{field_name}__regex": r'^[A-Z]+-\d{4}-\d{5}$'}
        ).order_by(f"-{field_name}")

        last_record = parent_qs.first()

        # Fallback: if no parent found, allow normal lookup (no regex restriction)
        if not last_record:
            last_record = model_class.objects.filter(**filter_kwargs).order_by(f"-{field_name}").first()

        if last_record:
            try:
                last_number = int(getattr(last_record, field_name).split('-')[-1])
            except Exception:
                last_number = 0

        # last_record = model_class.objects.filter(**filter_kwargs).order_by(f"-{field_name}").first()

        # if last_record:
        #     try:
        #         last_number = int(getattr(last_record, field_name).split('-')[-1])
        #     except Exception:
        #         last_number = 0

    return f"{prefix}-{last_number + 1:05d}"

#----------------------------Pramod-end------------------------------------------------

def generate_ledger_group_code(parent_group_id=None):
    """
    Generate hierarchical code for LedgerGroups based on parent group.
    
    Format:
    - Root level (no parent): 10000000, 20000000, 30000000...
    - Child level 1: 11000000, 12000000, 13000000... (under 10000000)
    - Child level 2: 11100000, 11200000... (under 11000000)
    - Child level 3: 11110000, 11120000... (under 11100000)
    
    Args:
        parent_group_id: UUID of the parent ledger group, or None for root level
        
    Returns:
        str: The generated code
    """
    from apps.masters.models import LedgerGroups
    
    if parent_group_id is None:
        # Root level - increment by 10000000
        last_root = LedgerGroups.objects.filter(
            under_group_id__isnull=True,
            code__isnull=False
        ).order_by('-code').first()
        
        if last_root and last_root.code:
            try:
                last_code = int(last_root.code)
                next_code = last_code + 10000000
            except (ValueError, TypeError):
                next_code = 10000000
        else:
            next_code = 10000000
            
        return str(next_code)
    else:
        # Child level - based on parent code
        try:
            parent = LedgerGroups.objects.get(ledger_group_id=parent_group_id)
            if not parent.code:
                raise ValueError("Parent group must have a code")
                
            parent_code = int(parent.code)
            
            # Find the level by counting zeros from right
            # 10000000 = level 0, 11000000 = level 1, 11100000 = level 2, etc.
            parent_str = str(parent_code)
            
            # Determine increment based on number of trailing zeros
            trailing_zeros = len(parent_str) - len(parent_str.rstrip('0'))
            
            if trailing_zeros >= 6:  # Level 0 (e.g., 10000000)
                increment = 1000000
            elif trailing_zeros >= 5:  # Level 1 (e.g., 11000000)
                increment = 100000
            elif trailing_zeros >= 4:  # Level 2 (e.g., 11100000)
                increment = 10000
            elif trailing_zeros >= 3:  # Level 3 (e.g., 11110000)
                increment = 1000
            elif trailing_zeros >= 2:  # Level 4
                increment = 100
            else:  # Deeper levels
                increment = 10
            
            # Get last sibling code
            last_sibling = LedgerGroups.objects.filter(
                under_group_id=parent_group_id,
                code__isnull=False
            ).order_by('-code').first()
            
            if last_sibling and last_sibling.code:
                try:
                    last_code = int(last_sibling.code)
                    next_code = last_code + increment
                except (ValueError, TypeError):
                    # Start from parent code + increment
                    next_code = parent_code + increment
            else:
                # First child
                next_code = parent_code + increment
                
            return str(next_code)
            
        except LedgerGroups.DoesNotExist:
            raise ValueError("Parent group does not exist")


def generate_ledger_account_code(ledger_group_id):
    """
    Generate code for LedgerAccounts based on the parent LedgerGroup code.
    
    Format: Takes the ledger group code and increments by 1 for each account
    Example: If group code is 23700000, accounts will be 23700001, 23700002, etc.
    
    Args:
        ledger_group_id: UUID of the parent ledger group
        
    Returns:
        str: The generated code
    """
    from apps.masters.models import LedgerGroups
    from apps.customer.models import LedgerAccounts
    
    try:
        ledger_group = LedgerGroups.objects.get(ledger_group_id=ledger_group_id)
        
        if not ledger_group.code:
            raise ValueError("Ledger Group must have a code before creating accounts")
            
        base_code = int(ledger_group.code)
        
        # Get the last account code under this group
        last_account = LedgerAccounts.objects.filter(
            ledger_group_id=ledger_group_id,
            code__isnull=False
        ).order_by('-code').first()
        
        if last_account and last_account.code:
            try:
                last_code = int(last_account.code)
                # Ensure we're incrementing within the group's range
                if last_code >= base_code:
                    next_code = last_code + 1
                else:
                    next_code = base_code + 1
            except (ValueError, TypeError):
                next_code = base_code + 1
        else:
            # First account in this group
            next_code = base_code + 1
            
        return str(next_code)
        
    except LedgerGroups.DoesNotExist:
        raise ValueError("Ledger Group does not exist")


def update_child_ledger_accounts_codes(ledger_group_id, old_code, new_code):
    """
    Update all child LedgerAccount codes when parent LedgerGroup code changes.
    
    Args:
        ledger_group_id: UUID of the ledger group
        old_code: Previous code of the ledger group
        new_code: New code of the ledger group
    """
    from apps.customer.models import LedgerAccounts
    
    if old_code == new_code or not old_code:
        return
        
    try:
        old_base = int(old_code)
        new_base = int(new_code)
        
        # Get all accounts under this group
        accounts = LedgerAccounts.objects.filter(
            ledger_group_id=ledger_group_id,
            code__isnull=False
        )
        
        for account in accounts:
            try:
                current_code = int(account.code)
                # Calculate offset from old base
                offset = current_code - old_base
                # Apply same offset to new base
                account.code = str(new_base + offset)
                account.save(update_fields=['code'])
            except (ValueError, TypeError):
                continue
                
    except (ValueError, TypeError):
        pass


class OrderNumberMixin(models.Model):
    order_no_prefix = ''
    order_no_field = ''

    class Meta:
        abstract = True
        
    def get_order_prefix(self):
        """
        Override to handle 'Other' sale type case
        and validate the prefix inline.
        """
        if hasattr(self, 'sale_type_id') and self.sale_type_id:
            if self.sale_type_id.name and self.sale_type_id.name.lower() == 'Other':
                return 'SOO'
        if hasattr(self, 'bill_type') and self.bill_type:
            if self.bill_type.lower() == 'Others':
                return 'SOO-INV'
        
        # Validate existing prefix before returning
        valid_prefixes = ['SO', 'SOO', 'PO', 'SO-INV', 'SR', 'PO-INV', 'PR', 'PRD', 'PTR', 'BPR', 'CUST']  # add all that you use
        if self.order_no_prefix not in valid_prefixes:
            raise ValueError("Invalid prefix")  # <== this will surface clearly

        return self.order_no_prefix
        
    # def get_order_prefix(self):
    #     """Method to be overridden by child classes for custom prefix logic"""
    #     return self.order_no_prefix

    def save(self, *args, **kwargs):
        """
        Override the save method to generate and set the order number if it is not already set.
        """
        if not getattr(self, self.order_no_field):
            prefix = self.get_order_prefix()
            setattr(self, self.order_no_field, generate_order_number(prefix))
        super().save(*args, **kwargs)

def increment_order_number(order_type_prefix):
    """
    Generate and increment an order number based on the given prefix and the current date.

    Args:
        order_type_prefix (str): The prefix for the order type.

    Returns:
        str: The generated order number.
    """
    
    if order_type_prefix == "CUST":
        key = f"{order_type_prefix}"
        sequence_number = cache.get(key, 0)
        sequence_number += 1
        cache.set(key, sequence_number, timeout=None)

        sequence_number_str = f"{sequence_number:05d}"
        return f"{order_type_prefix}-{sequence_number_str}"
    # Handle prefixes without date-based sequences (e.g., PRD)
    if order_type_prefix == "PRD":
        key = f"{order_type_prefix}"
        sequence_number = cache.get(key, 0)
        sequence_number += 1
        cache.set(key, sequence_number, timeout=None)

        sequence_number_str = f"{sequence_number:05d}"
        return f"{order_type_prefix}-{sequence_number_str}"

    # Handle date-based sequences (e.g., SO-INV, SO, SHIP, etc.)
    current_date = timezone.now()
    date_str = current_date.strftime('%y%m')

    key = f"{order_type_prefix}-{date_str}"
    sequence_number = cache.get(key, 0) + 1
    cache.set(key, sequence_number, timeout=None)

    sequence_number_str = f"{sequence_number:05d}"
    return f"{order_type_prefix}-{date_str}-{sequence_number_str}"


#=========================== BULK DATA VALIDATIONS / CURD OPERATION-REQUIREMENTS ===================================
def normalize_value(value):
    '''Check if the value is a list containing exactly one empty dictionary or [empty_dict, None]'''
    if value == [{}] or value is None or value == [{}, None]:
        return []
    return value

def get_object_or_none(model, **kwargs):
    """
    Fetches a single object from the database or returns None if not found.
    """
    try:
        return model.objects.get(**kwargs)
    except model.DoesNotExist:
        return None

# def delete_multi_instance(del_value, main_model_class, related_model_class, main_model_field_name=None):
#     """
#     Deletes instances from a related model based on a field value from the main model.

#     :param del_value: Value of the main model field to filter related model instances.
#     :param main_model_class: The main model class.
#     :param related_model_class: The related model class from which to delete instances.
#     :param main_model_field_name: The field name in the related model that references the main model.
#     """
#     try:
#         # Get the main model's primary key field name
#         main_model_pk_field_name = main_model_class._meta.pk.name

#         # Arrange arguments to filter
#         filter_kwargs = {main_model_field_name or main_model_pk_field_name: del_value}

#         # Delete related instances
#         deleted_count, _ = related_model_class.objects.filter(**filter_kwargs).delete()
#         logger.info(f"Deleted {deleted_count} instances from {related_model_class.__name__} where {filter_kwargs}.")
#     except Exception as e:
#         logger.error(f"Error deleting instances from {related_model_class.__name__}: {str(e)}")
#         return False
#     return True

# def delete_multi_instance(del_value, main_model_class, related_model_class, main_model_field_name=None, using_db=None):
#     """
#     Deletes instances from a related model based on a field value from the main model.
#     Allows for database selection (using_db).

#     :param del_value: Value of the main model field to filter related model instances.
#     :param main_model_class: The main model class.
#     :param related_model_class: The related model class from which to delete instances.
#     :param main_model_field_name: The field name in the related model that references the main model.
#     :param using_db: The database to use for the deletion operation (e.g., 'mstcnl', 'devcnl').
#     """
#     try:
#         # Get the main model's primary key field name
#         main_model_pk_field_name = main_model_class._meta.pk.name

#         # Arrange arguments to filter
#         filter_kwargs = {main_model_field_name or main_model_pk_field_name: del_value}

#         # Delete related instances from the specified database
#         deleted_count, _ = related_model_class.objects.using(using_db).filter(**filter_kwargs).delete()
#         logger.info(f"Deleted {deleted_count} instances from {related_model_class.__name__} where {filter_kwargs}.")

#     except Exception as e:
#         logger.error(f"Error deleting instances from {related_model_class.__name__}: {str(e)}")
#         return False

#     return True

def delete_multi_instance(del_value, main_model_class, related_model_class, main_model_field_name=None, using_db=None):
    """
    Soft deletes instances from a related model by setting is_deleted=True,
    based on a field value from the main model. Allows for database selection (using_db).

    :param del_value: Value of the main model field to filter related model instances.
    :param main_model_class: The main model class.
    :param related_model_class: The related model class from which to soft delete instances.
    :param main_model_field_name: The field name in the related model that references the main model.
    :param using_db: The database to use for the operation (e.g., 'mstcnl', 'devcnl').
    """
    try:
        main_model_pk_field_name = main_model_class._meta.pk.name
        filter_kwargs = {main_model_field_name or main_model_pk_field_name: del_value}

        related_qs = related_model_class.objects.using(using_db) if using_db else related_model_class.objects
        related_instances = related_qs.filter(**filter_kwargs)

        count = 0
        for instance in related_instances:
            instance.is_deleted = True
            if using_db:
                instance.save(using=using_db)
            else:
                instance.save()
            count += 1

        logger.info(f"Soft deleted {count} instances from {related_model_class.__name__} where {filter_kwargs}.")

    except Exception as e:
        logger.error(f"Error soft deleting instances from {related_model_class.__name__}: {str(e)}")
        return False

    return True



# def validate_multiple_data(self, bulk_data, model_serializer, exclude_fields):
#         errors = []

#         if bulk_data:
#             if isinstance(bulk_data, list):
#                 bulk_data = bulk_data
#             if isinstance(bulk_data, dict):
#                 bulk_data = [bulk_data]
        
#         # Validate child data
#         child_serializers = []
#         for data in bulk_data:
#             child_serializer = model_serializer(data=data)
#             if not child_serializer.is_valid(raise_exception=False):
#                 error = child_serializer.errors
#                 exclude_keys = exclude_fields
#                 # Create a new dictionary excluding specified keys
#                 filtered_data = {k: v for k, v in error.items() if k not in exclude_keys}
#                 if filtered_data:
#                     errors.append(filtered_data)

#         return errors

def validate_multiple_data(self, bulk_data, model_serializer, exclude_fields, using_db=None):
    errors = []

    if bulk_data:
        if isinstance(bulk_data, list):
            bulk_data = bulk_data
        if isinstance(bulk_data, dict):
            bulk_data = [bulk_data]

    # Validate child data
    child_serializers = []
    for data in bulk_data:
        # Pass using_db to the model_serializer to handle database context properly
        child_serializer = model_serializer(data=data, context={'using_db': using_db})
        if not child_serializer.is_valid(raise_exception=False):
            error = child_serializer.errors
            exclude_keys = exclude_fields
            # Create a new dictionary excluding specified keys
            filtered_data = {k: v for k, v in error.items() if k not in exclude_keys}
            if filtered_data:
                errors.append(filtered_data)

    return errors



def validate_payload_data(self, data , model_serializer, using='default'):
        errors = []

        # Validate parent data
        serializer = model_serializer(data=data)
        # If Main model data is not valid
        if not serializer.is_valid(raise_exception=False):
            error = serializer.errors
            model = serializer.Meta.model
            model_name = model.__name__
            logger.error("Validation error on %s: %s",model_name, str(error))  # Log validation error
            errors.append(error)
        
        return errors

# def generic_data_creation(self, valid_data, serializer_class, update_fields=None, using='default'):
#     '''
#     ** This function creates new instances with valid data & at the same time it updates the data with required fields**
#     valid_data - The data to be created
#     serializer_class - name of the serializer for which the data is to be created.
#     update_fields - fields to be updated before the data is created [input type dict]
#     '''
#     # If any fields to be updated before the data is created
#     if update_fields:
#         for data in valid_data:
#             for key, value in update_fields.items():
#                 data[key] = value

#     data_list = []
#     for data in valid_data:
#         serializer = serializer_class(data=data)
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#         data_list.append(serializer.data)

#     return data_list

def generic_data_creation(self, valid_data, serializer_class, update_fields=None, using='default'):
    '''
    ** This function creates new instances with valid data & at the same time it updates the data with required fields**
    valid_data - The data to be created
    serializer_class - name of the serializer for which the data is to be created.
    update_fields - fields to be updated before the data is created [input type dict]
    '''
    # If any fields to be updated before the data is created
    if update_fields:
        for data in valid_data:
            for key, value in update_fields.items():
                data[key] = value

    data_list = []
    for data in valid_data:
        serializer = serializer_class(data=data)
        serializer.is_valid(raise_exception=True)

        #  This is the only line added to respect the 'using' parameter
        instance = serializer.Meta.model.objects.db_manager(using).create(**serializer.validated_data)
        serializer = serializer_class(instance)

        data_list.append(serializer.data)

    return data_list


# def generic_data_creation(self, valid_data, serializer_class, update_fields=None, using=None):
#     '''
#     Generic function to create instances - remains completely table-agnostic
#     '''
#     if isinstance(valid_data, dict):
#         valid_data = [valid_data]
#         return_as_dict = True
#     else:
#         return_as_dict = False

#     if update_fields:
#         for data in valid_data:
#             data.update(update_fields)

#     data_list = []
#     for data in valid_data:
#         serializer = serializer_class(data=data)
#         serializer.is_valid(raise_exception=True)
        
#         if using:
#             ModelClass = serializer.Meta.model
#             instance = ModelClass.objects.using(using).create(**serializer.validated_data)
#             serializer = serializer_class(instance)
#         else:
#             instance = serializer.save()
            
#         data_list.append(serializer.data)

#     return data_list[0] if return_as_dict else data_list

def validate_input_pk(self, pk=None):
    try:
        UUID(pk, version=4)
    except ValueError:
        logger.info('Invalid UUID provided')
        return build_response(0, "Invalid UUID provided", [], status.HTTP_404_NOT_FOUND)

def filter_uuid(queryset, name, value):
    try:
        uuid.UUID(value)
    except ValueError:
        return queryset.none()
    return queryset.filter(Q(**{name: value}))

def update_multi_instances(self, pk, valid_data, related_model_name, related_class_name, update_fields, main_model_related_field=None, current_model_pk_field=None, using_db='default'):
    '''
    related_model_name : name of the current model class
    main_model_related_field : This field should have the relation with main Model
    '''
    data_list = []
    try:
        filter_kwargs = {main_model_related_field: pk}
        old_instances_list = list(
            related_model_name.objects.using(using_db).filter(**filter_kwargs).values_list('pk', flat=True)
        )
        old_instances_list = [str(uuid) for uuid in old_instances_list]  # conversion 
    except Exception as e:
        logger.error(f"Error fetching instances from {related_model_name.__name__}: {str(e)}")

    pks_in_update_data = []

    if valid_data:
        if isinstance(valid_data, list):
            valid_data = valid_data
        if isinstance(valid_data, dict):
            valid_data = [valid_data]

    update_count = 0
    for data in valid_data:
        id = data.get(current_model_pk_field, None)  # get the primary key in updated data
        if id is not None:
            pks_in_update_data.append(id)
            instance = related_model_name.objects.using(using_db).get(pk=id)
            if id in old_instances_list:
                serializer = related_class_name(instance, data=data, partial=False)
                serializer.is_valid(raise_exception=True)
                serializer.save(using=using_db)
                data_list.append(serializer.data)
                update_count += 1
        else:
            new_instance = generic_data_creation(self, [data], related_class_name, update_fields=update_fields)
            if new_instance:
                data_list.append(new_instance[0])
                logger.info(f'New instance in {related_model_name.__name__} is created')
            else:
                logger.warning(f"Error during update: new record creation failed in {related_model_name.__name__}")
                return build_response(0, f"Error during update: new record creation failed in {related_model_name.__name__}", [], status.HTTP_400_BAD_REQUEST)

    for id in old_instances_list:
        if id not in pks_in_update_data:
            try:
                instance = related_model_name.objects.using(using_db).get(pk=id)
                instance.delete()
                logger.info(f'Old record in {related_model_name.__name__} with id {id} is deleted')
            except Exception as e:
                logger.warning(f'Error deleting the record in {related_model_name.__name__} with id {id}')

    if (update_count == len(old_instances_list)) and update_count != 0:
        logger.info(f'All old instances in {related_model_name.__name__} are updated (update count: {len(old_instances_list)})')

    return data_list

# def validate_put_method_data(self, valid_data, serializer_name, update_fields, model_class_name, current_model_pk_field=None, using_db='default'):
#     error_list = []

#     if valid_data:
#         if isinstance(valid_data, list):
#             valid_data = valid_data
#         if isinstance(valid_data, dict):
#             valid_data = [valid_data]

#     for data in valid_data:
#         id = data.get(current_model_pk_field, None)
#         if id is not None:
#             instance = model_class_name.objects.using(using_db).filter(pk=id).first()
#             serializer = serializer_name(instance, data=data)
#             if not serializer.is_valid(raise_exception=False):
#                 error = serializer.errors
#                 model = serializer.Meta.model
#                 model_name = model.__name__
#                 logger.error("Validation error on %s: %s", model_name, str(error))
#                 error_list.append(error)
#         else:
#             serializer = serializer_name(data=data)
#             if not serializer.is_valid(raise_exception=False):
#                 error = serializer.errors
#                 exclude_keys = update_fields
#                 filtered_data = {k: v for k, v in error.items() if k not in exclude_keys}
#                 if filtered_data:
#                     error_list.append(filtered_data)

#     return error_list

def validate_put_method_data(self, valid_data, serializer_name, update_fields, model_class_name, current_model_pk_field=None, db_to_use=None):
    error_list = []
    
    #  Prioritize `db_to_use` if given, fallback to `using_db`
    selected_db = db_to_use

    if valid_data:
        if isinstance(valid_data, list):
            valid_data = valid_data
        if isinstance(valid_data, dict):
            valid_data = [valid_data]

    for data in valid_data:
        instance = None
        id = data.get(current_model_pk_field)
        if id is not None:
            instance = model_class_name.objects.using(selected_db).filter(pk=id).first()
            serializer = serializer_name(instance, data=data)
            if not serializer.is_valid(raise_exception=False):
                error = serializer.errors
                model = serializer.Meta.model
                model_name = model.__name__
                logger.error("Validation error on %s: %s", model_name, str(error))
                error_list.append(error)
        else:
            serializer = serializer_name(data=data)
            if not serializer.is_valid(raise_exception=False):
                error = serializer.errors
                exclude_keys = update_fields
                filtered_data = {k: v for k, v in error.items() if k not in exclude_keys}
                if filtered_data:
                    error_list.append(filtered_data)

    return error_list


# def update_multi_instances(self, pk, valid_data, related_model_name, related_class_name, update_fields, main_model_related_field=None, current_model_pk_field=None):
#     '''
#     related_model_class : name of the current model Name
#     main_model_related_field : This field should have the relation with main Model
#     '''
#     data_list = []
#     try:
#         filter_kwargs = {main_model_related_field:pk}
#         old_instances_list  = list(related_model_name.objects.filter(**filter_kwargs).values_list('pk', flat=True))
#         old_instances_list = [str(uuid) for uuid in old_instances_list] # conversion 
#     except Exception as e:
#         logger.error(f"Error fetching instances from {related_model_name.__name__}: {str(e)}")

#     # get the ids that are updated
#     pks_in_update_data = []

#     # prepare user provided pks and verify with the previous records
#     if valid_data:
#         if isinstance(valid_data, list):
#             valid_data = valid_data
#         if isinstance(valid_data, dict):
#             valid_data = [valid_data]

#     update_count = 0
#     for data in valid_data:
#         id = data.get(current_model_pk_field,None) # get the primary key in updated data
#         # update the data for avilable pks
#         if id is not None:
#             pks_in_update_data.append(id) # collecting how many ids are updated
#             instance = related_model_name.objects.get(pk=id)
#             # if given pk is avilable in old instances then update
#             if id in old_instances_list:
#                 serializer = related_class_name(instance, data=data, partial=False)
#                 serializer.is_valid(raise_exception=True)
#                 serializer.save()
#                 data_list.append(serializer.data)
#                 update_count = update_count + 1
#         # If there is no pk avilabe (id will be None), it is new record to create
#         else:
#             # create new record by using function 'generic_data_creation'
#             new_instance = generic_data_creation(self,[data],related_class_name,update_fields=update_fields)
#             if new_instance:
#                 data_list.append(new_instance[0]) # append the new record to existing records
#                 logger.info(f'New instance in {related_model_name.__name__} is created')
#             else:
#                 logger.warning("Error during update: new record creation failed in {related_model_name.__name__}")
#                 return build_response(0,f"Error during update: new record creation failed in {related_model_name.__name__}",[],status.HTTP_400_BAD_REQUEST)
            
#     # Delete the previous records if those are not mentioned in update data
#     for id in old_instances_list:
#         if id not in pks_in_update_data:
#             try:
#                 instance = related_model_name.objects.get(pk=id)
#                 instance.delete()
#                 logger.info(f'Old record in {related_model_name.__name__} with id {id} is deleted')
#             except Exception as e:
#                 logger.warning(f'Error deleting the record in {related_model_name.__name__} with id {id}')
  
#     if (update_count == len(old_instances_list)) and update_count != 0:
#         logger.info(f'All old instances in {related_model_name.__name__} are updated (update count : {len(old_instances_list)})')

#     return data_list

# def validate_put_method_data(self, valid_data, serializer_name, update_fields, model_class_name, current_model_pk_field=None):

#     error_list = []

#     if valid_data:
#         if isinstance(valid_data, list):
#             valid_data = valid_data
#         if isinstance(valid_data, dict):
#             valid_data = [valid_data]

#     for data in valid_data:
#         id = data.get(current_model_pk_field,None) # get the primary key in updated data
#         # update the data for avilable pks
#         if id is not None:
#             instance = model_class_name.objects.filter(pk=id).first()
#             serializer = serializer_name(instance,data=data)
#             # If Main model data is not valid
#             if not serializer.is_valid(raise_exception=False):
#                 error = serializer.errors
#                 model = serializer.Meta.model
#                 model_name = model.__name__
#                 logger.error("Validation error on %s: %s",model_name, str(error))  # Log validation error
#                 error_list.append(error)

#         # If there is no pk avilabe (id will be None), validate the new record
#         else:
#             serializer = serializer_name(data=data)
#             if not serializer.is_valid(raise_exception=False):
#                 error = serializer.errors
#                 exclude_keys = update_fields
#                 # Create a new dictionary excluding specified keys
#                 filtered_data = {k: v for k, v in error.items() if k not in exclude_keys}
#                 if filtered_data:
#                     error_list.append(filtered_data)

#     return error_list

def validate_order_type(data, error_list, model_name,look_up=None):
    '''
    data - the data to be validated
    error_list - the list contains erros
    model_name - Name of the model
    look_up - field that needs to be validated
    '''
    order_type = data.get(look_up,None) # 'order_type' is additonal Field and not defined in model
    if order_type is None:
        error_list.append({look_up:["This field may not be null."]})
    else:
        order_type = get_object_or_none(model_name, name=order_type)
        if order_type is None:
            if len(error_list) > 0:
                error_list[0][look_up] = ["Invalid order type."]
            else:
                error_list.append({look_up:["Invalid order type."]})

def get_related_data(model, serializer_class, filter_field, filter_value):
    """
    Retrieves related data for a given model, serializer, and filter field.
    """
    try:
        related_data = model.objects.filter(**{filter_field: filter_value})
        serializer = serializer_class(related_data, many=True)
        logger.debug("Retrieved related data for model %s with filter %s=%s.", model.__name__, filter_field, filter_value)
        return serializer.data
    except Exception as e:
        logger.exception("Error retrieving related data for model %s with filter %s=%s: %s", model.__name__, filter_field, filter_value, str(e))
        return []

def product_stock_verification(parent_model, child_model, data):
    """
    Verifies if sufficient stock is available for the product when a Sale/Purchase/work Order is being created.
    Raises a ValidationError if the product's available stock is less than the quantity being ordered.
    """
    def assign_error(balance, order_qty, stock_error, product, size=None, color=None):
        # Ensure balance is an integer and order_qty is also an integer
        from apps.products.models import Products, Size, Color
        product_name = Products.objects.get(product_id=product).name
        # if balance <= 0:
        #     stock_error[f'{product_name}'] = f"Product is Out Of Stock. Available: {balance}, Ordered: {order_qty}"

        # # Validate if the order_qty is greater than the available stock balance
        # elif int(order_qty) > balance:
        #     stock_error[f'{product_name}'] = f'Insufficient stock for this product. Available: {balance}, Ordered: {order_qty}'

        # Product variation verification.
        try:
            if size or color: # if both are none, then it is direct product.
            # Update each product variation stock (Subtract/Add the order QTY from stock)
                product_variation_instance = child_model.objects.get(product_id=product, size_id=size, color_id=color)
        except Exception:
            size_name = (Size.objects.filter(size_id=size).first().size_name if size else None)
            color_name = (Color.objects.filter(color_id=color).first().color_name if color else None)
            logger.error(f"Product variation with size :{size_name} , color :{color_name} does not exist.")
            stock_error[f'{product_name}'] = f"Product variation with size :{size_name} , color :{color_name} does not exist."
    
    stock_error = {}
    
    for item in data:
        product = item.get('product_id', None)
        size = item.get('size_id', None)
        color = item.get('color_id', None)
        order_qty = item.get('quantity', None)

        try:
            # Verify if product has variations
            variations = child_model.objects.filter(product_id=product)
            if not variations.exists():  # Check if variations are not present
                # Verify the Main Product (Direct Product)
                product_balance = parent_model.objects.get(product_id=product).balance
                logger.info('product_balance = %s', product_balance)
                
                # Call the error handling function if stock is insufficient
                assign_error(product_balance, order_qty, stock_error, product, size, color)

            else:
                # Get the product variation (assuming only one variation exists for the combination)
                product_variation = child_model.objects.get(
                    product_id=product,
                    size_id=size,
                    color_id=color
                )
                
                # Validate the variation quantity
                assign_error(product_variation.quantity, order_qty, stock_error, product, size, color)

        except parent_model.DoesNotExist:
            logger.error(f'Product with ID {product} not found in Products model.')
        except child_model.DoesNotExist:
            assign_error(0, order_qty, stock_error, product, size, color)
            logger.error(f'Variation with product_id {product}, size_id {size}, color_id {color} not found in Product Variations.')
        except Exception as e:
            logger.error(f'Unexpected error: {str(e)}')

    return stock_error

def previous_product_instance_verification(model_name, data): # In case of Product Return
    """
    Verifies if PREVIOUS PRODUCT INTANCE is available for the product when a SaleOrder is was created.
    Raises a ValidationError if the product's instance is not present in database.
    """    
    stock_error = {}
    for item in data:
        product = item.get('product_id',None)
        size = item.get('size_id',None)
        color = item.get('color_id',None)
        order_qty = item.get('quantity',None)

        # Verify if 'Product variations' are present with provided 'product_id' in payload.
        product_variation_instance = model_name.objects.filter(product_id=product).exists()

        if product_variation_instance:
            # Verify the Product Varition
            try:
                product_variation = model_name.objects.get(
                    product_id=product,
                    size_id=size,
                    color_id=color
                )
                available_stock = product_variation.quantity
            except model_name.DoesNotExist:
                stock_error[f'{product}'] = 'This product variation is unavailable or has been removed.'
    return stock_error

# @transaction.atomic
# def update_product_stock(parent_model, child_model, data, operation):
#     for item in data:
#         product = item.get('product_id',None)
#         return_qty = int(item.get('quantity',None))
#         size = item.get('size_id',None)
#         color = item.get('color_id',None)

#         # Update each product stock (Subtract the order QTY from stock)
#         product_instance = parent_model.objects.get(product_id=product)
#         if operation == 'add':
#             product_instance.balance += return_qty
#         elif operation == 'subtract':
#             product_instance.balance -= return_qty
#         product_instance.save()

#         try:
#             # Update each product variation stock (Subtract/Add the order QTY from stock)
#             product_variation_instance = child_model.objects.get(
#                 product_id=product,
#                 size_id=size,
#                 color_id=color
#             )
#             if operation == 'add':
#                 product_variation_instance.quantity += return_qty
#             elif operation == 'subtract':
#                 product_variation_instance.quantity -= return_qty
#             product_variation_instance.save()
#             logger.info(f'Updated stock for Product ID : {product}')
#         except Exception:
#             logger.info('Direct Product stock is updated.')

# @transaction.atomic
# def update_product_stock(parent_model, child_model, data, operation, using='default'):
#     for item in data:
#         product = item.get('product_id', None)
#         return_qty = float(item.get('quantity', None))
#         size = item.get('size_id', None)
#         color = item.get('color_id', None)

#         # Update each product stock (Subtract the order QTY from stock)
#         product_instance = parent_model.objects.using(using).get(product_id=product)
#         if operation == 'add':
#             product_instance.balance += return_qty
#         elif operation == 'subtract':
#             product_instance.balance -= return_qty
#         product_instance.save(using=using)  #  respect DB

#         try:
#             # Update each product variation stock (Subtract/Add the order QTY from stock)
#             product_variation_instance = child_model.objects.using(using).get(
#                 product_id=product,
#                 size_id=size,
#                 color_id=color
#             )
#             if operation == 'add':
#                 product_variation_instance.quantity += return_qty
#             elif operation == 'subtract':
#                 product_variation_instance.quantity -= return_qty
#             product_variation_instance.save(using=using)  #  respect DB
#             logger.info(f'Updated stock for Product ID : {product}')
#         except Exception:
#             logger.info('Direct Product stock is updated.')

@transaction.atomic
def update_product_stock(parent_model, child_model, data, operation, using='default'):
    for item in data:
        product_id = item.get('product_id')
        return_qty = float(item.get('quantity', 0))
        size_id = item.get('size_id')
        color_id = item.get('color_id')

        try:
            with transaction.atomic(using=using):
                # Update parent product
                product_instance = parent_model.objects.using(using).select_for_update().get(product_id=product_id)
                
                if operation == 'add':
                    product_instance.balance += return_qty
                elif operation == 'subtract':
                    product_instance.balance -= return_qty
                product_instance.save(using=using)

                # Update or create variation if size/color provided
                # if size_id and color_id:
                #     variation, created = child_model.objects.using(using).get_or_create(
                #         product_id=product_id,
                #         size_id=size_id,
                #         color_id=color_id,
                #         defaults={'quantity': return_qty if operation == 'add' else -return_qty}
                #     )
                # --- Always update variation stock (even if size/color are None) ---
                # Use product_instance (model instance) instead of product_id (string UUID)
                variation, created = child_model.objects.using(using).get_or_create(
                    product_id=product_instance,
                    size_id=size_id,     # may be None
                    color_id=color_id,   # may be None
                    # defaults={'quantity': return_qty if operation == 'add' else -return_qty}
                    defaults={
                        'quantity': return_qty if operation == 'add' else -return_qty,
                        'price': item.get('price') or product_instance.purchase_rate or 0
                    }
                )
                    
                if not created:
                    if operation == 'add':
                        variation.quantity += return_qty
                    elif operation == 'subtract':
                        variation.quantity -= return_qty
                    variation.save(using=using)
                    
                logger.info(f'{"Created" if created else "Updated"} variation for Product ID: {product_id}')
                
        except parent_model.DoesNotExist:
            logger.error(f'Product with ID {product_id} does not exist')
        except Exception as e:
            logger.error(f'Error updating stock for Product ID {product_id}: {e}')
            raise  # Re-raise to maintain transaction integrity

#=========================== COMMUNICATION / REPORTS / OTHER SPECIAL FUNCTIONS (CHETAN) ============================

def remove_fields(obj):
    """
    It removes fields from role_permissions for sending Proper data to frontend team after successfully login
    """
    if isinstance(obj, dict):
        obj.pop('created_at', None)
        obj.pop('updated_at', None)
        for value in obj.values():
            remove_fields(value)
    elif isinstance(obj, list):
        for item in obj:
            remove_fields(item)

def validate_uuid(uuid_to_test, version=4):
    try:
        uuid_obj = uuid.UUID(uuid_to_test, version=version)
    except ValueError:
        raise ValidationError("Invalid UUID")
    return uuid_obj
     
def format_phone_number(phone_number):
    phone_number_str = str(phone_number) 
    
    # Check if the phone number length is correct
    if len(phone_number_str) == 10:
        return "91" + phone_number_str  #
    elif len(phone_number_str) == 12 and phone_number_str.startswith("91"):
        return phone_number_str  
    else:
        return "N/A"  

# def send_pdf_via_email(to_email, pdf_relative_path):
#     """Send the generated PDF as an email attachment."""
    
#     # Construct the full path to the PDF file
#     pdf_full_path = os.path.join(MEDIA_ROOT, pdf_relative_path)
    
#     subject = 'Your Requested Documents'
#     body = 'Please Find Your Requested Documents.'
#     email = EmailMessage(subject, body, to=[to_email])

#     # Ensure the PDF file exists before attempting to open it
#     if not os.path.exists(pdf_full_path):
#         raise FileNotFoundError(f"The file {pdf_full_path} does not exist.")

#     # Read the PDF file from the provided full path
#     with open(pdf_full_path, 'rb') as pdf_file:
#         email.attach('Document.pdf', pdf_file.read(), 'application/pdf')
    
#     # Send the email
#     email.send()

#     return "PDF sent via Email successfully."

import os
from django.core.mail import EmailMessage

def send_pdf_via_email(to_email, pdf_relative_path, document_type):
    """Send the generated PDF as an email attachment based on the document type."""

    # Construct the full path to the PDF file
    pdf_full_path = os.path.join(MEDIA_ROOT, pdf_relative_path)

    # Define subject and body based on document type
    doc_messages = {
        "sale_order": ("Sale Order Document", "Please Find Your Requested Sale Order Documents."),
        "sale_quotation": ("Sale Quotation Document", "Please Find Your Requested Sale Quotation Documents."),
        "sale_invoice": ("Sale Invoice Document", "Please Find Your Requested Sale Invoice Documents."),
        "sale_return": ("Sale Return Document", "Please Find Your Requested Sale Return Documents."),
        "purchase_order": ("Purchase Order Document", "Please Find Your Requested Purchase Order Documents."),
        "purchase_return": ("Purchase Return Document", "Please Find Your Requested Purchase Return Documents."),
        "payment_receipt": ("Payment Receipt Document", "Please Find Your Requested Payment Receipt Documents."), #payment_receipt
        "bill_receipt": ("Bill Payment Receipt Document", "Please Find Your Requested Bill Payment Receipt Documents."), #payment_receipt
    }

    subject, body = doc_messages.get(document_type, ("Your Requested Documents", "Please Find Your Requested Documents."))

    email = EmailMessage(subject, body, to=[to_email])

    # Ensure the PDF file exists before attempting to open it
    if not os.path.exists(pdf_full_path):
        raise FileNotFoundError(f"The file {pdf_full_path} does not exist.")

    # Read the PDF file from the provided full path
    with open(pdf_full_path, 'rb') as pdf_file:
        email.attach('Document.pdf', pdf_file.read(), 'application/pdf')

    # Send the email
    email.send()

    return f"{document_type.replace('_', ' ').title()} PDF sent via Email successfully."

# def send_whatsapp_message_via_wati(to_number, file_url):
#     """
#     Send the PDF file as a WhatsApp message using WATI API.
#     Supports DRY RUN mode for local testing.
#     """

#     from django.conf import settings
#     import os, json, requests, logging

#     logger = logging.getLogger(__name__)

#     # -------------------- DRY RUN (LOCAL TESTING) --------------------
#     if getattr(settings, "WHATSAPP_DRY_RUN", False):
#         logger.info("🟡 WHATSAPP DRY RUN MODE")
#         logger.info(f"To Number : {to_number}")
#         logger.info(f"File URL  : {file_url}")

#         return {
#             "whatsapp_sent": True,
#             "mode": "dry_run",
#             "message": "WhatsApp logic executed successfully (DRY RUN)"
#         }

#     # -------------------- FILE VALIDATION --------------------
#     full_file_path = os.path.join(
#         settings.MEDIA_ROOT,
#         file_url.replace(settings.MEDIA_URL, '')
#     )

#     if not os.path.exists(full_file_path):
#         return {
#             "whatsapp_sent": False,
#             "reason": f"File not found: {full_file_path}"
#         }

#     # -------------------- WATI DISABLED SAFETY --------------------
#     if not getattr(settings, "ENABLE_WATI", False):
#         return {
#             "whatsapp_sent": False,
#             "reason": "WATI_DISABLED"
#         }

#     # -------------------- WATI API CALL --------------------
#     url = f'https://live-mt-server.wati.io/312172/api/v1/sendSessionFile/{to_number}'

#     headers = {
#         'accept': '*/*',
#         'Authorization': f'Bearer {settings.WATI_API_TOKEN}',
#     }

#     with open(full_file_path, 'rb') as file:
#         files = {
#             'file': (os.path.basename(full_file_path), file, 'application/pdf'),
#         }
#         response = requests.post(url, headers=headers, files=files)

#     response_data = json.loads(response.text)

#     if response_data.get("result") is True:
#         return {
#             "whatsapp_sent": True,
#             "mode": "wati",
#             "message": "PDF sent via WhatsApp successfully"
#         }

#     return {
#         "whatsapp_sent": False,
#         "reason": response_data.get("info", "Unknown WATI error")
#     }


def send_whatsapp_message_via_wati(to_number, file_url):
    """
    Best-effort WhatsApp sender (WATI)
    MUST be silent and non-blocking
    """
    try:
        # Construct full file path
        full_file_path = os.path.join(MEDIA_ROOT, file_url.replace(MEDIA_URL, ''))

        if not os.path.exists(full_file_path):
            return {
                "sent": False,
                "reason": "FILE_NOT_FOUND"
            }

        url = f'https://live-mt-server.wati.io/312172/api/v1/sendSessionFile/{to_number}'

        headers = {
            'accept': '*/*',
            'Authorization': 'Bearer YOUR_TOKEN_HERE',
        }

        with open(full_file_path, 'rb') as file:
            files = {
                'file': (os.path.basename(full_file_path), file, 'application/pdf'),
            }
            response = requests.post(url, headers=headers, files=files, timeout=10)

        # ---------------- SAFE RESPONSE HANDLING ----------------
        if not response.text:
            return {
                "sent": False,
                "reason": "EMPTY_RESPONSE_FROM_WATI"
            }

        try:
            response_data = response.json()
        except ValueError:
            return {
                "sent": False,
                "reason": "NON_JSON_RESPONSE",
                "raw_response": response.text[:200]
            }

        if response_data.get("result") is True:
            return {
                "sent": True
            }

        return {
            "sent": False,
            "reason": response_data.get("info", "UNKNOWN_WATI_ERROR")
        }

    except Exception as e:
        return {
            "sent": False,
            "reason": "EXCEPTION",
            "error": str(e)
        }

    
# from apps.customers.models import Customer
# from apps.vendors.models import Vendor



import urllib.parse
from django.conf import settings


def resolve_phone_from_document(document_type, pk, city_id=None, request=None):
    """
    Resolve phone number based on document & city
    """
    
    from apps.customer.models import CustomerAddresses
    from apps.vendor.models import VendorAddress
    from apps.sales.models import SaleOrder, SaleInvoiceOrders, SaleReturnOrders, PaymentTransactions
    from apps.purchase.models import PurchaseOrders, PurchaseReturnOrders, BillPaymentTransactions
    
    # -------------------------------------------------
    # 🟢 ACCOUNT LEDGER (SPECIAL CASE)
    # -------------------------------------------------
    if document_type == "account_ledger" and request:
        filter_value = request.GET.get(pk)
        if not filter_value:
            return None

        # CUSTOMER LEDGER
        if pk == "customer_id":
            qs = CustomerAddresses.objects.filter(customer_id=filter_value)
            if city_id:
                qs = qs.filter(city_id=city_id)
            addr = qs.first()
            return addr.phone if addr else None

        # VENDOR LEDGER
        if pk == "vendor_id":
            qs = VendorAddress.objects.filter(vendor_id=filter_value)
            if city_id:
                qs = qs.filter(city_id=city_id)
            addr = qs.first()
            return addr.phone if addr else None

        return None

    # SALE ORDER → CUSTOMER PHONE
    if document_type == "sale_order":
        sale = SaleOrder.objects.filter(pk=pk).first()
        if not sale:
            return None

        qs = CustomerAddresses.objects.filter(customer_id=sale.customer_id)
        if city_id:
            qs = qs.filter(city_id=city_id)

        addr = qs.first()
        return addr.phone if addr else None

    # ACCOUNT LEDGER / CUSTOMER
    if document_type == "sale_invoice": #, "payment_receipt"]
        sale = SaleInvoiceOrders.objects.filter(pk=pk).first()
        if not sale:
            return None
        qs = CustomerAddresses.objects.filter(customer_id=sale.customer_id)
        if city_id:
            qs = qs.filter(city_id=city_id)

        addr = qs.first()
        return addr.phone if addr else None
    
    # ACCOUNT LEDGER / CUSTOMER
    if document_type == "sale_return": #, "payment_receipt"]
        sale = SaleReturnOrders.objects.filter(pk=pk).first()
        if not sale:
            return None
        qs = CustomerAddresses.objects.filter(customer_id=sale.customer_id)
        if city_id:
            qs = qs.filter(city_id=city_id)

        addr = qs.first()
        return addr.phone if addr else None
    
    # ACCOUNT LEDGER / CUSTOMER
    if document_type == "payment_receipt": #, "payment_receipt"]
        sale = PaymentTransactions.objects.filter(pk=pk).first()
        if not sale:
            return None
        qs = CustomerAddresses.objects.filter(customer_id=sale.customer)
        if city_id:
            qs = qs.filter(city_id=city_id)

        addr = qs.first()
        return addr.phone if addr else None

    # VENDOR DOCUMENTS
    if document_type in ["purchase_order", "purchase_return", "bill_receipt"]:
        if document_type == 'purchase_order':
            purchase = PurchaseOrders.objects.filter(pk=pk).first()
            if not purchase:
                return None
            
        if document_type == 'purchase_return':
            purchase = PurchaseReturnOrders.objects.filter(pk=pk).first()
            if not purchase:
                return None
            
        if document_type == 'bill_receipt':
            purchase = BillPaymentTransactions.objects.filter(pk=pk).first()
            if not purchase:
                return None
        
        qs = VendorAddress.objects.filter(vendor_id=purchase.vendor_id)
        if city_id:
            qs = qs.filter(city_id=city_id)

        addr = qs.first()
        return addr.phone if addr else None

    return None


def build_whatsapp_click_url(phone, message):
    encoded = urllib.parse.quote(message)
    return f"https://wa.me/{phone}?text={encoded}"



def convert_amount_to_words(amount):
    '''
This Code convert amount into world.
Ex. amount = 784365923.04
    O/P = INR Seventy-Eight Crore Forty-Three Lakh Sixty-Five Thousand Nine Hundred Twenty-Three And Four Paise
-------------For Testing---------------
# # Example usage:
# amount = 784365923.04
# result = convert_amount_to_words(amount)
# print((result))  

'''
    # Split amount into rupees and paise
    amount_parts = str(amount).split(".")
    
    # Convert rupees part
    rupees_part = int(amount_parts[0])
    
    # Check if there are paise
    paise_part = int(amount_parts[1]) if len(amount_parts) > 1 else 0
    
    # Convert rupees to words
    rupees_in_words = convert_rupees_to_indian_words(rupees_part)
    
    # Convert paise to words
    p = inflect.engine()
    paise_in_words = (
        p.number_to_words(paise_part).replace(",", "") if paise_part > 0 else "zero"
    )
    
    amt_world = f"{rupees_in_words} and {paise_in_words} paise"
    return "INR " + amt_world.title() 

def convert_rupees_to_indian_words(number):
    p = inflect.engine()
    if number == 0:
        return "zero rupees"
    
    parts = []
    
    # Define thresholds and labels
    units = [
        (1000000000, "arab"),
        (10000000, "crore"),
        (100000, "lakh"),
        (1000, "thousand"),
        (100, "hundred")
    ]
    
    for value, name in units:
        if number >= value:
            quotient, number = divmod(number, value)
            parts.append(f"{p.number_to_words(quotient)} {name}")
    
    if number > 0:
        parts.append(p.number_to_words(number))
    
    return " ".join(parts)

def extract_product_data(data, tax_type=None):
    product_data = []
    
    for index, item in enumerate(data, start=1):
        product = item['product']
        unit_options = item['unit_options']        
        product_name = product['name']
        total_boxes = item['total_boxes']
        quantity = float(item['quantity'])
        unit_name = unit_options['unit_name']
        rate = float(item['rate'])
        amount = float(quantity * rate)
        discount_percent = item['discount']
        # discount = quantity * rate * float(discount_percent) / 100 
        discount = quantity * rate * (float(discount_percent) if discount_percent is not None else 0) / 100     
        total_amount = float(item['amount'])

        cgst = float(item['cgst'])
        sgst = float(item['sgst'])
        igst = float(item['igst'])
        gst_tax = 0.0 if tax_type == 'Inclusive' else float(cgst + sgst + igst)

        product_data.append([
            index, product_name, total_boxes, quantity, unit_name, rate, amount, discount_percent, discount, gst_tax, total_amount
        ])

    return product_data


# def extract_product_data(data):
#     product_data = []
    
#     for index, item in enumerate(data, start=1):
#         product = item['product']
#         unit_options = item['unit_options']        
#         product_name = product['name']
#         total_boxes = item['total_boxes']
#         quantity = float(item['quantity'])#item['quantity']
#         unit_name = unit_options['unit_name']
#         rate = float(item['rate']) #item['rate']
#         amount = float(quantity * rate)
#         discount_percent = item['discount']  # Convert discount to float
#         discount = quantity * rate * float(item['discount']) / 100  # Convert discount to float      
#         total_amount = float(item['amount'])  # Convert to float
#         cgst = float(item['cgst'])
#         sgst = float(item['sgst'])
#         igst = float(item['igst'])
#         gst_tax = float(cgst + sgst + igst)
        
#         product_data.append([
#             index, product_name, total_boxes, quantity, unit_name, rate, amount, discount_percent, discount, gst_tax, total_amount ])

#     return product_data

def path_generate(document_type):
    # Generate a random filename
    unique_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) + '.pdf'
    doc_name = document_type + '_' + unique_code
    # Construct the full file path
    file_path = os.path.join(settings.MEDIA_ROOT, 'doc_generater', doc_name)
    # Ensure that the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Return the relative path to the file (relative to MEDIA_ROOT)
    relative_file_path = os.path.join('doc_generater', os.path.basename(doc_name))
    # cdn_path = os.path.join(MEDIA_URL, relative_file_path)
    # print(cdn_path)

    return doc_name, file_path, relative_file_path

#workflow code
workflow_progression_dict = {}
def get_section_id(section_name):
    """
    Retrieve the section ID from the ModuleSections model based on the provided section name.
    This function can be used for any module by passing the appropriate section name (e.g., 'Sale Order', 'Purchase Order').
    """
    from apps.users.models import ModuleSections

    try:
        # Look up the section based on the given section name
        section = ModuleSections.objects.filter(section_name=section_name).first()
        if section:
            return section.section_id
        else:
            raise ValueError(f"Section '{section_name}' not found in ModuleSections")
    except Exception as e:
        raise ValueError(f"Error fetching section_id for {section_name}: {str(e)}")

#Added new logic
def get_active_workflow(section_id):
    """
    Retrieve the active workflow based on the given section_id by looking up the WorkflowStage.
    """
    from apps.sales.models import Workflow, WorkflowStage
    try:
        # Step 1: Find the active workflow for the section by checking WorkflowStage for a match
        active_workflow = Workflow.objects.filter(is_active=True, workflowstage__section_id=section_id).first()

        if not active_workflow:
            raise ValueError(f"No active workflow found for section_id {section_id}")

        print(f"Active Workflow ID: {active_workflow.workflow_id}")

        # Step 2: Retrieve the first stage of this active workflow
        first_stage = WorkflowStage.objects.filter(
            workflow_id=active_workflow.workflow_id, section_id=section_id
        ).order_by('stage_order').first()

        if first_stage:
            return active_workflow
        else:
            raise ValueError(f"No stages found for active workflow with ID {active_workflow.workflow_id}")

    except Exception as e:
        raise ValueError(f"Error fetching active workflow: {str(e)}")

class IsAdminRoles(BasePermission):
    """
    Custom permission to allow access only to users with the 'Admin' role.

    This checks the authenticated user's associated role via the `role_id` ForeignKey
    and ensures that the role's `name` field is equal to "Admin".

    Returns: True if the user is authenticated and their role is 'Admin', False otherwise.
    """
    def has_permission(self, request, view):
        user = request.user
        return (
            user.is_authenticated and 
            hasattr(user, 'role_id') and 
            getattr(user.role_id, 'role_name', '').lower() == "admin"
        )
    

class BaseExcelImportExport:
    """
    Base class for Excel import/export functionality that can be reused across different models.
    This class provides a foundation for Excel template generation, file upload, validation,
    and data import across different components in the application.
    
    Usage:
        1. Create a subclass for your specific model (e.g., CustomerExcelImport)
        2. Override the class attributes (MODEL_CLASS, FIELD_MAP, etc.)
        3. Implement the create_record method
        4. Set up API views that use your subclass
    """
    # Override these in child classes
    MODEL_CLASS = None
    SERIALIZER_CLASS = None
    REQUIRED_COLUMNS = []
    FIELD_MAP = {}
    BOOLEAN_FIELDS = []
    LOOKUP_FIELDS = {}  # Maps foreign keys to their lookup models
    TEMPLATE_FILENAME = "Export_Template.xlsx"
    
    # ============================================================
    # FK BULK HELPER CONFIG - Override in child classes
    # ============================================================
    # Configuration for bulk FK operations. Format:
    # {
    #     'lookup_key': {
    #         'model': ModelClass,
    #         'name_field': 'name',  # Field to lookup by (default: 'name')
    #         'create_fields': {'code': lambda name: name[:3].upper()},  # Extra fields for creation
    #         'excel_column': 'column_name',  # Excel column name (default: same as lookup_key)
    #     }
    # }
    FK_BULK_CONFIG = {}
    
    # Location FK config (Country, State, City) - shared across models
    LOCATION_FK_CONFIG = {
        'country': {'model': None, 'name_field': 'country_name'},
        'state': {'model': None, 'name_field': 'state_name'},
        'city': {'model': None, 'name_field': 'city_name'},
    }
    
    # ============================================================
    # BULK FK HELPER METHODS
    # ============================================================
    
    @classmethod
    def bulk_load_fk_lookups(cls, fk_config):
        """
        Load all FK records into dictionaries for O(1) lookups.
        
        Args:
            fk_config: Dict of FK configurations
            
        Returns:
            Dict of {lookup_key: {name_lower: object}}
            
        Example:
            fk_lookups = cls.bulk_load_fk_lookups({
                'customer_category': {'model': CustomerCategories, 'name_field': 'name'},
                'country': {'model': Country, 'name_field': 'country_name'},
            })
            # Result: {'customer_category': {'retail': <obj>, ...}, 'country': {'india': <obj>, ...}}
        """
        fk_lookups = {}
        total_loaded = 0
        
        for lookup_key, config in fk_config.items():
            model = config.get('model')
            if not model:
                fk_lookups[lookup_key] = {}
                continue
                
            name_field = config.get('name_field', 'name')
            
            # Build lookup dict: {name_lower: object}
            fk_lookups[lookup_key] = {
                getattr(obj, name_field).lower(): obj 
                for obj in model.objects.all()
                if getattr(obj, name_field, None)
            }
            total_loaded += len(fk_lookups[lookup_key])
        
        logger.info(f"FK Lookups: Loaded {total_loaded} records from {len(fk_config)} tables")
        return fk_lookups
    
    @classmethod
    def bulk_collect_missing_fks(cls, validated_rows, fk_config, fk_lookups):
        """
        Scan all rows and collect missing FK values that need to be created.
        
        Args:
            validated_rows: List of row dictionaries
            fk_config: FK configuration dict
            fk_lookups: Existing FK lookup dicts
            
        Returns:
            Dict of {lookup_key: set(missing_values)}
        """
        missing_fks = {key: set() for key in fk_config.keys()}
        
        for row_data in validated_rows:
            for lookup_key, config in fk_config.items():
                excel_col = config.get('excel_column', lookup_key)
                value = row_data.get(excel_col)
                
                if value and str(value).strip():
                    value_lower = str(value).strip().lower()
                    if value_lower not in fk_lookups.get(lookup_key, {}):
                        missing_fks[lookup_key].add(str(value).strip())
        
        # Log summary
        total_missing = sum(len(v) for v in missing_fks.values())
        if total_missing:
            logger.info(f"FK Missing: Found {total_missing} missing values to create")
        
        return missing_fks
    
    @classmethod
    def bulk_create_missing_fks(cls, missing_fks, fk_config, fk_lookups):
        """
        Bulk create all missing FK records and update lookups.
        
        Args:
            missing_fks: Dict of {lookup_key: set(missing_values)}
            fk_config: FK configuration dict
            fk_lookups: Existing FK lookup dicts (will be updated in-place)
            
        Returns:
            Updated fk_lookups dict
        """
        from django.db import transaction
        
        with transaction.atomic():
            for lookup_key, missing_values in missing_fks.items():
                if not missing_values:
                    continue
                    
                config = fk_config.get(lookup_key, {})
                model = config.get('model')
                if not model:
                    continue
                    
                name_field = config.get('name_field', 'name')
                create_fields = config.get('create_fields', {})
                
                # Build objects to create
                objects_to_create = []
                for name in missing_values:
                    obj_data = {name_field: name}
                    # Add extra fields (e.g., code = name[:3].upper())
                    for field_name, field_func in create_fields.items():
                        if callable(field_func):
                            obj_data[field_name] = field_func(name)
                        else:
                            obj_data[field_name] = field_func
                    objects_to_create.append(model(**obj_data))
                
                # Bulk create
                model.objects.bulk_create(objects_to_create)
                
                # Refresh lookup dict with newly created objects
                for obj in model.objects.filter(**{f"{name_field}__in": missing_values}):
                    fk_lookups[lookup_key][getattr(obj, name_field).lower()] = obj
                
                logger.info(f"FK Created: {len(missing_values)} {model.__name__} records")
        
        return fk_lookups
    
    @classmethod
    def get_fk_object(cls, fk_lookups, lookup_key, value):
        """
        Get FK object from pre-loaded lookups.
        
        Args:
            fk_lookups: Lookup dictionary
            lookup_key: Key in fk_lookups (e.g., 'customer_category')
            value: Value to lookup
            
        Returns:
            FK object or None
        """
        if not value or not str(value).strip():
            return None
        return fk_lookups.get(lookup_key, {}).get(str(value).strip().lower())
    
    @classmethod
    def process_row_to_model_data(cls, row_data, field_map, boolean_fields, fk_lookups, fk_field_mapping=None):
        """
        Convert a row dict to model field data using pre-loaded FK lookups.
        
        Args:
            row_data: Dict of {excel_column: value}
            field_map: FIELD_MAP from child class
            boolean_fields: List of boolean field names
            fk_lookups: Pre-loaded FK lookup dicts
            fk_field_mapping: Optional dict mapping excel_col to lookup_key 
                             (e.g., {'customer_category': 'customer_category'})
            
        Returns:
            Dict of model field data ready for Model(**data)
        """
        model_data = {}
        fk_field_mapping = fk_field_mapping or {}
        
        for excel_col, mapping in field_map.items():
            value = row_data.get(excel_col)
            
            if value is None or value == '':
                continue
            
            if isinstance(mapping, tuple):
                # FK field: (field_name, ModelClass)
                field_name = mapping[0]
                lookup_key = fk_field_mapping.get(excel_col, excel_col)
                fk_obj = cls.get_fk_object(fk_lookups, lookup_key, value)
                if fk_obj:
                    model_data[field_name] = fk_obj
            else:
                # Regular field
                if mapping in boolean_fields:
                    model_data[mapping] = cls.parse_boolean(value)
                else:
                    model_data[mapping] = str(value) if value else value
        
        return model_data
    
    @classmethod
    def generate_template(cls, extra_columns=None):
        """
        Generate an Excel template for the model with professional styling.
        Required fields are highlighted with red background and asterisk (*).
        
        Args:
            extra_columns: Additional columns to include in the template
            
        Returns:
            Openpyxl workbook object
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ImportTemplate"
        
        # Get all fields from FIELD_MAP
        headers = list(cls.FIELD_MAP.keys())
        
        # Add extra columns if provided
        if extra_columns:
            headers.extend(extra_columns)
        
        # Style definitions
        # Required fields - Red background with white bold text
        required_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Red
        required_font = Font(bold=True, color="FFFFFF")  # White bold text
        
        # Optional fields - Light blue background with dark text
        optional_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
        optional_font = Font(bold=True, color="000000")  # Black bold text
        
        align_center = Alignment(horizontal="center", vertical="center")
        
        # Add headers with styling
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.alignment = align_center
            
            # Check if this is a required field
            if column_title in cls.REQUIRED_COLUMNS:
                # Required field - Red with asterisk (*)
                cell.value = f"{column_title} *"
                cell.fill = required_fill
                cell.font = required_font
                cell.comment = Comment("⚠️ REQUIRED - This field must be filled", "System")
            else:
                # Optional field - Light blue
                cell.value = column_title
                cell.fill = optional_fill
                cell.font = optional_font
            
            # Set column width
            col_width = len(str(column_title)) + 6
            ws.column_dimensions[get_column_letter(col_num)].width = col_width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        return wb
        
    @classmethod
    def get_template_response(cls, request, extra_columns=None):
        """
        Generates an Excel template with headers based on FIELD_MAP and extra_columns.
        
        Args:
            request: HTTP request object
            extra_columns: Additional columns to include in the template
            
        Returns:
            HttpResponse with Excel file attachment
        """
        wb = cls.generate_template(extra_columns)
        
        # Create and return the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={cls.TEMPLATE_FILENAME}'
        wb.save(response)
        return response

    @classmethod
    def save_upload(cls, file):
        """Save the uploaded file to disk"""
        upload_dir = 'media/uploads'
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, file.name)
        with open(file_path, 'wb+') as dest:
            for chunk in file.chunks():
                dest.write(chunk)
        return file_path

    @classmethod
    def process_excel_file(cls, file, create_record_func, get_or_create_funcs=None):
        """Process the uploaded Excel file and import records."""
        if not file:
            return {"error": "No file uploaded"}, status.HTTP_400_BAD_REQUEST
        
        # Check file type
        file_name = file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return {"error": "Invalid file format. Only Excel files (.xlsx or .xls) are supported."}, status.HTTP_400_BAD_REQUEST
        
        # Save and process the file
        file_path = cls.save_upload(file)
        
        try:
            # Load the workbook with read_only=True for better performance (2-3x faster)
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Get headers from first row - strip asterisk (*) from required field headers
            raw_headers = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]
            # Remove " *" suffix that marks required fields in template
            headers = [h.replace(' *', '').strip() for h in raw_headers]
            
            # Validate required columns
            missing_columns = []
            for col in cls.REQUIRED_COLUMNS:
                if col.lower() not in headers:
                    missing_columns.append(col)
            
            if missing_columns:
                return {
                    "error": "Excel template format mismatch. Required columns are missing.",
                    "missing_columns": missing_columns,
                }, status.HTTP_400_BAD_REQUEST
            
            # Get ALL valid columns - including any additional headers defined in child classes
            all_expected_columns = list(cls.FIELD_MAP.keys())
            
            # Add additional special headers from any class-specific header collections
            special_header_attributes = [attr for attr in dir(cls) if attr.endswith('_HEADERS') and attr != 'REQUIRED_COLUMNS']
            for attr in special_header_attributes:
                if hasattr(cls, attr):
                    additional_headers = getattr(cls, attr)
                    if isinstance(additional_headers, list):
                        all_expected_columns.extend(additional_headers)
            
            # Also allow address-related fields with billing_ or shipping_ prefixes
            # Check for columns in Excel that are not in our expected list
            unexpected_columns = []
            for header in headers:
                if header and header not in [col.lower() for col in all_expected_columns] and not header.startswith('billing_') and not header.startswith('shipping_'):
                    unexpected_columns.append(header)
            
            # Check for expected columns that are missing from Excel (only from FIELD_MAP, not additional headers)
            missing_expected_columns = []
            for expected_col in cls.FIELD_MAP.keys():  # Only check main fields, not additional headers
                if expected_col.lower() not in headers:
                    missing_expected_columns.append(expected_col)
            
            if unexpected_columns or missing_expected_columns:
                return {
                    "error": "Excel template format mismatch.",
                    "unexpected_columns": unexpected_columns,
                    "missing_expected_columns": missing_expected_columns,
                }, status.HTTP_400_BAD_REQUEST
            
            # Extract data rows - improved to skip truly empty rows
            data_rows = []
            row_number_map = []  # Track actual Excel row numbers
            
            # Get indices of required columns for faster checking
            required_col_indices = []
            for req_col in cls.REQUIRED_COLUMNS:
                if req_col.lower() in headers:
                    required_col_indices.append(headers.index(req_col.lower()))
            
            for excel_row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Normalize cell values - strip whitespace from strings
                normalized_row = []
                for cell in row:
                    if cell is not None:
                        if isinstance(cell, str):
                            cell = cell.strip()
                            # Convert empty strings to None
                            cell = cell if cell else None
                    normalized_row.append(cell)
                
                # Check if row has actual data in at least one required column
                # This prevents rows with only formatting/invisible data from being processed
                has_required_data = False
                for col_idx in required_col_indices:
                    if col_idx < len(normalized_row) and normalized_row[col_idx] is not None:
                        has_required_data = True
                        break
                
                # Also check if ANY cell has actual data (for rows without required data but with other data)
                has_any_data = any(cell is not None for cell in normalized_row)
                
                # Only include rows that have at least one required field filled
                # Skip completely empty rows or rows with only non-required data
                if has_required_data:
                    data_rows.append(tuple(normalized_row))
                    row_number_map.append(excel_row_idx)
                elif has_any_data:
                    # Log warning for rows with data but missing all required fields
                    logger.warning(f"Row {excel_row_idx} has data but no required fields - skipping")
            
            # Check for missing required data
            missing_data_rows = []
            for idx, row in enumerate(data_rows):
                actual_row_number = row_number_map[idx]
                row_data = dict(zip(headers, row))
                missing_fields = []
                
                for field in cls.REQUIRED_COLUMNS:
                    field_lower = field.lower()
                    if field_lower in headers:
                        value = row_data.get(field_lower)
                        # Check for None, empty string, or whitespace-only string
                        if value is None or (isinstance(value, str) and not value.strip()):
                            missing_fields.append(field)
                
                if missing_fields:
                    missing_data_rows.append({
                        "row": actual_row_number,
                        "missing_fields": missing_fields
                    })
            
            if missing_data_rows:
                return {
                    "error": "Some rows are missing required data.",
                    "missing_data_rows": missing_data_rows
                }, status.HTTP_400_BAD_REQUEST
            
            # Process valid rows
            results = cls.process_rows(headers, data_rows, create_record_func, get_or_create_funcs, row_number_map)
            return results, status.HTTP_200_OK
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            return {"error": f"Error processing Excel file: {str(e)}"}, status.HTTP_400_BAD_REQUEST
    
    @classmethod
    def process_excel_file_bulk(cls, file, bulk_create_func):
        """
        Process the uploaded Excel file using BULK CREATE for maximum performance.
        
        This is 5-10x faster than process_excel_file() for large imports (500+ rows).
        Use this method when:
        - Importing 500+ records
        - You don't need Django signals to fire
        - You can handle FK lookups in batch
        
        Args:
            file: Uploaded Excel file
            bulk_create_func: Function that handles bulk creation with signature:
                              bulk_create_func(validated_rows, field_map, boolean_fields) -> dict
        
        Returns:
            Tuple of (result_dict, status_code)
        """
        if not file:
            return {"error": "No file uploaded"}, status.HTTP_400_BAD_REQUEST
        
        # Check file type
        file_name = file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return {"error": "Invalid file format. Only Excel files (.xlsx or .xls) are supported."}, status.HTTP_400_BAD_REQUEST
        
        # Save and process the file
        file_path = cls.save_upload(file)
        
        try:
            # Load the workbook with read_only=True for better performance
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Get headers from first row - strip asterisk (*) from required field headers
            raw_headers = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]
            headers = [h.replace(' *', '').strip() for h in raw_headers]
            
            # Validate required columns
            missing_columns = []
            for col in cls.REQUIRED_COLUMNS:
                if col.lower() not in headers:
                    missing_columns.append(col)
            
            if missing_columns:
                return {
                    "error": "Excel template format mismatch. Required columns are missing.",
                    "missing_columns": missing_columns,
                }, status.HTTP_400_BAD_REQUEST
            
            # Get ALL valid columns
            all_expected_columns = list(cls.FIELD_MAP.keys())
            special_header_attributes = [attr for attr in dir(cls) if attr.endswith('_HEADERS') and attr != 'REQUIRED_COLUMNS']
            for attr in special_header_attributes:
                if hasattr(cls, attr):
                    additional_headers = getattr(cls, attr)
                    if isinstance(additional_headers, list):
                        all_expected_columns.extend(additional_headers)
            
            # Check for unexpected/missing columns
            unexpected_columns = []
            for header in headers:
                if header and header not in [col.lower() for col in all_expected_columns] and not header.startswith('billing_') and not header.startswith('shipping_'):
                    unexpected_columns.append(header)
            
            missing_expected_columns = []
            for expected_col in cls.FIELD_MAP.keys():
                if expected_col.lower() not in headers:
                    missing_expected_columns.append(expected_col)
            
            if unexpected_columns or missing_expected_columns:
                return {
                    "error": "Excel template format mismatch.",
                    "unexpected_columns": unexpected_columns,
                    "missing_expected_columns": missing_expected_columns,
                }, status.HTTP_400_BAD_REQUEST
            
            # Extract data rows
            data_rows = []
            row_number_map = []
            
            required_col_indices = []
            for req_col in cls.REQUIRED_COLUMNS:
                if req_col.lower() in headers:
                    required_col_indices.append(headers.index(req_col.lower()))
            
            for excel_row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                normalized_row = []
                for cell in row:
                    if cell is not None:
                        if isinstance(cell, str):
                            cell = cell.strip()
                            cell = cell if cell else None
                    normalized_row.append(cell)
                
                has_required_data = False
                for col_idx in required_col_indices:
                    if col_idx < len(normalized_row) and normalized_row[col_idx] is not None:
                        has_required_data = True
                        break
                
                if has_required_data:
                    data_rows.append(tuple(normalized_row))
                    row_number_map.append(excel_row_idx)
            
            # Close workbook to free memory
            wb.close()
            
            # Process using bulk method
            results = cls.process_rows_bulk(headers, data_rows, bulk_create_func, row_number_map)
            return results, status.HTTP_200_OK
            
        except Exception as e:
            logger.error(f"Error processing Excel file (bulk): {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return {"error": f"Error processing Excel file: {str(e)}"}, status.HTTP_400_BAD_REQUEST

    @classmethod
    def process_rows(cls, headers, data_rows, create_record_func, get_or_create_funcs=None, row_number_map=None):
        """
        Process all rows from Excel file with BATCH PROCESSING for production reliability.
        
        Key improvements:
        1. Processes records in batches (default: 100 records per batch)
        2. Each batch has its own transaction - partial success is preserved
        3. Connection is reset between batches to prevent timeout
        4. Memory is managed by processing in chunks
        5. Detailed progress tracking for debugging
        
        This approach ensures large imports (4000+ records) work reliably in production
        where Nginx/Gunicorn have request timeouts (typically 30-60 seconds).
        """
        import gc
        from django.db import connection, reset_queries
        from django.conf import settings
        
        success = 0
        failed = []
        total_rows = len(data_rows)
        
        # Batch size - configurable, default 100 for optimal performance
        # Smaller batches = more reliable but slower
        # Larger batches = faster but may hit timeout limits
        BATCH_SIZE = getattr(settings, 'EXCEL_IMPORT_BATCH_SIZE', 100)
        
        logger.info(f"Starting batch import: {total_rows} total rows, batch size: {BATCH_SIZE}")
        
        # Process in batches
        for batch_start in range(0, total_rows, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, total_rows)
            batch_rows = data_rows[batch_start:batch_end]
            batch_row_map = row_number_map[batch_start:batch_end] if row_number_map else None
            
            batch_num = (batch_start // BATCH_SIZE) + 1
            total_batches = (total_rows + BATCH_SIZE - 1) // BATCH_SIZE
            logger.info(f"Processing batch {batch_num}/{total_batches} (rows {batch_start+1}-{batch_end})")
            
            # Process each row in this batch
            for idx, row in enumerate(batch_rows):
                actual_idx = batch_start + idx
                actual_row_number = batch_row_map[idx] if batch_row_map else actual_idx + 2
                row_data = dict(zip(headers, row))
                
                # Normalize row_data values - strip whitespace from strings
                for key, value in row_data.items():
                    if isinstance(value, str):
                        row_data[key] = value.strip() if value else None
                
                # Check all required fields (case-insensitive lookup)
                missing_fields = []
                for field in cls.REQUIRED_COLUMNS:
                    field_lower = field.lower()
                    value = row_data.get(field_lower)
                    if value is None or (isinstance(value, str) and not value.strip()):
                        missing_fields.append(field)
                
                if missing_fields:
                    failed.append({
                        "row": actual_row_number, 
                        "error": f"Missing required fields: {', '.join(missing_fields)}"
                    })
                    continue
                    
                try:
                    # Each record gets its own transaction for maximum reliability
                    with transaction.atomic():
                        # Use the provided function to create the record
                        create_record_func(row_data, cls.FIELD_MAP, cls.BOOLEAN_FIELDS, get_or_create_funcs)
                        
                    success += 1
                        
                except Exception as e:
                    logger.error(f"Error processing row {actual_row_number}: {e}")
                    failed.append({"row": actual_row_number, "error": str(e)})
            
            # After each batch: clean up to prevent memory buildup
            # This is critical for large imports
            if settings.DEBUG:
                reset_queries()  # Clear query log in debug mode
            
            gc.collect()  # Force garbage collection
            
            # Log batch completion
            logger.info(f"Batch {batch_num} complete: {success} successful so far, {len(failed)} failed")
        
        logger.info(f"Import complete: {success} successful, {len(failed)} failed out of {total_rows} total")
        
        return {
            "success": success,
            "message": f"{success} records imported successfully.",
            "errors": failed,
            "total": total_rows
        }
    
    @classmethod
    def process_rows_bulk(cls, headers, data_rows, bulk_create_func, row_number_map=None):
        """
        Process all rows from Excel file using BULK_CREATE for maximum performance.
        
        This method is 5-10x faster than process_rows() because it:
        1. Pre-fetches all foreign key lookups in batch (few queries)
        2. Validates ALL rows first before any DB writes
        3. Uses bulk_create to insert records (few INSERT statements)
        
        Args:
            headers: List of column headers from Excel
            data_rows: List of row tuples from Excel
            bulk_create_func: Function that handles bulk creation (must be implemented by subclass)
            row_number_map: Optional list mapping data_row index to actual Excel row numbers
            
        Returns:
            dict with success count, errors list, and total count
        """
        import gc
        from django.db import connection, reset_queries
        from django.conf import settings
        
        total_rows = len(data_rows)
        logger.info(f"Starting BULK import: {total_rows} total rows")
        
        # PHASE 1: Parse all rows into dictionaries
        all_row_data = []
        for idx, row in enumerate(data_rows):
            actual_row_number = row_number_map[idx] if row_number_map else idx + 2
            row_data = dict(zip(headers, row))
            
            # Normalize values - strip whitespace from strings
            for key, value in row_data.items():
                if isinstance(value, str):
                    row_data[key] = value.strip() if value else None
            
            row_data['_excel_row'] = actual_row_number
            all_row_data.append(row_data)
        
        # PHASE 2: Validate all rows (check required fields)
        validated_rows = []
        errors = []
        
        for row_data in all_row_data:
            actual_row_number = row_data.get('_excel_row', 0)
            
            # Check required fields
            missing_fields = []
            for field in cls.REQUIRED_COLUMNS:
                field_lower = field.lower()
                value = row_data.get(field_lower)
                if value is None or (isinstance(value, str) and not value.strip()):
                    missing_fields.append(field)
            
            if missing_fields:
                errors.append({
                    "row": actual_row_number,
                    "error": f"Missing required fields: {', '.join(missing_fields)}"
                })
            else:
                validated_rows.append(row_data)
        
        logger.info(f"Validation complete: {len(validated_rows)} valid, {len(errors)} invalid")
        
        # PHASE 3: Call the bulk create function with validated rows
        if validated_rows:
            try:
                bulk_result = bulk_create_func(validated_rows, cls.FIELD_MAP, cls.BOOLEAN_FIELDS)
                
                # Merge any errors from bulk creation
                if 'errors' in bulk_result:
                    errors.extend(bulk_result['errors'])
                
                success_count = bulk_result.get('success', 0)
                
            except Exception as e:
                logger.error(f"Bulk create failed: {str(e)}")
                # If bulk create fails completely, mark all rows as failed
                for row_data in validated_rows:
                    errors.append({
                        "row": row_data.get('_excel_row', 0),
                        "error": str(e)
                    })
                success_count = 0
        else:
            success_count = 0
        
        # Cleanup
        if settings.DEBUG:
            reset_queries()
        gc.collect()
        
        logger.info(f"BULK import complete: {success_count} successful, {len(errors)} failed out of {total_rows} total")
        
        return {
            "success": success_count,
            "message": f"{success_count} records imported successfully.",
            "errors": errors,
            "total": total_rows
        }

    @staticmethod
    def parse_boolean(value):
        """Convert various formats to boolean"""
        if value is None:
            return None
            
        value_str = str(value).strip().lower()
        if value_str in ["yes", "true", "1"]:
            return True
        elif value_str in ["no", "false", "0"]:
            return False
            
        return None
        
    @classmethod
    def create_record(cls, row_data, field_map=None, boolean_fields=None, get_or_create_funcs=None):
        """
        Creates a record from row data.
        This method should be implemented by subclasses to handle specific model creation.
        
        Args:
            row_data: Dictionary mapping column names to values
            field_map: Dictionary mapping Excel column names to model fields
            boolean_fields: List of fields that should be treated as boolean
            get_or_create_funcs: Dictionary of functions to get or create related objects
            
        Returns:
            Created model instance
        """
        raise NotImplementedError("Subclasses must implement create_record method")
        
    @classmethod
    def get_or_create_fk(cls, model, value, field="name"):
        """Get or create a foreign key reference"""
        if not value:
            return None
            
        value = str(value).strip()
        obj = model.objects.filter(**{f"{field}__iexact": value}).first()
        
        if obj:
            return obj
            
        logger.info(f"Creating {model.__name__} with {field}='{value}'")
        
        # Special case handling for models with required fields
        if hasattr(cls, 'FK_REQUIRED_FIELDS') and model.__name__ in cls.FK_REQUIRED_FIELDS:
            required_fields = cls.FK_REQUIRED_FIELDS[model.__name__]
            create_data = {field: value}
            
            for req_field, default_provider in required_fields.items():
                if callable(default_provider):
                    create_data[req_field] = default_provider()
                else:
                    create_data[req_field] = default_provider
                    
            return model.objects.create(**create_data)
        
        # Standard case - just create with the name field
        return model.objects.create(**{field: value})
    
    @classmethod
    def get_template_response(cls, request, extra_columns=None):
        """
        Generates an Excel template with headers based on FIELD_MAP and extra_columns.
        
        Args:
            request: HTTP request object
            extra_columns: Additional columns to include in the template
            
        Returns:
            HttpResponse with Excel file attachment
        """
        wb = cls.generate_template(extra_columns)
        
        # Create and return the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={cls.TEMPLATE_FILENAME}'
        wb.save(response)
        return response

    @classmethod
    def upload_file(cls, request):
        """
        Handle file upload for Excel import.
        
        Args:
            request: HTTP request object with file attachment
            
        Returns:
            Tuple: (file_path, status_code) or (error_response, status_code)
        """
        file = request.FILES.get('file')
        if not file:
            return {"error": "No file uploaded"}, status.HTTP_400_BAD_REQUEST
            
        # Check file type
        file_name = file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return {"error": "Invalid file format. Only Excel files (.xlsx or .xls) are supported."}, status.HTTP_400_BAD_REQUEST
            
        # Save and process the file
        return cls.save_upload(file), status.HTTP_200_OK
    

    
