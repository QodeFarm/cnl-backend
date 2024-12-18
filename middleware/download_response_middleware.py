# Description :

# This script facilitates downloading data in various formats: JSON, CSV, and Excel. 
# It iterates through all URLs to verify if they end with 'download/{file_format}/'. 
# It proceeds to download model data, filter data, and customized filter data accordingly.
#---------------------------------------------------------------------------------------------

import logging
import csv
import json
import decimal
import openpyxl
from datetime import date
from urllib.parse import urlparse
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from .fields import *

# Set up basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Get an instance of a logger
logger = logging.getLogger(__name__)

# Below function is useful in such cases if some response data is still in 'Decimal Format' which is not co-operative to dump in json.
# It converts Decimal data type to float data type
def convert_decimal_to_float(obj):
    if isinstance(obj, date):
            return obj.isoformat()
    if isinstance(obj, dict):
        return {key: convert_decimal_to_float(value) for key, value in obj.items()}
    elif isinstance(obj, decimal.Decimal):
        return float(obj)
    else:
        return obj

# Below function is useful when the response data in dict type, but in the dictionary required data avilable at key 'data'.
# This required data is type of <class dict_values>  which is not possible to process, so this function is useful to process the all objects
# and if any object data is in Decimal format then by using function it converts to float data type  [function name: convert_decimal_to_float]
def process_the_dict_values_data(response):
    content = response.content
    data = content.decode('utf-8')
    data = response.data
    data = data['data']
    new_data = []
    for v in data:
        sub_data = {}
        for k,m in v.items():
            sub_data[k] = m
        else:
            float_values = convert_decimal_to_float(sub_data)
            new_data.append(float_values)
    else:
        return new_data

# From the url below function will fetch the model name, 
# Based on model_name takes list of fields from fields.py this required columns will appear on excel and csv files.
def fetch_model_fields(original_path):
    if original_path:
        #get the model name from URL
        split_path = original_path.split('/')
        if 'download' in split_path:
            download_index = split_path.index("download")
            model_name_index = download_index - 1
            model_name = split_path[model_name_index]
        else:
            model_name = split_path[-2]
        #get Required fields from fields.py
        fields = all_model_fields[model_name]

        return model_name, fields
    
#download Excel data with required fields metioned in fields.py
def download_fields_excel_data(response,original_path):
    try:
        content = response.content
        data = content.decode('utf-8') #decode the data in content
        data = response.data
        extracted_data = (data['data'])

    except json.JSONDecodeError:
        return response  # Return the original response if JSON decoding fails
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="csv_data_file.csv"'

        # get fields and model
        model_name,fields = fetch_model_fields(original_path)

        writer = csv.writer(response)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{model_name}"

        if data:
            # Write the header
            for col_num, field in enumerate(fields, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = field
                # Excel decoration
                cell = ws[f'{col_letter}1']
                cell.font = bold_white_font
                cell.fill = blue_fill

            # Write the data rows
            for row_num, item in enumerate(extracted_data, 2):
                for col_num, field in enumerate(fields, 1):
                    col_letter = get_column_letter(col_num)
                    val = item.get(field, '')
                    ws[f'{col_letter}{row_num}'] = next((val[key] for key in val if 'name' in key), None) if isinstance(val, dict) else str(val)
                    cell = ws[f'{col_letter}{row_num}']
                    # Excel decoration
                    cell.fill = blue_light_fill

        #Give title to the worksheet
        ws.insert_rows(1)
        second_row_length = len(list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=second_row_length)
        ws.cell(row=1, column=1, value=f'{model_name}'.upper().replace('_',' '))
        cell = ws.cell(row=1, column=1)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        first_cell = ws['A1']
        first_cell.font = Font(bold=True, size=14)
        first_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Yellow color
        ws.insert_rows(2) # add second row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=second_row_length) # merge the second row
        ws.cell(
            row=2, 
            column=1, 
            value= "ABC Industry Pvt. Ltd.\n456 Tech Hub Road\nBlock C, Building 9\nWhitefield, Bangalore, Karnataka 560066\nIndia"
            )
        # Apply text wrapping to the cell
        cell = ws.cell(row=2, column=1)
        # Adjust the height as needed
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 79

        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
        wb.save(response)
    return response
    
# For Excel decoration
# White font color - Bold
bold_white_font = Font(bold=True,size=12, color="FFFFFF")    
# Define the fill color
blue_fill = PatternFill(start_color="5D86B4", end_color="5D86B4", fill_type="solid") # Blue background
blue_light_fill = PatternFill(start_color="CEDAE8", end_color="CEDAE8", fill_type="solid") # Light Blue background

class StripDownloadJsonMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        if request.method == 'GET' and ('download' in request.path_info or 'download' in request.get_full_path()):
            logger.info("\n|----Hello i'm Download Middleware------|")
            # Store the original path
            original_path = request.path_info
            # Store the fill path including filtering options
            filter_path = request.get_full_path()
            parsed_url = urlparse(request.get_full_path())
            # collects applied filter name just before the '=' symbol
            filter_name = parsed_url.query.split("=")[0]  # Ex:  ?phone_number=9848012345  , takes phone_number as filter name


            #------------------- MODEL DATA - EXCEL FORMAT -----------------------------
            # Description: Below code downloads data in EXCEL format
            # The script verifies each URL to determine if it ends with '/download/excel/'
            #----------------------------------------------------------------------------        

            if original_path.endswith('download/excel/'):
                # Strip the '/download/excel/' part from the URL
                modified_path = original_path[:-len('download/excel/')]
                # Modify both path and path_info
                request.path_info = modified_path
                request.path = modified_path
                request.download_excel = True
            else:
                request.download_excel = False

            # Get the response from the next middleware or view
            response = self.get_response(request)
            
            # Restore the original path_info to avoid affecting other middlewares or views
            request.path_info = original_path
            request.path = original_path

            # If the original request URL ended with '/download/excel/'
            if getattr(request, 'download_excel', False):
                # Ensure the response is JSON
                if response.status_code == 200:
                    response = download_fields_excel_data(response,original_path)
                    return response
                
            #------------------- FILTER DATA - EXCEL FORMAT ------------------------------
            # Description: Below code downloads data in EXCEL format
            # The script verifies each URL to determine if it ends with 'download/excel/'
            #-----------------------------------------------------------------------------

            if filter_path.endswith('download/excel/'):
                request.download_excel = True
            else:
                request.download_excel = False

            # Get the response from the next middleware or view
            response = self.get_response(request)

            # If the original request URL ended with '/download/excel/'
            if getattr(request, 'download_excel', False):
                # Ensure the response status_code is 200
                if response.status_code == 200:
                    # collect the list type processed data after processiong from dict_type
                    new_data = process_the_dict_values_data(response)

                    writer = csv.writer(response)
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = f'{filter_name}'

                    if filter_name not in custom_filters:
                        logger.info("Normal filter data is fetched.")
                        response = download_fields_excel_data(response,original_path)
                        return response
    
                    if new_data and filter_name in custom_filters: #if new_data:
                        logger.info("Custom Filter data is fetched.")
                        # Write the headers
                        header = new_data[0].keys()
                        writer.writerow(header)
                        for col_num, header in enumerate(header, 1):
                            col_letter = get_column_letter(col_num)
                            ws[f'{col_letter}1'] = header
                            cell = ws[f'{col_letter}1']
                            cell.fill = blue_fill
                            cell.font = bold_white_font

                        # Write the data rows
                        for row_num, item in enumerate(new_data, 2):
                            for col_num, (key, value) in enumerate(item.items(), 1):
                                col_letter = get_column_letter(col_num)
                                ws[f'{col_letter}{row_num}'] = value if isinstance(value, (int, float)) else str(value)  # Convert value to string, if number dont convert
                                cell = ws[f'{col_letter}{row_num}']
                                cell.fill = blue_light_fill

                        #Give title to the worksheet
                        ws.insert_rows(1)
                        second_row_length = len(list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0])
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=second_row_length)
                        ws.cell(row=1, column=1, value=f'{filter_name}'.upper().replace('_',' '))
                        first_cell = ws['A1']
                        first_cell.font = Font(bold=True, size=14)
                        first_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color
                        first_cell.alignment = Alignment(horizontal='center', vertical='center')                         

                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = f'attachment; filename="{filter_name}.xlsx"'
                        wb.save(response)

                        return response
            
        response = self.get_response(request)
        return response # This will return original response as it is if no '/download/{format_name}/' is detected
